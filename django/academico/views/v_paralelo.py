from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import ProtectedError, Count

from academico.models import Paralelo, PeriodoDeNivelacion, MatriculaParalelo, Carrera, Campus, Horario
from academico.services import servicio_generar_paralelos, servicio_mover_estudiante, servicio_recalcular_cohorte_de_carrera, servicio_retirar_estudiante_de_paralelo, servicio_agregar_estudiante_a_paralelo, periodo_permite_gestion_matriculas
from usuarios.models import PerfilEstudiante
from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
from poo.clases.enums.estado_de_matricula import EstadoDeMatricula
from poo.clases.enums.jornada import Jornada
from usuarios.utils import (
    requiere_perfil,
    usuario_es_solo_lectura,
    obtener_rol_usuario,
    ROL_COORDINADOR_DAN,
    ROL_DIRECTOR_DAN,
    ROL_RECTOR,
    ROL_VICERRECTOR,
    ROL_COORDINADOR_UA,
    ROL_DOCENTE,
)

ROLES_VISUALIZAN = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR, ROL_COORDINADOR_UA, ROL_DOCENTE)
ROLES_MODIFICAN = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_COORDINADOR_UA)


def _obtener_universidad_usuario(user):
    """Get universidad from perfil_administrativo or perfil_docente."""
    perfil_admin = getattr(user, 'perfil_administrativo', None)
    if perfil_admin:
        return perfil_admin.universidad
    perfil_docente = getattr(user, 'perfil_docente', None)
    if perfil_docente:
        return perfil_docente.universidad
    return None


def _numero_para_orden(nombre):
    # Soporta "Paralelo N" (legacy) y letras "A", "B", ..., "A1" etc.
    nombre = str(nombre).strip()
    try:
        return int(nombre.split()[-1])
    except (ValueError, IndexError):
        pass
    # Orden por letra: A=0, B=1, ..., Z=25, A1=26, B1=27...
    if len(nombre) == 1 and nombre.isalpha():
        return ord(nombre.upper()) - ord('A')
    if len(nombre) >= 2 and nombre[0].isalpha() and nombre[1:].isdigit():
        return 26 + (int(nombre[1:]) - 1) * 26 + (ord(nombre[0].upper()) - ord('A'))
    return 0


def _paralelos_del_grupo(representativo):
    carrera = representativo.unidad_curricular.malla_curricular.carrera
    return Paralelo.objects.filter(
        periodo_de_nivelacion=representativo.periodo_de_nivelacion,
        jornada=representativo.jornada,
        nombre=representativo.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    )

@requiere_perfil(*ROLES_VISUALIZAN)
def listar_paralelos(request):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    periodos = PeriodoDeNivelacion.objects.filter(universidad=universidad_usuario)
    carreras = Carrera.objects.filter(campus__universidad=universidad_usuario).order_by("nombre")
    campus_disponibles = Campus.objects.filter(universidad=universidad_usuario).order_by("nombre")

    paralelos = Paralelo.objects.filter(
        periodo_de_nivelacion__universidad=universidad_usuario
    ).select_related(
        "periodo_de_nivelacion",
        "unidad_curricular__malla_curricular__carrera__campus",
    ).annotate(_n_estudiantes=Count("estudiantes_matriculados"))

    periodo_id = request.GET.get("periodo")
    carrera_id = request.GET.get("carrera")
    jornada_filtro = request.GET.get("jornada")
    campus_filtro = request.GET.get("campus", "")
    periodo_seleccionado = None
    carrera_seleccionada = None

    if periodo_id:
        paralelos = paralelos.filter(periodo_de_nivelacion__id=periodo_id)
        periodo_seleccionado = periodos.filter(id=periodo_id).first()
    if campus_filtro:
        paralelos = paralelos.filter(unidad_curricular__malla_curricular__carrera__campus__id=campus_filtro)
    if carrera_id:
        paralelos = paralelos.filter(unidad_curricular__malla_curricular__carrera__id=carrera_id)
        carrera_seleccionada = carreras.filter(id=carrera_id).first()
    if jornada_filtro:
        paralelos = paralelos.filter(jornada=jornada_filtro)

    # Role-based filtering
    rol = obtener_rol_usuario(request.user)
    es_coordinador_ua = (rol == ROL_COORDINADOR_UA)
    puede_distribuir = rol in (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN)

    if es_coordinador_ua:
        perfil_admin = getattr(request.user, 'perfil_administrativo', None)
        if perfil_admin and perfil_admin.carrera_asignada:
            paralelos = paralelos.filter(unidad_curricular__malla_curricular__carrera=perfil_admin.carrera_asignada)
            # Pre-load campus and carrera filters for coordinador UA
            if not campus_filtro and perfil_admin.carrera_asignada.campus:
                campus_filtro = str(perfil_admin.carrera_asignada.campus.id)
            if not carrera_id:
                carrera_seleccionada = perfil_admin.carrera_asignada
    elif rol == ROL_DOCENTE:
        perfil_docente = getattr(request.user, 'perfil_docente', None)
        if perfil_docente:
            paralelos = paralelos.filter(docente_responsable=perfil_docente)

    grupos = {}
    for p in paralelos:
        carrera = p.unidad_curricular.malla_curricular.carrera
        clave = (p.periodo_de_nivelacion_id, carrera.id, p.jornada, p.nombre)
        if clave not in grupos:
            grupos[clave] = {
                "representativo_id": p.id,
                "codigo": p.codigo_de_paralelo,
                "nombre": p.nombre,
                "campus": carrera.campus.nombre,
                "carrera": carrera.nombre,
                "periodo": p.periodo_de_nivelacion.periodo,
                "jornada": p.get_jornada_display(),
                "capacidad": p.capacidad_maxima,
                "unidades": 0,
                "estudiantes": p._n_estudiantes,
                "_orden": (carrera.nombre, _numero_para_orden(p.nombre)),
            }
        grupos[clave]["unidades"] += 1

    paralelos_agrupados = sorted(grupos.values(), key=lambda g: g["codigo"])

    from academico.permisos import obtener_permisos_periodo
    permisos = obtener_permisos_periodo(universidad_usuario)

    return render(request, "academico/listar_paralelos.html", {
        "paralelos": paralelos_agrupados,
        "periodos": periodos,
        "carreras": carreras,
        "campus_disponibles": campus_disponibles,
        "campus_filtro": campus_filtro,
        "jornadas": [j.value for j in Jornada],
        "periodo_seleccionado": periodo_seleccionado,
        "carrera_seleccionada": carrera_seleccionada,
        "jornada_filtro": jornada_filtro or "",
        "solo_lectura": usuario_es_solo_lectura(request.user),
        "puede_distribuir": puede_distribuir and permisos["puede_distribuir_paralelos"],
        "puede_eliminar_paralelo": permisos["puede_eliminar_paralelo"] and puede_distribuir,
        "es_coordinador_ua": es_coordinador_ua,
        "titulo_pagina": "Paralelo - NIVEC",
        "titulo": "Paralelos",
        "url_volver": "panel_principal",
    })

@requiere_perfil(ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN)
def generar_paralelos(request):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    periodos = PeriodoDeNivelacion.objects.filter(
        universidad=universidad_usuario,
        estado=EstadoDePeriodo.PLANIFICACION.value,
    )

    if not periodos.exists():
        messages.warning(request, "No existen Periodos de nivelación en planificación")
        return redirect("panel_dan")

    if request.method == "POST":
        periodo_id = request.POST.get("periodo") or None
        periodo = periodos.filter(id=periodo_id).first() if periodo_id else None
        if not periodo:
            messages.error(request, "Especifique un Periodo de nivelación en planificación")
            return redirect("generar_paralelos")


        capacidad = request.POST.get("capacidad") or 35
        try:
            capacidad = int(capacidad)
        except (TypeError, ValueError):
            capacidad = 35
        if capacidad < 20 or capacidad > 50:
            messages.error(request, "La capacidad máxima debe ser entre 20 y 50 estudiantes")
            return redirect("generar_paralelos")

        resumen = servicio_generar_paralelos(periodo, capacidad)

        for advertencia in resumen["advertencias"]:
            messages.warning(request, advertencia)

        if resumen["paralelos_creados"] > 0 or resumen["estudiantes_distribuidos"] > 0:
            messages.success(
                request,
                f"{resumen['grupos_creados']} Paralelos creados. {resumen['estudiantes_distribuidos']} Estudiantes distribuidos"
            )
        else:
            messages.warning(
                request,
                "No quedan Estudiantes pendientes por distribuir actualmente"
            )

        return redirect("listar_paralelos")

    return render(request, "academico/generar_paralelos.html", {
        "periodos": periodos,
        "titulo_pagina": "Paralelo - NIVEC",
        "titulo": "Distribuir Paralelos",
        "url_cancelar": "listar_paralelos",
    })

@requiere_perfil(*ROLES_MODIFICAN)
def eliminar_paralelo(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    periodo = representativo.periodo_de_nivelacion
    carrera = representativo.unidad_curricular.malla_curricular.carrera
    paralelos_grupo = _paralelos_del_grupo(representativo)

    with transaction.atomic():
        MatriculaParalelo.objects.filter(paralelo__in=paralelos_grupo).delete()
        paralelos_grupo.delete()

    servicio_recalcular_cohorte_de_carrera(periodo, carrera)

    messages.success(request, f"El Paralelo ha sido eliminado correctamente")
    return redirect("listar_paralelos")

@requiere_perfil(*ROLES_VISUALIZAN)
def detalle_paralelo(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    carrera = representativo.unidad_curricular.malla_curricular.carrera

    unidades = _paralelos_del_grupo(representativo).select_related(
        "unidad_curricular",
        "docente_responsable__usuario_de_sistema",
    ).order_by("unidad_curricular__nombre")

    # Check if horario exists for the paralelo group
    tiene_horario = Horario.objects.filter(
        paralelo__in=unidades
    ).exists()

    # Check period state for conditional buttons
    from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
    periodo = representativo.periodo_de_nivelacion
    es_evaluacion_o_cerrado = periodo.estado in (
        EstadoDePeriodo.EVALUACION.value, EstadoDePeriodo.CERRADO.value
    )

    # Role-based access control
    rol = obtener_rol_usuario(request.user)
    if rol == ROL_COORDINADOR_UA:
        perfil_admin = getattr(request.user, 'perfil_administrativo', None)
        if perfil_admin and perfil_admin.carrera_asignada and carrera != perfil_admin.carrera_asignada:
            messages.error(request, "No tiene acceso a este paralelo")
            return redirect("panel_principal")
    elif rol == ROL_DOCENTE:
        perfil_docente = getattr(request.user, 'perfil_docente', None)
        if perfil_docente and not unidades.filter(docente_responsable=perfil_docente).exists():
            messages.error(request, "No tiene acceso a este paralelo")
            return redirect("panel_principal")

    return render(request, "academico/detalle_paralelo.html", {
        "representativo": representativo,
        "carrera": carrera,
        "unidades": unidades,
        "tiene_horario": tiene_horario,
        "es_evaluacion_o_cerrado": es_evaluacion_o_cerrado,
        "solo_lectura": usuario_es_solo_lectura(request.user),
        "titulo_pagina": "Paralelo - NIVEC",
        "titulo": f"{representativo.nombre} - {carrera.nombre} ({representativo.get_jornada_display()})",
    })

@requiere_perfil(*ROLES_VISUALIZAN)
def listar_estudiantes_paralelo(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    carrera = paralelo.unidad_curricular.malla_curricular.carrera

    # Role-based access control
    rol = obtener_rol_usuario(request.user)
    if rol == ROL_DOCENTE:
        perfil_docente = getattr(request.user, 'perfil_docente', None)
        if perfil_docente and paralelo.docente_responsable != perfil_docente:
            messages.error(request, "No tiene acceso a este paralelo")
            return redirect("panel_principal")

    matriculas = MatriculaParalelo.objects.filter(
        paralelo=paralelo
    ).select_related("estudiante__usuario_de_sistema")

    periodo = paralelo.periodo_de_nivelacion
    ocupacion = matriculas.count()
    capacidad = paralelo.capacidad_maxima
    puede_gestionar = periodo_permite_gestion_matriculas(periodo)

    otros = Paralelo.objects.filter(
        periodo_de_nivelacion=paralelo.periodo_de_nivelacion,
        jornada=paralelo.jornada,
        unidad_curricular__malla_curricular__carrera=carrera,
    ).exclude(nombre=paralelo.nombre).annotate(_n_estudiantes=Count("estudiantes_matriculados"))

    representativos = {}
    for otro in otros:
        if otro.nombre not in representativos:
            representativos[otro.nombre] = otro

    destinos = []
    for nombre_grupo in sorted(representativos.keys()):
        rep = representativos[nombre_grupo]
        ocupacion_destino = rep._n_estudiantes
        destinos.append({
            "nombre": nombre_grupo,
            "paralelo_id": rep.id,
            "ocupacion": ocupacion_destino,
            "capacidad": rep.capacidad_maxima,
            "lleno": ocupacion_destino >= rep.capacidad_maxima,
        })

    # Check if there are students available to add
    hay_estudiantes_disponibles = PerfilEstudiante.objects.filter(
        carrera_registrada=carrera,
        jornada=paralelo.jornada,
        periodo_de_nivelacion=periodo,
    ).exclude(
        estado_de_matricula__in=[EstadoDeMatricula.RETIRADO.value, EstadoDeMatricula.ANULADO.value]
    ).exclude(
        estudiantes_matriculados__paralelo__periodo_de_nivelacion=periodo
    ).exists()

    return render(request, "academico/estudiantes_paralelo.html", {
        "paralelo": paralelo,
        "matriculas": matriculas,
        "destinos": destinos,
        "puede_gestionar": puede_gestionar,
        "hay_estudiantes_disponibles": hay_estudiantes_disponibles,
        "ocupacion": ocupacion,
        "capacidad": capacidad,
        "lleno": ocupacion >= capacidad,
        "solo_lectura": usuario_es_solo_lectura(request.user),
        "titulo_pagina": "Paralelo - NIVEC",
        "titulo": f"{paralelo.nombre} ({paralelo.unidad_curricular.nombre})",
    })

@requiere_perfil(*ROLES_MODIFICAN)
def mover_estudiante(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    if request.method == "POST":
        estudiante_id = request.POST.get("estudiante") or None
        destino_id = request.POST.get("paralelo_destino") or None

        if not estudiante_id or not destino_id:
            messages.error(request, "Especifique un Estudiante y un Paralelo de destino")
            return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)

        estudiante = get_object_or_404(
            PerfilEstudiante,
            id=estudiante_id,
            carrera_registrada__campus__universidad=universidad_usuario,
        )
        paralelo_destino = get_object_or_404(
            Paralelo,
            id=destino_id,
            periodo_de_nivelacion__universidad=universidad_usuario,
        )

        ok, mensaje = servicio_mover_estudiante(estudiante, paralelo_destino)
        if ok:
            messages.success(request, mensaje)
        else:
            messages.error(request, mensaje)

    return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)


@requiere_perfil(*ROLES_MODIFICAN)
def retirar_estudiante(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    if request.method == "POST":
        estudiante_id = request.POST.get("estudiante") or None
        if not estudiante_id:
            messages.error(request, "Especifique un Estudiante")
            return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)

        estudiante = get_object_or_404(
            PerfilEstudiante,
            id=estudiante_id,
            carrera_registrada__campus__universidad=universidad_usuario,
        )
        ok, mensaje = servicio_retirar_estudiante_de_paralelo(estudiante, paralelo)
        (messages.success if ok else messages.error)(request, mensaje)

    return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)


@requiere_perfil(*ROLES_MODIFICAN)
def estudiantes_disponibles(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    carrera = paralelo.unidad_curricular.malla_curricular.carrera
    periodo = paralelo.periodo_de_nivelacion

    if not periodo_permite_gestion_matriculas(periodo):
        messages.warning(request, "Solo se puede gestionar la matrícula en Periodos en planificación o en curso")
        return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)

    ocupacion = MatriculaParalelo.objects.filter(paralelo=paralelo).count()
    capacidad = paralelo.capacidad_maxima

    if ocupacion >= capacidad:
        messages.warning(request, "El Paralelo no presenta cupos disponibles")
        return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)

    no_asignados = PerfilEstudiante.objects.filter(
        carrera_registrada=carrera,
        jornada=paralelo.jornada,
        periodo_de_nivelacion=periodo,
    ).exclude(
        estado_de_matricula__in=[EstadoDeMatricula.RETIRADO.value, EstadoDeMatricula.ANULADO.value]
    ).exclude(
        estudiantes_matriculados__paralelo__periodo_de_nivelacion=periodo
    ).distinct().select_related("usuario_de_sistema").order_by(
        "usuario_de_sistema__apellidos", "usuario_de_sistema__nombres"
    )

    if not no_asignados.exists():
        messages.warning(request, "No existen estudiantes sin asignar para esta Carrera y Jornada")
        return redirect("listar_estudiantes_paralelo", paralelo_id=paralelo.id)

    return render(request, "academico/estudiantes_disponibles.html", {
        "paralelo": paralelo,
        "no_asignados": no_asignados,
        "ocupacion": ocupacion,
        "capacidad": capacidad,
        "titulo_pagina": "Paralelo - NIVEC",
        "titulo": f"Incorporar Estudiante - {paralelo.nombre} ({paralelo.unidad_curricular.nombre})",
    })


@requiere_perfil(*ROLES_MODIFICAN)
def agregar_estudiante(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    if request.method == "POST":
        estudiante_id = request.POST.get("estudiante") or None
        if not estudiante_id:
            messages.error(request, "Especifique un Estudiante")
            return redirect("estudiantes_disponibles", paralelo_id=paralelo.id)

        estudiante = get_object_or_404(
            PerfilEstudiante,
            id=estudiante_id,
            carrera_registrada__campus__universidad=universidad_usuario,
        )
        ok, mensaje = servicio_agregar_estudiante_a_paralelo(estudiante, paralelo)
        (messages.success if ok else messages.error)(request, mensaje)

    return redirect("estudiantes_disponibles", paralelo_id=paralelo.id)



@requiere_perfil(*ROLES_VISUALIZAN)
def descargar_info_paralelo(request, paralelo_id):
    import openpyxl
    from django.http import HttpResponse
    from academico.models import Horario

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    carrera = representativo.unidad_curricular.malla_curricular.carrera
    periodo = representativo.periodo_de_nivelacion

    # Todas las filas del paralelo lógico
    unidades_rows = Paralelo.objects.filter(
        periodo_de_nivelacion=periodo,
        jornada=representativo.jornada,
        nombre=representativo.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    ).select_related(
        "unidad_curricular", "docente_responsable__usuario_de_sistema"
    ).order_by("unidad_curricular__nombre")

    wb = openpyxl.Workbook()

    # === HOJA 1: Información general ===
    ws = wb.active
    ws.title = "Información general"

    ws.append(["Información registrada del Paralelo"])
    ws.append([])
    ws.append(["Nombre", representativo.nombre])
    ws.append(["Código de Paralelo", representativo.codigo_de_paralelo])
    ws.append(["Carrera asignada", carrera.nombre])
    ws.append(["Jornada", representativo.get_jornada_display()])
    ws.append(["Modalidad", representativo.get_modalidad_display()])
    ws.append(["Periodo de nivelación", periodo.periodo])
    ws.append(["Año", periodo.anio])
    ws.append(["Fecha de inicio", str(periodo.fecha_inicio)])
    ws.append(["Fecha de finalización", str(periodo.fecha_fin)])
    ws.append(["Capacidad máxima de estudiantes especificada", representativo.capacidad_maxima])
    estudiantes_count = MatriculaParalelo.objects.filter(paralelo=representativo).count()
    ws.append(["Número de estudiantes registrados", estudiantes_count])
    ws.append([])

    # Unidades con docente y horario
    ws.append(["Unidades curriculares"])
    ws.append(["Código de Unidad curricular", "Nombre", "Docente designado", "Horario"])
    for row in unidades_rows:
        docente = ""
        if row.docente_responsable:
            d = row.docente_responsable.usuario_de_sistema
            docente = f"{d.nombres} {d.apellidos}"
        horarios_txt = ", ".join(
            f"{h.get_dia_semana_display()} {h.hora_inicio.strftime('%H:%M')}-{h.hora_fin.strftime('%H:%M')}"
            for h in Horario.objects.filter(paralelo=row).order_by("dia_semana", "hora_inicio")
        )
        ws.append([
            row.unidad_curricular.codigo_de_unidad,
            row.unidad_curricular.nombre,
            docente,
            horarios_txt,
        ])

    for col in range(1, 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 35

    # === HOJAS 2+: Una por unidad con listado de estudiantes ===
    matriculas = MatriculaParalelo.objects.filter(
        paralelo=representativo
    ).select_related(
        "estudiante__usuario_de_sistema",
    ).order_by(
        "estudiante__usuario_de_sistema__apellidos",
        "estudiante__usuario_de_sistema__nombres",
    )

    for row in unidades_rows:
        titulo_hoja = row.unidad_curricular.nombre[:31]
        ws_u = wb.create_sheet(title=titulo_hoja)

        ws_u.append([
            "Tipo de identificación",
            "Número de identificación",
            "Apellidos",
            "Nombres",
            "Correo institucional",
            "Número de matrícula",
            "Registro de cupo",
            "Estado de matrícula",
        ])

        for mat in matriculas:
            est = mat.estudiante
            usr = est.usuario_de_sistema
            ws_u.append([
                usr.tipo_de_identificacion,
                usr.identificacion,
                usr.apellidos,
                usr.nombres,
                usr.correo_institucional,
                est.numero_de_matricula,
                est.registro_de_cupo,
                est.estado_de_matricula,
            ])

        for col in range(1, 9):
            ws_u.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Nombre del archivo: paralelo_letra_carrera (todo en minúscula)
    nombre_archivo = f"{representativo.nombre}_{carrera.nombre}".lower().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response