from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from academico.models import Paralelo, Horario, PeriodoDeNivelacion, Carrera
from academico.services import servicio_registrar_horario, servicio_horas_agendadas_paralelo, servicio_obtener_matriz_de_horarios, servicio_generar_horario_sugerido, servicio_editar_horario, periodo_en_planificacion, _horas_sincronicas_semanales
from poo.clases.enums.dia_de_semana import DiaDeSemana
from poo.clases.enums.jornada import Jornada
from poo.clases.franja_horaria import texto_franja, obtener_franja
from usuarios.utils import (
    requiere_perfil,
    usuario_es_solo_lectura,
    ROL_COORDINADOR_DAN,
    ROL_DIRECTOR_DAN,
    ROL_RECTOR,
    ROL_VICERRECTOR,
    ROL_COORDINADOR_UA,
    ROL_DOCENTE,
)

ROLES_VISUALIZAN = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR, ROL_COORDINADOR_UA, ROL_DOCENTE)
ROLES_MODIFICAN = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_COORDINADOR_UA)

def _parsear_hora(valor):
    try:
        return datetime.strptime(valor, "%H:%M").time()
    except (TypeError, ValueError):
        return None

def _unidades_del_grupo(representativo):
    carrera = representativo.unidad_curricular.malla_curricular.carrera
    return Paralelo.objects.filter(
        periodo_de_nivelacion=representativo.periodo_de_nivelacion,
        jornada=representativo.jornada,
        nombre=representativo.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    ).select_related("unidad_curricular", "docente_responsable__usuario_de_sistema").order_by("unidad_curricular__nombre")


def _obtener_universidad_usuario(user):
    """Get universidad from perfil_administrativo or perfil_docente."""
    perfil_admin = getattr(user, 'perfil_administrativo', None)
    if perfil_admin:
        return perfil_admin.universidad
    perfil_docente = getattr(user, 'perfil_docente', None)
    if perfil_docente:
        return perfil_docente.universidad
    return None


@requiere_perfil(*ROLES_VISUALIZAN)
def listar_horarios_paralelo(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    unidades_rows = list(_unidades_del_grupo(representativo))

    unidades = []
    todas_completas = True
    for row in unidades_rows:
        agendadas = servicio_horas_agendadas_paralelo(row)
        requeridas = _horas_sincronicas_semanales(row.unidad_curricular, row.periodo_de_nivelacion)
        completo = agendadas >= requeridas
        todas_completas = todas_completas and completo
        unidades.append({
            "id": row.id,
            "codigo": row.unidad_curricular.codigo_de_unidad,
            "nombre": row.unidad_curricular.nombre,
            "docente": row.docente_responsable,
            "agendadas": agendadas,
            "requeridas": requeridas,
            "restantes": round(max(requeridas - agendadas, 0), 2),
            "completo": completo,
        })

    horarios = Horario.objects.filter(paralelo__in=unidades_rows).select_related(
        "paralelo__unidad_curricular"
    ).order_by("paralelo__unidad_curricular__nombre", "dia_semana", "hora_inicio")

    jornada_enum = Jornada(representativo.jornada)

    # Construir datos de la grilla visual (Día × Hora)
    franja = obtener_franja(jornada_enum)
    slots_hora = []
    if franja:
        h = franja[0].hour
        while h < franja[1].hour:
            slots_hora.append(h)
            h += 1

    # Asignar colores en escalas de gris por unidad
    _COLORES_UNIDAD = [
        "#e8e8ed", "#d2d2d7", "#c7c7cc", "#b0b0b5", "#9a9a9f",
        "#aeaeb2", "#dcdce0", "#c4c4c8", "#bababe", "#8e8e93",
    ]
    nombres_unidades = sorted(set(h.paralelo.unidad_curricular.nombre for h in horarios))
    mapa_colores = {nombre: _COLORES_UNIDAD[i % len(_COLORES_UNIDAD)] for i, nombre in enumerate(nombres_unidades)}

    dias_semana = [d.value for d in DiaDeSemana]  # Lunes a Domingo (7 días)
    grilla = []  # lista de {hora, celdas: [{bloque o None} por día]}
    for slot in slots_hora:
        fila = {"hora": f"{slot:02d}:00", "celdas": []}
        for dia in dias_semana:
            bloque = None
            for h in horarios:
                if h.dia_semana == dia and h.hora_inicio.hour <= slot < h.hora_fin.hour:
                    bloque = {
                        "nombre": h.paralelo.unidad_curricular.nombre,
                        "hora_inicio": h.hora_inicio.strftime("%H:%M"),
                        "hora_fin": h.hora_fin.strftime("%H:%M"),
                        "slot_inicio": f"{slot:02d}:00",
                        "slot_fin": f"{slot+1:02d}:00",
                        "color": mapa_colores.get(h.paralelo.unidad_curricular.nombre, "#e8e8ed"),
                    }
                    break
            fila["celdas"].append(bloque)
        grilla.append(fila)

    return render(request, "academico/horarios_paralelo.html", {
        "representativo": representativo,
        "unidades": unidades,
        "horarios": horarios,
        "todas_completas": todas_completas,
        "es_planificacion": periodo_en_planificacion(representativo.periodo_de_nivelacion),
        "franja": texto_franja(jornada_enum),
        "dias": [dia.value for dia in DiaDeSemana],
        "dias_semana": dias_semana,
        "grilla": grilla,
        "mapa_colores": mapa_colores,
        "solo_lectura": usuario_es_solo_lectura(request.user),
        "titulo_pagina": "Horario - NIVEC",
        "titulo": f"Horarios - {representativo.nombre} ({representativo.unidad_curricular.malla_curricular.carrera.nombre})",
    })

@requiere_perfil(*ROLES_MODIFICAN)
def registrar_horario(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    if request.method == "POST":
        unidad_id = request.POST.get("unidad_paralelo") or None
        fila = _unidades_del_grupo(representativo).filter(id=unidad_id).first() if unidad_id else None
        if not fila:
            messages.error(request, "Especifique una Unidad curricular válida")
            return redirect("listar_horarios_paralelo", paralelo_id=representativo.id)

        dia_semana = request.POST.get("dia_semana")
        espacio = (request.POST.get("espacio_de_imparticion") or "").strip()
        hora_inicio = _parsear_hora(request.POST.get("hora_inicio"))
        hora_fin = _parsear_hora(request.POST.get("hora_fin"))

        if not hora_inicio or not hora_fin:
            messages.error(request, "Las horas de inicio y finalización son requeridas")
            return redirect("listar_horarios_paralelo", paralelo_id=representativo.id)

        ok, mensaje = servicio_registrar_horario(
            fila, dia_semana, hora_inicio, hora_fin, espacio
        )
        (messages.success if ok else messages.error)(request, mensaje)

    return redirect("listar_horarios_paralelo", paralelo_id=representativo.id)

@requiere_perfil(*ROLES_MODIFICAN)
def editar_horario(request, horario_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    horario = get_object_or_404(
        Horario, id=horario_id,
        paralelo__periodo_de_nivelacion__universidad=universidad_usuario,
    )
    paralelo = horario.paralelo

    if request.method == "POST":
        dia_semana = request.POST.get("dia_semana")
        espacio = (request.POST.get("espacio_de_imparticion") or "").strip()
        hora_inicio = _parsear_hora(request.POST.get("hora_inicio"))
        hora_fin = _parsear_hora(request.POST.get("hora_fin"))

        if not hora_inicio or not hora_fin:
            messages.error(request, "Las horas de inicio y finalización son requeridas")
            return redirect("editar_horario", horario_id=horario.id)

        ok, mensaje = servicio_editar_horario(
            horario, dia_semana, hora_inicio, hora_fin, espacio
        )
        if ok:
            messages.success(request, mensaje)
            return redirect("listar_horarios_paralelo", paralelo_id=paralelo.id)
        messages.error(request, mensaje)

    jornada_enum = Jornada(paralelo.jornada)
    return render(request, "academico/editar_horario.html", {
        "horario": horario,
        "paralelo": paralelo,
        "franja": texto_franja(jornada_enum),
        "dias": [dia.value for dia in DiaDeSemana],
        "titulo_pagina": "Horario - NIVEC",
        "titulo": f"Editar sesión - {paralelo.unidad_curricular.nombre}",
    })

@requiere_perfil(*ROLES_MODIFICAN)
def generar_horario_sugerido(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    representativo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    if request.method == "POST":
        ok, mensaje = servicio_generar_horario_sugerido(representativo)
        (messages.success if ok else messages.error)(request, mensaje)

    return redirect("listar_horarios_paralelo", paralelo_id=representativo.id)

@requiere_perfil(*ROLES_MODIFICAN)
def eliminar_horario(request, horario_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    horario = get_object_or_404(
        Horario, id=horario_id,
        paralelo__periodo_de_nivelacion__universidad=universidad_usuario,
    )
    paralelo_id = horario.paralelo_id
    horario.delete()
    messages.success(request, "El horario ha sido eliminado correctamente")
    return redirect("listar_horarios_paralelo", paralelo_id=paralelo_id)

@requiere_perfil(*ROLES_VISUALIZAN)
def matriz_horarios(request):
    from academico.models import Carrera
    from poo.clases.franja_horaria import obtener_franja

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    periodos = PeriodoDeNivelacion.objects.filter(
        universidad=universidad_usuario
    ).order_by("-anio", "numero_periodo")
    carreras = Carrera.objects.filter(campus__universidad=universidad_usuario).order_by("nombre")
    jornadas = [j.value for j in Jornada]

    periodo_id = request.GET.get("periodo") or None
    carrera_id = request.GET.get("carrera") or None
    jornada_filtro = request.GET.get("jornada") or None
    paralelo_id = request.GET.get("paralelo") or None

    periodo_seleccionado = None
    carrera_seleccionada = None
    grupos = []
    paralelo_seleccionado = None
    grilla = []
    dias_semana = []
    mapa_colores = {}
    franja = None

    if periodo_id:
        periodo_seleccionado = PeriodoDeNivelacion.objects.filter(
            id=periodo_id, universidad=universidad_usuario
        ).first()
    if carrera_id:
        carrera_seleccionada = carreras.filter(id=carrera_id).first()

    # Build paralelo groups based on filters
    if periodo_seleccionado:
        paralelos_qs = Paralelo.objects.filter(
            periodo_de_nivelacion=periodo_seleccionado
        ).select_related("unidad_curricular__malla_curricular__carrera")

        if carrera_seleccionada:
            paralelos_qs = paralelos_qs.filter(
                unidad_curricular__malla_curricular__carrera=carrera_seleccionada
            )
        if jornada_filtro:
            paralelos_qs = paralelos_qs.filter(jornada=jornada_filtro)

        grupos_vistos = {}
        for p in paralelos_qs:
            carrera_p = p.unidad_curricular.malla_curricular.carrera
            clave = f"{carrera_p.id}|{p.jornada}|{p.nombre}"
            if clave not in grupos_vistos:
                grupos_vistos[clave] = {
                    "clave": clave,
                    "carrera": carrera_p.nombre,
                    "jornada_display": p.get_jornada_display(),
                    "nombre": p.nombre,
                    "representativo_id": p.id,
                }
        grupos = sorted(grupos_vistos.values(), key=lambda g: (g["carrera"], g["nombre"]))

        # If a paralelo is selected, build grilla
        if paralelo_id:
            paralelo_seleccionado = Paralelo.objects.filter(
                id=paralelo_id, periodo_de_nivelacion=periodo_seleccionado
            ).select_related("unidad_curricular__malla_curricular__carrera").first()

            if paralelo_seleccionado:
                unidades_rows = list(_unidades_del_grupo(paralelo_seleccionado))
                horarios = Horario.objects.filter(
                    paralelo__in=unidades_rows
                ).select_related("paralelo__unidad_curricular")

                jornada_enum = Jornada(paralelo_seleccionado.jornada)
                franja_rango = obtener_franja(jornada_enum)
                franja = texto_franja(jornada_enum)
                slots_hora = []
                if franja_rango:
                    h = franja_rango[0].hour
                    while h < franja_rango[1].hour:
                        slots_hora.append(h)
                        h += 1

                _COLORES_UNIDAD = [
                    "#e8e8ed", "#d2d2d7", "#c7c7cc", "#b0b0b5", "#9a9a9f",
                    "#aeaeb2", "#dcdce0", "#c4c4c8", "#bababe", "#8e8e93",
                ]
                nombres_unidades = sorted(set(h.paralelo.unidad_curricular.nombre for h in horarios))
                mapa_colores = {nombre: _COLORES_UNIDAD[i % len(_COLORES_UNIDAD)] for i, nombre in enumerate(nombres_unidades)}

                dias_semana = [d.value for d in DiaDeSemana]
                for slot in slots_hora:
                    fila = {"hora": f"{slot:02d}:00", "celdas": []}
                    for dia in dias_semana:
                        bloque = None
                        for h_obj in horarios:
                            if h_obj.dia_semana == dia and h_obj.hora_inicio.hour <= slot < h_obj.hora_fin.hour:
                                bloque = {
                                    "nombre": h_obj.paralelo.unidad_curricular.nombre,
                                    "hora_inicio": h_obj.hora_inicio.strftime("%H:%M"),
                                    "hora_fin": h_obj.hora_fin.strftime("%H:%M"),
                                    "slot_inicio": f"{slot:02d}:00",
                                    "slot_fin": f"{slot+1:02d}:00",
                                    "color": mapa_colores.get(h_obj.paralelo.unidad_curricular.nombre, "#e8e8ed"),
                                }
                                break
                        fila["celdas"].append(bloque)
                    grilla.append(fila)

    return render(request, "academico/matriz_horarios.html", {
        "periodos": periodos,
        "carreras": carreras,
        "jornadas": jornadas,
        "periodo_seleccionado": periodo_seleccionado,
        "carrera_seleccionada": carrera_seleccionada,
        "jornada_filtro": jornada_filtro or "",
        "grupos": grupos,
        "paralelo_seleccionado": paralelo_seleccionado,
        "grilla": grilla,
        "dias_semana": dias_semana,
        "mapa_colores": mapa_colores,
        "franja": franja,
        "titulo_pagina": "Horario - NIVEC",
        "titulo": "Matriz de horarios",
    })


@requiere_perfil(*ROLES_VISUALIZAN)
def descargar_horarios_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from academico.models import Carrera

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    periodo_id = request.GET.get("periodo")
    carrera_id = request.GET.get("carrera")

    if not periodo_id or not carrera_id:
        messages.error(request, "Debe especificar un Periodo y una Carrera")
        return redirect("matriz_horarios")

    periodo = get_object_or_404(PeriodoDeNivelacion, id=periodo_id, universidad=universidad_usuario)
    carrera = get_object_or_404(Carrera, id=carrera_id, campus__universidad=universidad_usuario)

    paralelos = Paralelo.objects.filter(
        periodo_de_nivelacion=periodo,
        unidad_curricular__malla_curricular__carrera=carrera,
    ).select_related("unidad_curricular").order_by("jornada", "nombre")

    # Group by jornada
    jornadas_dict = {}
    for p in paralelos:
        jornadas_dict.setdefault(p.jornada, []).append(p)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for jornada_valor, paralelos_jornada in jornadas_dict.items():
        try:
            jornada_nombre = Jornada(jornada_valor).value
        except (ValueError, KeyError):
            jornada_nombre = jornada_valor

        ws = wb.create_sheet(title=str(jornada_nombre)[:31])
        ws.append(["Código de paralelo", "Nombre de paralelo", "Unidad curricular", "Día", "Hora inicio", "Hora fin", "Espacio"])

        # Orden personalizado de días para ordenar Lunes primero
        _ORDEN_DIA = {"Lunes": 1, "Martes": 2, "Miércoles": 3, "Jueves": 4, "Viernes": 5, "Sábado": 6, "Domingo": 7}

        horarios = list(Horario.objects.filter(
            paralelo__in=paralelos_jornada
        ).select_related("paralelo__unidad_curricular"))

        horarios.sort(key=lambda h: (_ORDEN_DIA.get(h.dia_semana, 99), h.hora_inicio, h.paralelo.nombre))

        for h in horarios:
            ws.append([
                h.paralelo.codigo_de_paralelo,
                h.paralelo.nombre,
                h.paralelo.unidad_curricular.nombre,
                h.dia_semana,
                h.hora_inicio.strftime("%H:%M"),
                h.hora_fin.strftime("%H:%M"),
                h.espacio_de_imparticion or "",
            ])

        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sin datos")
        ws.append(["No se encontraron horarios para la carrera y periodo seleccionados"])

    nombre_archivo = f"horarios_{carrera.nombre}_{periodo.periodo}".lower().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response