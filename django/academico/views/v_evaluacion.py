from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl

from academico.models import Paralelo, EvaluacionAcademica, PeriodoDeNivelacion, MatriculaParalelo, UnidadCurricular
from academico.services import servicio_cargar_calificaciones_desde_excel
from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
from usuarios.utils import (
    requiere_perfil, usuario_es_solo_lectura, obtener_rol_usuario,
    ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR, ROL_COORDINADOR_UA, ROL_DOCENTE,
)

ROLES_CARGAR = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_COORDINADOR_UA, ROL_DOCENTE)
ROLES_VER = (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR, ROL_COORDINADOR_UA, ROL_DOCENTE, "ESTUDIANTE")


def _obtener_universidad_usuario(user):
    """Get universidad from perfil_administrativo, perfil_docente, or perfil_estudiante."""
    perfil_admin = getattr(user, 'perfil_administrativo', None)
    if perfil_admin:
        return perfil_admin.universidad
    perfil_docente = getattr(user, 'perfil_docente', None)
    if perfil_docente:
        return perfil_docente.universidad
    # For students
    from usuarios.models import PerfilEstudiante
    perfil_est = PerfilEstudiante.objects.filter(usuario_de_sistema=user).select_related("carrera_registrada__campus__universidad").first()
    if perfil_est and perfil_est.carrera_registrada:
        return perfil_est.carrera_registrada.campus.universidad
    return None


@requiere_perfil(*ROLES_VER)
def listar_evaluaciones_paralelo(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    periodo = paralelo.periodo_de_nivelacion

    if periodo.estado not in (EstadoDePeriodo.EVALUACION.value, EstadoDePeriodo.CERRADO.value):
        messages.warning(request, "Las calificaciones solo están disponibles cuando el Periodo está en evaluación o cerrado")
        return redirect("detalle_paralelo", paralelo_id=paralelo.id)

    evaluaciones = EvaluacionAcademica.objects.filter(
        unidad_curricular=paralelo.unidad_curricular,
        periodo_de_nivelacion=periodo,
        estudiante__estudiantes_matriculados__paralelo=paralelo,
    ).select_related("estudiante__usuario_de_sistema").order_by(
        "estudiante__usuario_de_sistema__apellidos"
    )

    es_evaluacion = (periodo.estado == EstadoDePeriodo.EVALUACION.value)

    rol = obtener_rol_usuario(request.user)
    es_docente = (rol == ROL_DOCENTE)
    es_coordinador_ua = (rol == ROL_COORDINADOR_UA)
    es_coordinador_o_director = rol in (ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN)
    es_administrativo = rol in (ROL_RECTOR, ROL_VICERRECTOR)

    # Determine calificaciones state
    todas_en_revision = evaluaciones.exists() and not evaluaciones.filter(estado_revision="Borrador").exists()
    todas_formalizadas = evaluaciones.exists() and not evaluaciones.exclude(estado_revision="Formalizado").exists()
    hay_en_revision = evaluaciones.filter(estado_revision="En revisión").exists()

    # Role-based permissions for calificaciones actions
    ya_tiene_calificaciones = evaluaciones.exists()
    
    # Point 1: Hide cargar once calificaciones have been processed (evaluaciones exist)
    puede_cargar = es_evaluacion and (es_docente or es_coordinador_ua) and not ya_tiene_calificaciones
    
    # Point 2: Docente can only pass to revisión if evaluaciones are in Borrador and exist
    puede_pasar_revision = es_evaluacion and es_docente and ya_tiene_calificaciones and not todas_en_revision and not todas_formalizadas
    
    # Coordinador UA can formalizar if evaluaciones exist and not all formalized
    puede_formalizar = es_evaluacion and es_coordinador_ua and ya_tiene_calificaciones and not todas_formalizadas
    
    # Docente can edit only in Borrador; Coordinador UA can edit in Borrador and En revisión (not Formalizado)
    # If coordinador UA uploaded (todas_en_revision without docente passing), docente can't edit
    puede_editar_calificacion = es_evaluacion and (
        (es_docente and ya_tiene_calificaciones and not todas_en_revision and not todas_formalizadas) or
        (es_coordinador_ua and ya_tiene_calificaciones and not todas_formalizadas)
    )

    return render(request, "academico/evaluaciones_paralelo.html", {
        "paralelo": paralelo,
        "evaluaciones": evaluaciones,
        "periodo": periodo,
        "es_evaluacion": es_evaluacion,
        "solo_lectura": usuario_es_solo_lectura(request.user),
        "puede_cargar": puede_cargar,
        "puede_pasar_revision": puede_pasar_revision,
        "puede_formalizar": puede_formalizar,
        "puede_editar_calificacion": puede_editar_calificacion,
        "todas_formalizadas": todas_formalizadas,
        "titulo_pagina": "Resultados - NIVEC",
        "titulo": f"Registro de calificaciones - {paralelo.nombre} ({paralelo.unidad_curricular.nombre})",
    })


@requiere_perfil(*ROLES_CARGAR)
def cargar_calificaciones(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    periodo = paralelo.periodo_de_nivelacion

    if periodo.estado != EstadoDePeriodo.EVALUACION.value:
        messages.error(request, "El Periodo de nivelación no se encuentra en estado de evaluación")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    if request.method == "POST":
        if "archivo_excel" not in request.FILES:
            messages.error(request, "Debe seleccionar un documento para procesar")
            return redirect("cargar_calificaciones", paralelo_id=paralelo.id)
        
        archivo = request.FILES["archivo_excel"]
        if not archivo.name.endswith(".xlsx"):
            messages.error(request, "Documento con formato no válido")
            return redirect("cargar_calificaciones", paralelo_id=paralelo.id)

        resultado = servicio_cargar_calificaciones_desde_excel(
            archivo, paralelo, paralelo.unidad_curricular, periodo
        )

        if resultado["error"]:
            messages.error(request, resultado["error"])
        for adv in resultado["advertencias"]:
            messages.warning(request, adv)
        if resultado["exitosos"] > 0:
            messages.success(request, f"{resultado['exitosos']} calificaciones procesadas correctamente")

        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    return render(request, "academico/cargar_calificaciones.html", {
        "paralelo": paralelo,
        "titulo_pagina": "Resultados - NIVEC",
        "titulo": f"Procesar calificaciones - {paralelo.nombre} ({paralelo.unidad_curricular.nombre})",
    })


@requiere_perfil(*ROLES_CARGAR)
def descargar_plantilla_calificaciones(request, paralelo_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calificaciones"
    ws.append(["Número de identificación", "Apellidos", "Nombres", "Correo institucional", "Número de matrícula", "Calificación parcial 1 (0-10)", "Calificación parcial 2 (0-10)", "Porcentaje de asistencia (0-100)"])

    # Pre-fill with student data
    matriculas = MatriculaParalelo.objects.filter(
        paralelo=paralelo
    ).select_related("estudiante__usuario_de_sistema").order_by(
        "estudiante__usuario_de_sistema__apellidos"
    )
    for m in matriculas:
        ws.append([
            m.estudiante.usuario_de_sistema.identificacion,
            m.estudiante.usuario_de_sistema.apellidos,
            m.estudiante.usuario_de_sistema.nombres,
            m.estudiante.usuario_de_sistema.correo_institucional,
            m.estudiante.numero_de_matricula,
            "",
            "",
            "",
        ])

    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="calificaciones_{paralelo.nombre}_{paralelo.unidad_curricular.nombre}.xlsx"'.replace(" ", "_").lower()
    wb.save(response)
    return response


@requiere_perfil(*ROLES_VER)
def detalle_evaluacion(request, evaluacion_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    evaluacion = get_object_or_404(EvaluacionAcademica, id=evaluacion_id)
    unidad = evaluacion.unidad_curricular
    estudiante = evaluacion.estudiante
    usuario = estudiante.usuario_de_sistema

    # Métricas de aprobación definidas
    criterio_nota = unidad.criterio_de_aprobacion
    criterio_asistencia = unidad.porcentaje_minimo_asistencia

    # Métricas actuales del estudiante
    nota_actual = evaluacion.nota_final
    asistencia_actual = evaluacion.porcentaje_asistencia

    # Motivo detallado
    cumple_nota = nota_actual >= criterio_nota
    cumple_asistencia = asistencia_actual >= criterio_asistencia

    if evaluacion.estado_de_aprobacion == "Aprobado":
        motivo = "El Estudiante cumple con el criterio de calificación y porcentaje mínimo de asistencia"
    elif evaluacion.estado_de_aprobacion == "Retirado":
        motivo = "El Estudiante fue retirado del proceso de nivelación"
    elif evaluacion.estado_de_aprobacion == "Anulado":
        motivo = "La matrícula del Estudiante fue anulada"
    else:
        razones = []
        if not cumple_nota:
            razones.append(f"El registro de la nota final es inferior al criterio mínimo de aprobación establecido")
        if not cumple_asistencia:
            razones.append(f"El registro del porcentaje de asistencia es inferior al criterio mínimo de aprobación establecido")
        motivo = ". ".join(razones) if razones else "El estudiante no cumple con los criterios de aprobación establecidos"

    return render(request, "academico/detalle_evaluacion.html", {
        "evaluacion": evaluacion,
        "estudiante": estudiante,
        "usuario": usuario,
        "unidad": unidad,
        "criterio_nota": criterio_nota,
        "criterio_asistencia": criterio_asistencia,
        "nota_actual": nota_actual,
        "asistencia_actual": asistencia_actual,
        "cumple_nota": cumple_nota,
        "cumple_asistencia": cumple_asistencia,
        "motivo": motivo,
        "titulo_pagina": "Evaluación - NIVEC",
        "titulo": f"{usuario.apellidos} {usuario.nombres}",
    })



@requiere_perfil(*ROLES_CARGAR)
def editar_evaluacion(request, evaluacion_id):
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Institución no ha sido registrada actualmente")
        return redirect("panel_principal")

    evaluacion = get_object_or_404(EvaluacionAcademica, id=evaluacion_id)
    paralelo = Paralelo.objects.filter(
        unidad_curricular=evaluacion.unidad_curricular,
        periodo_de_nivelacion=evaluacion.periodo_de_nivelacion,
        estudiantes_matriculados__estudiante=evaluacion.estudiante,
    ).first()

    if not paralelo:
        messages.error(request, "No se encontró el paralelo asociado")
        return redirect("panel_principal")

    if evaluacion.estado_revision == "Formalizado":
        messages.error(request, "Las calificaciones formalizadas no se pueden editar")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    # Docentes can only edit Borrador; Coordinador UA can edit Borrador and En revisión
    rol = obtener_rol_usuario(request.user)
    if rol == ROL_DOCENTE and evaluacion.estado_revision != "Borrador":
        messages.error(request, "Las calificaciones en revisión no se pueden editar")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    periodo = evaluacion.periodo_de_nivelacion
    if periodo and periodo.estado != EstadoDePeriodo.EVALUACION.value:
        messages.error(request, "Solo se pueden editar calificaciones cuando el Periodo está en evaluación")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    if request.method == "POST":
        try:
            p1 = float(request.POST.get("parcial_1", 0))
            p2 = float(request.POST.get("parcial_2", 0))
            asist = float(request.POST.get("asistencia", 0))
        except (ValueError, TypeError):
            messages.error(request, "Los valores especificados no son válidos")
            return redirect("editar_evaluacion", evaluacion_id=evaluacion.id)

        if not (0 <= p1 <= 10) or not (0 <= p2 <= 10):
            messages.error(request, "Las calificaciones deben estar entre 0 y 10")
            return redirect("editar_evaluacion", evaluacion_id=evaluacion.id)
        if not (0 <= asist <= 100):
            messages.error(request, "La asistencia debe estar entre 0 y 100")
            return redirect("editar_evaluacion", evaluacion_id=evaluacion.id)

        unidad = evaluacion.unidad_curricular
        nota_final = round((p1 + p2) / 2, 2)
        
        if asist < unidad.porcentaje_minimo_asistencia:
            estado = "Reprobado"
        elif nota_final >= unidad.criterio_de_aprobacion:
            estado = "Aprobado"
        else:
            estado = "Reprobado"

        evaluacion.calificacion_parcial_1 = p1
        evaluacion.calificacion_parcial_2 = p2
        evaluacion.nota_final = nota_final
        evaluacion.porcentaje_asistencia = asist
        evaluacion.estado_de_aprobacion = estado
        evaluacion.save()

        messages.success(request, "La calificación ha sido actualizada correctamente")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    return render(request, "academico/editar_evaluacion.html", {
        "evaluacion": evaluacion,
        "paralelo": paralelo,
        "titulo_pagina": "Calificación - NIVEC",
        "titulo": f"Modificar calificación - {evaluacion.estudiante.usuario_de_sistema.apellidos} {evaluacion.estudiante.usuario_de_sistema.nombres}",
    })



@requiere_perfil(*ROLES_CARGAR)
def pasar_a_revision(request, paralelo_id):
    from academico.services import servicio_pasar_a_revision
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    periodo = paralelo.periodo_de_nivelacion

    if periodo.estado != EstadoDePeriodo.EVALUACION.value:
        messages.error(request, "Solo se puede pasar a revisión cuando el Periodo está en evaluación")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    if request.method != "POST":
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    # Check there are evaluations in Borrador
    from academico.models import EvaluacionAcademica as EA
    borradores = EA.objects.filter(
        unidad_curricular=paralelo.unidad_curricular,
        periodo_de_nivelacion=periodo,
        estado_revision="Borrador",
        estudiante__estudiantes_matriculados__paralelo=paralelo,
    ).count()

    if borradores == 0:
        messages.warning(request, "No existen calificaciones en borrador para pasar a revisión")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    count = servicio_pasar_a_revision(paralelo)
    messages.success(request, f"{count} calificaciones pasadas a revisión correctamente")
    return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)


@requiere_perfil(*ROLES_CARGAR)
def formalizar_evaluaciones(request, paralelo_id):
    from academico.services import servicio_formalizar_evaluaciones
    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    periodo = paralelo.periodo_de_nivelacion

    if periodo.estado != EstadoDePeriodo.EVALUACION.value:
        messages.error(request, "Solo se puede formalizar cuando el Periodo está en evaluación")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    if request.method != "POST":
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    from academico.models import EvaluacionAcademica as EA
    en_revision = EA.objects.filter(
        unidad_curricular=paralelo.unidad_curricular,
        periodo_de_nivelacion=periodo,
        estado_revision="En revisión",
        estudiante__estudiantes_matriculados__paralelo=paralelo,
    ).count()

    if en_revision == 0:
        messages.warning(request, "No existen calificaciones en revisión para formalizar")
        return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)

    count = servicio_formalizar_evaluaciones(paralelo)
    messages.success(request, f"{count} calificaciones formalizadas correctamente")
    return redirect("listar_evaluaciones_paralelo", paralelo_id=paralelo.id)



@requiere_perfil(*ROLES_VER)
def descargar_calificaciones_paralelo(request, paralelo_id):
    import openpyxl
    from django.http import HttpResponse
    from academico.models import Horario

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        return redirect("panel_principal")

    paralelo = get_object_or_404(
        Paralelo, id=paralelo_id, periodo_de_nivelacion__universidad=universidad_usuario
    )
    carrera = paralelo.unidad_curricular.malla_curricular.carrera
    periodo = paralelo.periodo_de_nivelacion

    # Get all paralelos of this logical group
    unidades_rows = Paralelo.objects.filter(
        periodo_de_nivelacion=periodo,
        jornada=paralelo.jornada,
        nombre=paralelo.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    ).select_related("unidad_curricular").order_by("unidad_curricular__nombre")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for row in unidades_rows:
        titulo_hoja = row.unidad_curricular.nombre[:31]
        ws = wb.create_sheet(title=titulo_hoja)

        ws.append([
            "Tipo de identificación",
            "Número de identificación",
            "Apellidos",
            "Nombres",
            "Correo institucional",
            "Número de matrícula",
            "Calificación parcial 1",
            "Calificación parcial 2",
            "Nota final",
            "Porcentaje de asistencia",
            "Estado de aprobación",
            "Estado de revisión",
        ])

        evaluaciones = EvaluacionAcademica.objects.filter(
            unidad_curricular=row.unidad_curricular,
            periodo_de_nivelacion=periodo,
            estudiante__estudiantes_matriculados__paralelo=row,
        ).select_related("estudiante__usuario_de_sistema").order_by(
            "estudiante__usuario_de_sistema__apellidos"
        )

        for ev in evaluaciones:
            est = ev.estudiante
            usr = est.usuario_de_sistema
            ws.append([
                usr.tipo_de_identificacion,
                usr.identificacion,
                usr.apellidos,
                usr.nombres,
                usr.correo_institucional,
                est.numero_de_matricula,
                ev.calificacion_parcial_1,
                ev.calificacion_parcial_2,
                ev.nota_final,
                ev.porcentaje_asistencia,
                ev.estado_de_aprobacion,
                ev.estado_revision,
            ])

        for col in range(1, 13):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sin datos")
        ws.append(["No existen calificaciones registradas"])

    nombre_archivo = f"calificaciones_{paralelo.nombre}_{carrera.nombre}".lower().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response


@requiere_perfil(ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR)
def informe_general(request):
    from academico.services import servicio_generar_informe_general

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    from academico.models import PeriodoDeNivelacion
    periodos = PeriodoDeNivelacion.objects.filter(universidad=universidad_usuario).order_by("-anio", "-numero_periodo")

    periodo_id = request.GET.get("periodo")
    periodo_seleccionado = None
    informe = None

    if periodo_id:
        periodo_seleccionado = periodos.filter(id=periodo_id).first()
        if periodo_seleccionado:
            informe = servicio_generar_informe_general(periodo_seleccionado)

    return render(request, "academico/informe_general.html", {
        "periodos": periodos,
        "periodo_seleccionado": periodo_seleccionado,
        "informe": informe,
        "titulo_pagina": "Informe General - NIVEC",
        "titulo": "Informe General de Nivelación",
    })


@requiere_perfil(ROL_COORDINADOR_DAN, ROL_DIRECTOR_DAN, ROL_RECTOR, ROL_VICERRECTOR)
def descargar_informe_general(request):
    from academico.services import servicio_generar_informe_general
    import openpyxl
    from django.http import HttpResponse

    universidad_usuario = _obtener_universidad_usuario(request.user)
    if not universidad_usuario:
        messages.warning(request, "La Universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    periodo_id = request.GET.get("periodo")
    if not periodo_id:
        messages.error(request, "Debe especificar un Periodo")
        return redirect("informe_general")

    from academico.models import PeriodoDeNivelacion
    periodo = PeriodoDeNivelacion.objects.filter(id=periodo_id, universidad=universidad_usuario).first()
    if not periodo:
        messages.error(request, "Periodo no válido")
        return redirect("informe_general")

    informe = servicio_generar_informe_general(periodo)
    if not informe:
        messages.warning(request, "No existen datos para generar el informe (el periodo debe estar en evaluación o cerrado)")
        return redirect("informe_general")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe General"

    ws.append(["Informe General de Nivelación"])
    ws.append(["Periodo", periodo.periodo])
    ws.append(["Año", periodo.anio])
    ws.append(["Estado", periodo.estado])
    ws.append([])
    ws.append(["Carrera", "Total evaluaciones", "Aprobados", "Reprobados", "Retirados", "Anulados", "Pendientes", "Formalizados", "% Aprobación"])

    for fila in informe:
        ws.append([
            fila["carrera"], fila["total"], fila["aprobados"], fila["reprobados"],
            fila["retirados"], fila["anulados"], fila["pendientes"], fila["formalizados"],
            fila["porcentaje_aprobacion"],
        ])

    # Totals row
    ws.append([])
    total_eval = sum(f["total"] for f in informe)
    total_apr = sum(f["aprobados"] for f in informe)
    total_rep = sum(f["reprobados"] for f in informe)
    total_ret = sum(f["retirados"] for f in informe)
    total_anu = sum(f["anulados"] for f in informe)
    total_pen = sum(f["pendientes"] for f in informe)
    total_for = sum(f["formalizados"] for f in informe)
    pct = round(total_apr / total_eval * 100, 1) if total_eval > 0 else 0
    ws.append(["TOTAL", total_eval, total_apr, total_rep, total_ret, total_anu, total_pen, total_for, pct])

    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    nombre = f"informe_general_{periodo.periodo}".lower().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre}.xlsx"'
    wb.save(response)
    return response



@login_required
def descargar_constancia_estudiante(request):
    import openpyxl
    from django.http import HttpResponse
    from usuarios.models import PerfilEstudiante
    from academico.models import EvaluacionAcademica, MatriculaParalelo

    usuario = request.user
    perfil_estudiante = PerfilEstudiante.objects.filter(usuario_de_sistema=usuario).first()
    if not perfil_estudiante:
        messages.error(request, "No se encontró el perfil de estudiante")
        return redirect("panel_principal")

    periodo = perfil_estudiante.periodo_de_nivelacion
    if not periodo:
        messages.warning(request, "No tiene un periodo de nivelación asignado")
        return redirect("panel_principal")

    evaluaciones = EvaluacionAcademica.objects.filter(
        estudiante=perfil_estudiante,
        periodo_de_nivelacion=periodo,
    ).select_related("unidad_curricular").order_by("unidad_curricular__nombre")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Constancia"

    # Header
    ws.append(["Constancia de resultados académicos"])
    ws.append([])
    ws.append(["Estudiante", f"{usuario.apellidos} {usuario.nombres}"])
    ws.append(["Identificación", usuario.identificacion])
    ws.append(["Correo institucional", usuario.correo_institucional])
    ws.append(["Número de matrícula", perfil_estudiante.numero_de_matricula])
    ws.append(["Carrera", perfil_estudiante.carrera_registrada.nombre if perfil_estudiante.carrera_registrada else ""])
    ws.append(["Jornada", perfil_estudiante.jornada])
    ws.append(["Periodo de nivelación", periodo.periodo])
    ws.append(["Año", periodo.anio])
    ws.append([])

    # Results table
    ws.append(["Unidad curricular", "Parcial 1", "Parcial 2", "Nota final", "% Asistencia", "Estado de aprobación"])
    for ev in evaluaciones:
        ws.append([
            ev.unidad_curricular.nombre,
            ev.calificacion_parcial_1,
            ev.calificacion_parcial_2,
            ev.nota_final,
            ev.porcentaje_asistencia,
            ev.estado_de_aprobacion,
        ])

    ws.append([])
    # Overall status
    if evaluaciones.exists():
        todos_aprobados = all(ev.estado_de_aprobacion == "Aprobado" for ev in evaluaciones)
        ws.append(["Resultado general", "APROBADO" if todos_aprobados else "NO APROBADO"])

    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    nombre = f"constancia_{usuario.identificacion}_{periodo.periodo}".lower().replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre}.xlsx"'
    wb.save(response)
    return response
