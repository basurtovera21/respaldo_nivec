import openpyxl
import unicodedata
from datetime import date, datetime
from django.db import transaction
from django.http import HttpResponse

from academico.models import Campus, Carrera, PeriodoDeNivelacion, Paralelo, EvaluacionAcademica, CohorteDeMatricula, InformeGeneral, ConsolidadoAcademico, MatriculaParalelo, Horario, MallaCurricular
from usuarios.models import PerfilEstudiante
from usuarios.utils import generar_identificador_siguiente

# Enums
from poo.clases.enums.estado_de_matricula import EstadoDeMatricula
from poo.clases.enums.estado_de_aprobacion import EstadoDeAprobacion
from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
from poo.clases.enums.formato_de_exportacion import FormatoDeExportacion
from poo.clases.enums.modalidad import Modalidad
from poo.clases.enums.jornada import Jornada
from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte
from poo.clases.enums.dia_de_semana import DiaDeSemana
# POO
from poo.clases.carrera import Carrera as CarreraBase
from poo.clases.periodo_de_nivelacion import PeriodoDeNivelacion as PeriodoDeNivelacionBase
from poo.clases.evaluacion_academica import EvaluacionAcademica as EvaluacionAcademicaPOO # Evitar choque de nombres
from poo.clases.informe_general import InformeGeneral as InformeGeneralPOO
from poo.clases.servicios.procesador_de_informe import ProcesadorDeInforme
from poo.clases.cohorte_de_matricula import CohorteDeMatricula as CohorteDeMatriculaPOO
from poo.clases.horario import Horario as HorarioPOO
from poo.clases.consolidado_academico import ConsolidadoAcademico as ConsolidadoAcademicoPOO

from academico.models import (
    Campus, Carrera, PeriodoDeNivelacion, Paralelo, EvaluacionAcademica,
    CohorteDeMatricula, InformeGeneral, ConsolidadoAcademico, MatriculaParalelo,
    Horario, MallaCurricular, UnidadCurricular )

from poo.clases.enums.estado_de_malla import EstadoDeMalla


def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto

def obtener_enum_flexible(enum_class, valor_sucio):
    if not valor_sucio:
        return None
    valor_normalizado = normalizar_texto(valor_sucio)
    for opcion in enum_class:
        if normalizar_texto(opcion.value) == valor_normalizado:
            return opcion
    raise ValueError(f"'{valor_sucio}' registro no válido para {enum_class.__name__}")

#Campus
def servicio_campus_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.campus import Campus as CampusBase

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        nombres_registrados = {
            CampusBase.normalizar_nombre(nombre_existente)
            for nombre_existente in Campus.objects.filter(
                universidad=universidad_usuario
            ).values_list("nombre", flat=True)
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nombre, direccion, provincia = fila[0], fila[1], fila[2]
                if not nombre and not direccion and not provincia: continue
                if not nombre or not direccion or not provincia:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información")
                    continue

                nombre_normalizado = CampusBase.normalizar_nombre(nombre)
                if nombre_normalizado in nombres_registrados:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (el Campus ya existe)"
                    )
                    continue

                with transaction.atomic():
                    Campus.objects.create(
                        universidad=universidad_usuario,
                        codigo_de_campus=generar_identificador_siguiente(Campus, 'CAM', 'codigo_de_campus'),
                        nombre=nombre,
                        direccion_fisica=direccion,
                        provincia=provincia
                    )
                    resultado["exitosos"] += 1
                    nombres_registrados.add(nombre_normalizado)
            except Exception as e:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido({str(e)})")
                
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
        
    return resultado


#Carrera
def servicio_carrera_registrar_masivo_desde_excel(archivo, universidad_usuario):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        campus_existente = Campus.objects.filter(universidad=universidad_usuario)

        carreras_registradas = {
            (campus_id, CarreraBase.normalizar_nombre(nombre_existente))
            for campus_id, nombre_existente in Carrera.objects.filter(
                campus__universidad=universidad_usuario
            ).values_list("campus_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo_campus, nombre, modalidad, vigencia = fila[:4]
                
                if not codigo_campus and not nombre and not modalidad and not vigencia:
                    continue
                
                if not codigo_campus or not nombre or not modalidad or not vigencia:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información")
                    continue

                try:
                    enum_modalidad = obtener_enum_flexible(Modalidad, modalidad)
                except ValueError:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Modalidad no válida)")
                    continue

                if isinstance(vigencia, (datetime, date)):
                    vigencia_date = vigencia.date() if isinstance(vigencia, datetime) else vigencia
                elif isinstance(vigencia, str):
                    try:
                        vigencia_date = datetime.strptime(vigencia.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                        continue
                else:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                    continue

                campus_obj = campus_existente.filter(codigo_de_campus=codigo_campus).first()
                if not campus_obj:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (código de Campus no válido)")
                    continue
                
                carrera_poo = CarreraBase(
                    codigo_de_carrera="PENDIENTE",
                    nombre=nombre,
                    modalidad=enum_modalidad,
                    vigencia_sniese=vigencia_date
                )

                if not carrera_poo.esta_activa():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Carrera no vigente)")
                    continue

                clave_carrera = (campus_obj.id, CarreraBase.normalizar_nombre(nombre))
                if clave_carrera in carreras_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Carrera ya existe)"
                    )
                    continue

                with transaction.atomic():
                    Carrera.objects.create(
                        campus=campus_obj,
                        codigo_de_carrera=generar_identificador_siguiente(Carrera, 'CAR', 'codigo_de_carrera'),
                        nombre=nombre,
                        modalidad=enum_modalidad.value, 
                        vigencia_sniese=vigencia_date
                    )
                    resultado["exitosos"] += 1
                    carreras_registradas.add(clave_carrera)
            except Exception as e:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})")
                
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
        
    return resultado



#PeriodoDeNivelacion 
def _construir_periodo(periodo_db):
    return PeriodoDeNivelacionBase(
        codigo_periodo=periodo_db.codigo_periodo,
        anio=periodo_db.anio,
        periodo=periodo_db.periodo,
        fecha_inicio=periodo_db.fecha_inicio,
        fecha_fin=periodo_db.fecha_fin,
        modalidad=obtener_enum_flexible(Modalidad, periodo_db.modalidad),
        numero_periodo=periodo_db.numero_periodo,
        estado=obtener_enum_flexible(EstadoDePeriodo, periodo_db.estado)
    )

def servicio_iniciar_periodo_de_nivelacion(periodo_db):
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica
    periodo_poo = _construir_periodo(periodo_db)
    facade = CentroDeOperacionAcademica()
    if facade.iniciar_periodo(periodo_poo):
        periodo_db.estado = periodo_poo.estado.value
        periodo_db.save()
        return True
    return False

def servicio_finalizar_periodo_de_nivelacion(periodo_db):
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica
    periodo_poo = _construir_periodo(periodo_db)
    facade = CentroDeOperacionAcademica()
    if facade.finalizar_periodo(periodo_poo):
        periodo_db.estado = periodo_poo.estado.value
        periodo_db.save()
        return True
    return False


#Malla curricular
def servicio_malla_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.malla_curricular import MallaCurricular as MallaCurricularBase
    from poo.clases.enums.estado_de_malla import EstadoDeMalla

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        carreras_existentes = Carrera.objects.filter(campus__universidad=universidad_usuario)

        mallas_registradas = {
            (carrera_id, str(nombre_existente).strip().lower())
            for carrera_id, nombre_existente in MallaCurricular.objects.filter(
                carrera__campus__universidad=universidad_usuario
            ).values_list("carrera_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo_carrera, nombre = fila[:2]

                if not any([codigo_carrera, nombre]):
                    continue

                if not all([codigo_carrera, nombre]):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido por falta de información"
                    )
                    continue

                carrera_obj = carreras_existentes.filter(
                    codigo_de_carrera=str(codigo_carrera).strip()
                ).first()
                if not carrera_obj:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (Código de Carrera no válido)"
                    )
                    continue

                malla_poo = MallaCurricularBase(
                    codigo_de_malla="PENDIENTE",
                    nombre=str(nombre).strip(),
                    version_de_malla="PENDIENTE",
                )

                errores_poo = malla_poo.validar_datos_de_registro()
                if errores_poo:
                    primer_error = list(errores_poo.values())[0]
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido ({primer_error})"
                    )
                    continue

                clave_malla = (carrera_obj.id, malla_poo.nombre.strip().lower())
                if clave_malla in mallas_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (La Malla curricular ya existe)"
                    )
                    continue

                with transaction.atomic():
                    MallaCurricular.objects.create(
                        carrera=carrera_obj,
                        codigo_de_malla=generar_identificador_siguiente(
                            MallaCurricular, "MC", "codigo_de_malla"
                        ),
                        nombre=malla_poo.nombre,
                        version_de_malla=servicio_generar_version_malla(carrera_obj),
                        estado=EstadoDeMalla.DISENO.value,
                    )
                    resultado["exitosos"] += 1
                    mallas_registradas.add(clave_malla)

            except Exception as e:
                resultado["advertencias"].append(
                    f"El registro de la fila {numero_fila} fue omitido ({str(e)})"
                )

    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"

    return resultado



def servicio_unidad_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.unidad_curricular import UnidadCurricular as UnidadCurricularBase
    from poo.clases.enums.estado_de_malla import EstadoDeMalla
    from academico.models import UnidadCurricular, MallaCurricular
    from usuarios.utils import generar_identificador_siguiente

    estados_editables = (EstadoDeMalla.DISENO.value, EstadoDeMalla.ACTIVA.value)

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        mallas_existentes = MallaCurricular.objects.filter(
            carrera__campus__universidad=universidad_usuario
        )

        unidades_registradas = {
            (malla_id, str(nombre_u).strip().lower())
            for malla_id, nombre_u in UnidadCurricular.objects.filter(
                malla_curricular__carrera__campus__universidad=universidad_usuario
            ).values_list("malla_curricular_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (codigo_malla, nombre, horas_totales,
                 horas_sincronicas, horas_sincronicas_semanales, horas_asincronicas,
                 criterio, porcentaje_asistencia) = fila[:8]

                if not any([codigo_malla, nombre, horas_totales,
                            horas_sincronicas, horas_asincronicas]):
                    continue

                if not all([codigo_malla, nombre, horas_totales,
                            horas_sincronicas, horas_sincronicas_semanales, horas_asincronicas]):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido por falta de información"
                    )
                    continue

                try:
                    horas_totales_f = float(horas_totales)
                    horas_sincronicas_f = float(horas_sincronicas)
                    horas_sincronicas_semanales_f = float(horas_sincronicas_semanales)
                    horas_asincronicas_f = float(horas_asincronicas)
                    criterio_f = float(criterio) if criterio is not None else 7.0
                    porcentaje_f = float(porcentaje_asistencia) if porcentaje_asistencia is not None else 70.0
                except (ValueError, TypeError):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (registro(s) no válido(s))"
                    )
                    continue

                if (horas_sincronicas_semanales_f <= 0
                        or horas_sincronicas_semanales_f != int(horas_sincronicas_semanales_f)
                        or horas_sincronicas_semanales_f > horas_sincronicas_f):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (Horas sincrónicas semanales no válidas: entero, > 0 y <= sincrónicas totales)"
                    )
                    continue

                malla_obj = mallas_existentes.filter(
                    codigo_de_malla=str(codigo_malla).strip()
                ).first()
                if not malla_obj:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (Código de Malla no válido)"
                    )
                    continue

                if malla_obj.estado not in estados_editables:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (Estado no válido)"
                    )
                    continue

                clave_unidad = (malla_obj.id, str(nombre).strip().lower())
                if clave_unidad in unidades_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Unidad curricular ya existe en la Malla especificada)"
                    )
                    continue

                unidad_poo = UnidadCurricularBase(
                    codigo_de_unidad="PENDIENTE",
                    nombre=str(nombre).strip(),
                    horas_totales=horas_totales_f,
                    horas_sincronicas=horas_sincronicas_f,
                    horas_asincronicas=horas_asincronicas_f,
                    criterio_de_aprobacion=criterio_f,
                    porcentaje_minimo_asistencia=porcentaje_f,
                )

                errores_poo = unidad_poo.validar_datos_de_registro()
                if errores_poo:
                    primer_error = list(errores_poo.values())[0]
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido ({primer_error})"
                    )
                    continue

                with transaction.atomic():
                    UnidadCurricular.objects.create(
                        malla_curricular=malla_obj,
                        codigo_de_unidad=generar_identificador_siguiente(
                            UnidadCurricular, "UC", "codigo_de_unidad"
                        ),
                        nombre=unidad_poo.nombre,
                        horas_totales=unidad_poo.horas_totales,
                        horas_sincronicas=unidad_poo.horas_sincronicas,
                        horas_sincronicas_semanales=horas_sincronicas_semanales_f,
                        horas_asincronicas=unidad_poo.horas_asincronicas,
                        criterio_de_aprobacion=unidad_poo.criterio_de_aprobacion,
                        porcentaje_minimo_asistencia=unidad_poo.porcentaje_minimo_asistencia,
                    )
                    resultado["exitosos"] += 1
                    unidades_registradas.add(clave_unidad)

            except Exception as e:
                resultado["advertencias"].append(
                    f"El registro de la fila {numero_fila} fue omitido ({str(e)})"
                )

    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"

    return resultado





def servicio_registrar_evaluacion_academica(evaluacion_academica: EvaluacionAcademica):
    from poo.clases.unidad_curricular import UnidadCurricular as UnidadCurricularBase
    from usuarios.services import _construir_estudiante

    unidad_curricular_base = UnidadCurricularBase(
        codigo_de_unidad = evaluacion_academica.unidad_curricular.codigo_de_unidad,
        nombre = evaluacion_academica.unidad_curricular.nombre,
        horas_totales = evaluacion_academica.unidad_curricular.horas_totales,
        horas_sincronicas = evaluacion_academica.unidad_curricular.horas_sincronicas,
        horas_asincronicas = evaluacion_academica.unidad_curricular.horas_asincronicas,
        criterio_de_aprobacion = evaluacion_academica.unidad_curricular.criterio_de_aprobacion,
        porcentaje_minimo_asistencia = evaluacion_academica.unidad_curricular.porcentaje_minimo_asistencia
    )
    estudiante = _construir_estudiante(evaluacion_academica.estudiante)
    evaluacion = EvaluacionAcademicaPOO(estudiante = estudiante, unidad_curricular = unidad_curricular_base)
    
    evaluacion.registrar_calificacion(1, evaluacion_academica.calificacion_parcial_1)
    evaluacion.registrar_calificacion(2, evaluacion_academica.calificacion_parcial_2)
    evaluacion.registrar_asistencia_final(evaluacion_academica.porcentaje_asistencia)
    nota_final = evaluacion.calcular_nota_final()
    estado = evaluacion.verificar_aprobacion()

    evaluacion_academica.nota_final = nota_final
    evaluacion_academica.estado_de_aprobacion = estado.value
    evaluacion_academica.observacion = evaluacion._observacion
    evaluacion_academica.save()


def servicio_procesar_mtn(archivo, periodo_de_nivelacion: PeriodoDeNivelacion):
    from usuarios.services import servicio_estudiante_registrar_masivo_desde_excel
    from poo.clases.consolidado_academico import ConsolidadoAcademico as ConsolidadoAcademicoPOO

    resultado = servicio_estudiante_registrar_masivo_desde_excel(
        archivo, periodo_de_nivelacion.universidad, periodo_de_nivelacion=periodo_de_nivelacion
    )


    if resultado.get("error"):
        return resultado

    # Conteo dinámico: el consolidado refleja el estado real del periodo (cuenta de
    # estudiantes anclados), no acumula por carga. Así, re-cargar la misma MTN no
    # infla ni descuadra los totales.
    identificaciones = list(
        PerfilEstudiante.objects.filter(
            periodo_de_nivelacion=periodo_de_nivelacion
        ).values_list("usuario_de_sistema__identificacion", flat=True)
    )
    participantes = len(identificaciones)
    observados = resultado["observados"]

    matriz_procesada = (
        [{"identificacion": ident} for ident in identificaciones]
        + [{} for _ in range(observados)]
    )

    consolidado_poo = ConsolidadoAcademicoPOO(
        periodo_academico=None,
        fecha_de_corte=date.today(),
        total_de_cupos_aceptados=participantes,
    )
    consolidado_poo.cargar_matriz_de_cupos(matriz_procesada, participantes, observados)
    estadisticas = consolidado_poo.obtener_estadisticas_de_consolidado()

    consolidado_db, _ = ConsolidadoAcademico.objects.get_or_create(
        periodo_academico=periodo_de_nivelacion,
        defaults={"fecha_de_corte": date.today()},
    )
    consolidado_db.registros_validos = estadisticas["Registros válidos"]
    consolidado_db.total_cupos_aceptados = estadisticas["Cupos aceptados esperados"]
    consolidado_db.registros_observados = observados
    consolidado_db.registros_totales = participantes + observados
    consolidado_db.fecha_de_corte = date.today()
    consolidado_db.save()

    return resultado




def servicio_exportar_informe(informe_general: InformeGeneral, formato: str):
    evaluaciones = EvaluacionAcademica.objects.filter(
        unidad_curricular__malla_curricular__carrera__campus__universidad=informe_general.periodo_academico.universidad
    ).select_related("estudiante", "unidad_curricular")

    if formato == "excel":
        return _exportar_excel(informe_general, evaluaciones)
    return _exportar_txt(informe_general, evaluaciones)


def _exportar_excel(informe_django, evaluaciones):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = f"Informe {informe_django.periodo_academico.periodo}"
    hoja.append(["Identificación", "Nombres", "Apellidos", "Carrera registrada", "Campus registrado",
                 "Jornada", "Unidad curricular", "Calificación primer parcial", "Calificación segundo parcial",
                 "Calificacion final", "Porcentaje de asistencia final", "Estado"])
    for evaluacion in evaluaciones:
        evaluacion_estudiante = evaluacion.estudiante
        hoja.append([
            evaluacion_estudiante.usuario_de_sistema.identificacion,
            evaluacion_estudiante.usuario_de_sistema.nombres,
            evaluacion_estudiante.usuario_de_sistema.apellidos,
            evaluacion_estudiante.carrera_registrada.nombre,
            evaluacion_estudiante.campus_registrado.nombre,
            evaluacion_estudiante.jornada,
            evaluacion.unidad_curricular.nombre,
            evaluacion.calificacion_parcial_1,
            evaluacion.calificacion_parcial_2,
            evaluacion.nota_final,
            evaluacion.porcentaje_asistencia,
            evaluacion.estado_de_aprobacion,
        ])
    documento = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    documento["Content-Disposition"] = f'attachment; filename="informe_{informe_django.periodo_academico.periodo}.xlsx"'
    libro.save(documento)
    return documento


def _exportar_txt(informe_django, evaluaciones):
    lineas = [
        "INFORME GENERAL DE PERIODO DE NIVELACIÓN",
        f"Periodo de nivelación: {informe_django.periodo_academico.periodo}",
        f"Universidad: {informe_django.periodo_academico.universidad.nombre}",
        f"Fecha de emisión: {date.today()}",
        "─" * 60,
    ]
    for evaluacion in evaluaciones:
        evaluacion_estudiante = evaluacion.estudiante
        lineas.append(
            f"{evaluacion_estudiante.usuario_de_sistema.identificacion:<15} "
            f"{evaluacion_estudiante.usuario_de_sistema.nombres:<25} "
            f"{evaluacion.unidad_curricular.nombre:<30} "
            f"{evaluacion.nota_final:>6.2f} "
            f"{evaluacion.estado_de_aprobacion:<12}"
        )
    documento = HttpResponse("\n".join(lineas), content_type="text/plain; charset=utf-8")
    documento["Content-Disposition"] = f'attachment; filename="informe_{informe_django.periodo_academico.periodo}.txt"'
    return documento


#MallaCurricular
def _construir_unidad_poo(unidad_db):
    from poo.clases.unidad_curricular import UnidadCurricular as UnidadCurricularBase

    return UnidadCurricularBase(
        codigo_de_unidad=unidad_db.codigo_de_unidad,
        nombre=unidad_db.nombre,
        horas_totales=unidad_db.horas_totales,
        horas_sincronicas=unidad_db.horas_sincronicas,
        horas_asincronicas=unidad_db.horas_asincronicas,
        criterio_de_aprobacion=unidad_db.criterio_de_aprobacion,
        porcentaje_minimo_asistencia=unidad_db.porcentaje_minimo_asistencia,
    )
    
def servicio_generar_version_malla(carrera):
    import re
    maximo = 0
    for version in MallaCurricular.objects.filter(carrera=carrera).values_list("version_de_malla", flat=True):
        coincidencia = re.search(r"\d+", str(version or ""))
        if coincidencia:
            maximo = max(maximo, int(coincidencia.group()))
    return f"V{maximo + 1}"
    
def _construir_malla_poo(malla_db, cargar_unidades=False):
    from poo.clases.malla_curricular import MallaCurricular as MallaCurricularBase

    malla_poo = MallaCurricularBase(
        codigo_de_malla=malla_db.codigo_de_malla,
        nombre=malla_db.nombre,
        version_de_malla=malla_db.version_de_malla,
    )

    if cargar_unidades:
        for unidad_db in malla_db.unidades_curriculares.all():
            malla_poo.agregar_unidad_curricular(_construir_unidad_poo(unidad_db))

    malla_poo.establecer_estado(obtener_enum_flexible(EstadoDeMalla, malla_db.estado))
    return malla_poo

def servicio_recalcular_total_horas_malla(malla_db):
    malla_poo = _construir_malla_poo(malla_db, cargar_unidades=False)
    malla_poo.establecer_estado(EstadoDeMalla.DISENO)

    for unidad_db in malla_db.unidades_curriculares.all():
        malla_poo.agregar_unidad_curricular(_construir_unidad_poo(unidad_db))

    malla_db.total_horas_nivelacion = malla_poo.calcular_total_horas_nivelacion()
    malla_db.save(update_fields=["total_horas_nivelacion"])
    return malla_db.total_horas_nivelacion

def servicio_cambiar_estado_malla(malla_db, accion):
    malla_poo = _construir_malla_poo(malla_db)

    if accion == "activar":
        otra_activa = MallaCurricular.objects.filter(
            carrera=malla_db.carrera,
            estado=EstadoDeMalla.ACTIVA.value,
        ).exclude(pk=malla_db.pk).first()

        if otra_activa:
            return (
                False,
                f"La Carrera tiene una Malla curricular activa ({otra_activa.codigo_de_malla} - {otra_activa.version_de_malla})"
            )

        if not malla_poo.activar():
            return (False, "La Malla curricular no se ha podido habilitar")

    elif accion == "historica":
        if not malla_poo.marcar_historica():
            return (False, "La Malla curricular no se ha podido descontinuar")

    elif accion == "inactivar":
        if not malla_poo.inactivar():
            return (False, "La Malla curricular no se ha podido deshabilitar")

    else:
        return (False, "No válido")

    malla_db.estado = malla_poo.estado.value
    malla_db.save(update_fields=["estado"])
    return (True, "El estado de la Malla curricular ha sido actualizado correctamente")

def servicio_clonar_malla_curricular(id_malla_curricular_bd, nuevo_nombre=None):
    from academico.models import UnidadCurricular

    malla_curricular_db = MallaCurricular.objects.get(id=id_malla_curricular_bd)

    malla_poo = _construir_malla_poo(malla_curricular_db, cargar_unidades=True)

    nuevo_codigo = generar_identificador_siguiente(MallaCurricular, "MC", "codigo_de_malla")
    nueva_version = servicio_generar_version_malla(malla_curricular_db.carrera)
    clon_poo = malla_poo.clonar(nuevo_codigo, nueva_version)

    if nuevo_nombre and str(nuevo_nombre).strip():
        clon_poo.nombre = str(nuevo_nombre).strip()

    with transaction.atomic():
        nueva_malla_curricular_db = MallaCurricular.objects.create(
            carrera=malla_curricular_db.carrera,
            codigo_de_malla=clon_poo.codigo_de_malla,
            nombre=clon_poo.nombre,
            version_de_malla=clon_poo.version_de_malla,
            estado=clon_poo.estado.value,  # Diseño
            total_horas_nivelacion=clon_poo.total_horas_nivelacion,
        )

        for unidad_poo in clon_poo.obtener_unidades_curriculares():
            UnidadCurricular.objects.create(
                malla_curricular=nueva_malla_curricular_db,
                codigo_de_unidad=generar_identificador_siguiente(
                    UnidadCurricular, "UC", "codigo_de_unidad"
                ),
                nombre=unidad_poo.nombre,
                horas_totales=unidad_poo.horas_totales,
                horas_sincronicas=unidad_poo.horas_sincronicas,
                horas_asincronicas=unidad_poo.horas_asincronicas,
                criterio_de_aprobacion=unidad_poo.criterio_de_aprobacion,
                porcentaje_minimo_asistencia=unidad_poo.porcentaje_minimo_asistencia,
            )

    return nueva_malla_curricular_db


def _obtener_o_crear_cohorte(periodo_db, carrera):
    from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte

    cohorte = CohorteDeMatricula.objects.filter(
        periodo_de_nivelacion=periodo_db, carrera_registrada=carrera
    ).first()
    if cohorte:
        return cohorte

    return CohorteDeMatricula.objects.create(
        periodo_de_nivelacion=periodo_db,
        carrera_registrada=carrera,
        codigo_de_registro=generar_identificador_siguiente(CohorteDeMatricula, "COH", "codigo_de_registro"),
        nombre_cohorte=f"Cohorte {carrera.nombre} - {periodo_db.periodo}",
        fecha_de_cierre=periodo_db.fecha_fin,
        tipo_de_cohorte=TipoDeCohorte.PRIMERA_MATRICULA.value,
    )
def servicio_recalcular_cohorte_de_carrera(periodo_db, carrera):
    # Recalcula los totales de la cohorte (por tipo de cupo) desde los estudiantes
    # realmente matriculados. La lógica de conteo vive en la POO CohorteDeMatricula.
    from poo.clases.cohorte_de_matricula import CohorteDeMatricula as CohorteDeMatriculaPOO
    from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte

    cohorte = CohorteDeMatricula.objects.filter(
        periodo_de_nivelacion=periodo_db, carrera_registrada=carrera
    ).first()
    if not cohorte:
        return

    estudiantes = PerfilEstudiante.objects.filter(
        periodo_de_nivelacion=periodo_db,
        carrera_registrada=carrera,
        estudiantes_matriculados__paralelo__periodo_de_nivelacion=periodo_db,
    ).distinct()

    cohorte_poo = CohorteDeMatriculaPOO(
        codigo_de_registro=cohorte.codigo_de_registro,
        nombre_cohorte=cohorte.nombre_cohorte,
        carrera_registrada=None,
        fecha_de_cierre=periodo_db.fecha_fin,
        periodo_de_nivelacion=None,
        tipo_de_cohorte=obtener_enum_flexible(TipoDeCohorte, cohorte.tipo_de_cohorte),
    )
    for estudiante_db in estudiantes:
        cohorte_poo.registrar_estudiante_matriculado(estudiante_db)

    estadisticas = cohorte_poo.obtener_estadisticas_de_registro()
    cohorte.total_primera_matricula = estadisticas["Total primera matricula"]
    cohorte.total_segunda_matricula = estadisticas["Total segunda matricula"]
    cohorte.total_exonerados = estadisticas["Total exonerados"]
    cohorte.save()


def _nombre_paralelo_letra(indice):
    # 0->Paralelo A, 1->Paralelo B, ..., 25->Paralelo Z, 26->Paralelo A1...
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if indice < 26:
        return f"Paralelo {letras[indice]}"
    ciclo = (indice - 26) // 26 + 1
    pos = (indice - 26) % 26
    return f"Paralelo {letras[pos]}{ciclo}"


def servicio_generar_paralelos(periodo_db, capacidad=35):
    import math
    from poo.clases.paralelo import Paralelo as ParaleloBase
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica
    from poo.clases.cohorte_de_matricula import CohorteDeMatricula as CohorteDeMatriculaPOO
    from poo.clases.enums.jornada import Jornada
    from poo.clases.enums.modalidad import Modalidad as EnumModalidad
    from poo.clases.enums.estado_de_malla import EstadoDeMalla
    from poo.clases.enums.registro_de_cupo import RegistroDeCupo
    from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte

    resumen = {
        "grupos_creados": 0,
        "paralelos_creados": 0,
        "estudiantes_distribuidos": 0,
        "advertencias": [],
    }

    try:
        capacidad = int(capacidad)
    except (TypeError, ValueError):
        capacidad = 35
    if capacidad <= 0:
        capacidad = 35

    universidad = periodo_db.universidad
    facade = CentroDeOperacionAcademica()
    enum_modalidad = obtener_enum_flexible(EnumModalidad, periodo_db.modalidad)

    import re

    def _num_sufijo(cadena):
        coincidencia = re.search(r"(\d+)$", str(cadena or ""))
        return int(coincidencia.group(1)) if coincidencia else 0

    contador_codigo = max(
        [_num_sufijo(c) for c in Paralelo.objects.values_list("codigo_de_paralelo", flat=True).distinct()],
        default=0,
    )

    carreras = Carrera.objects.filter(campus__universidad=universidad)

    for carrera in carreras:
        malla = MallaCurricular.objects.filter(
            carrera=carrera, estado=EstadoDeMalla.ACTIVA.value
        ).first()
        if not malla:
            if PerfilEstudiante.objects.filter(
                carrera_registrada=carrera,
                periodo_de_nivelacion=periodo_db,
                estado_de_matricula=EstadoDeMatricula.MATRICULADO.value,
            ).exists():
                resumen["advertencias"].append(
                    f"La Carrera de {carrera.nombre} tiene estudiantes registrados pero no cuenta con una Malla curricular activa"
                )
            continue

        unidades = list(malla.unidades_curriculares.all())
        if not unidades:
            resumen["advertencias"].append(
                f"Los registros en la Carrera de {carrera.nombre} fueron omitidos (sin registros asociados)"
            )
            continue

        jornadas_presentes = (
            PerfilEstudiante.objects.filter(
                carrera_registrada=carrera, periodo_de_nivelacion=periodo_db
            )
            .values_list("jornada", flat=True).distinct()
        )


        for jornada_valor in jornadas_presentes:
            estudiantes = list(
                PerfilEstudiante.objects.filter(
                    carrera_registrada=carrera,
                    jornada=jornada_valor,
                    periodo_de_nivelacion=periodo_db,
                    estado_de_matricula=EstadoDeMatricula.MATRICULADO.value,
                ).exclude(
                    estudiantes_matriculados__paralelo__periodo_de_nivelacion=periodo_db
                ).distinct()
            )
            if not estudiantes:
                continue

            try:
                enum_jornada = obtener_enum_flexible(Jornada, jornada_valor)
            except ValueError:
                resumen["advertencias"].append(
                    f"El registro de Jornada fue omitido (registro no válido)"
                )
                continue

            cohorte = _obtener_o_crear_cohorte(periodo_db, carrera)
            estudiantes_a_contar = []

            def _numero_de_grupo(nombre):
                try:
                    return int(str(nombre).split()[-1])
                except (ValueError, IndexError):
                    return 0

            with transaction.atomic():
                paralelos_existentes = Paralelo.objects.filter(
                    periodo_de_nivelacion=periodo_db,
                    jornada=jornada_valor,
                    unidad_curricular__in=unidades,
                )
                grupos_existentes = {}
                for paralelo_db in paralelos_existentes:
                    grupos_existentes.setdefault(paralelo_db.nombre, []).append(paralelo_db)

                indice_max = 0
                indice_pendiente = 0

                for nombre_grupo in sorted(grupos_existentes.keys(), key=_numero_de_grupo):
                    paralelos_grupo = grupos_existentes[nombre_grupo]
                    indice_max = max(indice_max, _numero_de_grupo(nombre_grupo))

                    representativo = paralelos_grupo[0]
                    ocupacion = MatriculaParalelo.objects.filter(paralelo=representativo).count()
                    cupo_libre = representativo.capacidad_maxima - ocupacion
                    if cupo_libre <= 0:
                        continue

                    a_matricular = estudiantes[indice_pendiente:indice_pendiente + cupo_libre]
                    indice_pendiente += len(a_matricular)

                    for paralelo_db in paralelos_grupo:
                        for estudiante_db in a_matricular:
                            MatriculaParalelo.objects.create(
                                estudiante=estudiante_db,
                                paralelo=paralelo_db,
                                cohorte_de_matricula=cohorte,
                            )
                    estudiantes_a_contar.extend(a_matricular)

                estudiantes_restantes = estudiantes[indice_pendiente:]
                if estudiantes_restantes:
                    # El nombre reinicia por carrera: cada carrera empieza en "A".
                    nombres_existentes = list(
                        Paralelo.objects.filter(
                            periodo_de_nivelacion=periodo_db,
                            unidad_curricular__malla_curricular__carrera=carrera,
                        ).values_list("nombre", flat=True).distinct()
                    )
                    indice_base = len(set(nombres_existentes))
                    numero_de_grupos = math.ceil(len(estudiantes_restantes) / capacidad)
                    grupos_poo = [
                        ParaleloBase(
                            codigo_de_paralelo=f"G{indice}",
                            nombre=_nombre_paralelo_letra(indice_base + indice - 1),
                            jornada=enum_jornada,
                            modalidad=enum_modalidad,
                            capacidad_maxima=capacidad,
                        )
                        for indice in range(1, numero_de_grupos + 1)
                    ]

                    facade.distribuir_estudiantes(grupos_poo, estudiantes_restantes)

                    for indice, grupo_poo in enumerate(grupos_poo, start=1):
                        miembros = list(grupo_poo._estudiantes_matriculados)
                        if not miembros:
                            continue
                        nombre_nuevo = _nombre_paralelo_letra(indice_base + indice - 1)
                        # Un solo código para todo el paralelo lógico (todas sus unidades).
                        contador_codigo += 1
                        codigo_nuevo = f"PAR{contador_codigo:03d}"
                        for unidad in unidades:
                            paralelo_db = Paralelo.objects.create(
                                periodo_de_nivelacion=periodo_db,
                                unidad_curricular=unidad,
                                codigo_de_paralelo=codigo_nuevo,
                                nombre=nombre_nuevo,
                                jornada=jornada_valor,
                                modalidad=periodo_db.modalidad,
                                capacidad_maxima=capacidad,
                            )
                            resumen["paralelos_creados"] += 1
                            for estudiante_db in miembros:
                                MatriculaParalelo.objects.create(
                                    estudiante=estudiante_db,
                                    paralelo=paralelo_db,
                                    cohorte_de_matricula=cohorte,
                                )
                        resumen["grupos_creados"] += 1
                        estudiantes_a_contar.extend(miembros)

                # El conteo por tipo de cupo vive en la POO CohorteDeMatricula.
                cohorte_poo = CohorteDeMatriculaPOO(
                    codigo_de_registro=cohorte.codigo_de_registro,
                    nombre_cohorte=cohorte.nombre_cohorte,
                    carrera_registrada=None,
                    fecha_de_cierre=periodo_db.fecha_fin,
                    periodo_de_nivelacion=None,
                    tipo_de_cohorte=obtener_enum_flexible(TipoDeCohorte, cohorte.tipo_de_cohorte),
                )
                for estudiante_db in estudiantes_a_contar:
                    cohorte_poo.registrar_estudiante_matriculado(estudiante_db)

                estadisticas_cohorte = cohorte_poo.obtener_estadisticas_de_registro()
                cohorte.total_primera_matricula += estadisticas_cohorte["Total primera matricula"]
                cohorte.total_segunda_matricula += estadisticas_cohorte["Total segunda matricula"]
                cohorte.total_exonerados += estadisticas_cohorte["Total exonerados"]
                resumen["estudiantes_distribuidos"] += len(estudiantes_a_contar)

                cohorte.save()

    return resumen

def servicio_mover_estudiante(estudiante_db, paralelo_destino_db):
    periodo = paralelo_destino_db.periodo_de_nivelacion
    carrera = paralelo_destino_db.unidad_curricular.malla_curricular.carrera
    nombre_destino = paralelo_destino_db.nombre
    jornada = paralelo_destino_db.jornada

    paralelos_destino = list(Paralelo.objects.filter(
        periodo_de_nivelacion=periodo,
        jornada=jornada,
        nombre=nombre_destino,
        unidad_curricular__malla_curricular__carrera=carrera,
    ))
    if not paralelos_destino:
        return (False, "La especificación del Paralelo de destino no es válida")

    representativo = paralelos_destino[0]
    ocupacion_destino = MatriculaParalelo.objects.filter(
        paralelo=representativo
    ).exclude(estudiante=estudiante_db).count()
    if ocupacion_destino >= representativo.capacidad_maxima:
        return (False, "El Paralelo de destino no presenta cupos disponibles")

    matriculas_actuales = MatriculaParalelo.objects.filter(
        estudiante=estudiante_db,
        paralelo__periodo_de_nivelacion=periodo,
        paralelo__unidad_curricular__malla_curricular__carrera=carrera,
    )

    primera_matricula = matriculas_actuales.first()
    if primera_matricula and primera_matricula.paralelo.nombre == nombre_destino:
        return (False, "El Estudiante ya pertenece al Paralelo especificado")

    cohorte = (
        primera_matricula.cohorte_de_matricula
        if primera_matricula else _obtener_o_crear_cohorte(periodo, carrera)
    )

    with transaction.atomic():
        matriculas_actuales.delete()
        for paralelo_db in paralelos_destino:
            MatriculaParalelo.objects.get_or_create(
                estudiante=estudiante_db,
                paralelo=paralelo_db,
                defaults={"cohorte_de_matricula": cohorte},
            )

    return (True, "El Estudiante fue reasignado correctamente")


def _paralelos_del_grupo_de_estudiantes(paralelo_db):
    # Todas las filas Paralelo (una por unidad) del mismo paralelo lógico.
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera
    return Paralelo.objects.filter(
        periodo_de_nivelacion=paralelo_db.periodo_de_nivelacion,
        jornada=paralelo_db.jornada,
        nombre=paralelo_db.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    )


def periodo_permite_gestion_matriculas(periodo_db):
    # La matrícula manual solo se puede gestionar en Planificación o En curso.
    return periodo_db.estado in (EstadoDePeriodo.PLANIFICACION.value, EstadoDePeriodo.EN_CURSO.value)


def servicio_retirar_estudiante_de_paralelo(estudiante_db, paralelo_db):
    periodo = paralelo_db.periodo_de_nivelacion
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera

    if not periodo_permite_gestion_matriculas(periodo):
        return (False, "Solo se puede gestionar la matrícula en Periodos en planificación o en curso")

    paralelos_grupo = _paralelos_del_grupo_de_estudiantes(paralelo_db)
    matriculas = MatriculaParalelo.objects.filter(estudiante=estudiante_db, paralelo__in=paralelos_grupo)
    if not matriculas.exists():
        return (False, "El Estudiante no pertenece al Paralelo especificado")

    with transaction.atomic():
        matriculas.delete()

    servicio_recalcular_cohorte_de_carrera(periodo, carrera)
    return (True, "El Estudiante fue retirado del Paralelo correctamente")


def servicio_agregar_estudiante_a_paralelo(estudiante_db, paralelo_db):
    periodo = paralelo_db.periodo_de_nivelacion
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera

    if not periodo_permite_gestion_matriculas(periodo):
        return (False, "Solo se puede gestionar la matrícula en Periodos en planificación o en curso")

    # Coherencia: mismo periodo, misma carrera y misma jornada.
    if (estudiante_db.periodo_de_nivelacion_id != periodo.id
            or estudiante_db.carrera_registrada_id != carrera.id
            or estudiante_db.jornada != paralelo_db.jornada):
        return (False, "El Estudiante no es compatible con el Paralelo (Carrera o Jornada)")

    # Solo estudiantes con matrícula activa (no retirados ni anulados).
    if estudiante_db.estado_de_matricula != EstadoDeMatricula.MATRICULADO.value:
        return (False, "El Estudiante no tiene una matrícula activa")

    # No debe estar ya asignado a ningún paralelo del periodo.
    if MatriculaParalelo.objects.filter(estudiante=estudiante_db, paralelo__periodo_de_nivelacion=periodo).exists():
        return (False, "El Estudiante ya se encuentra asignado a un Paralelo")

    paralelos_grupo = list(_paralelos_del_grupo_de_estudiantes(paralelo_db))
    representativo = paralelos_grupo[0]
    ocupacion = MatriculaParalelo.objects.filter(paralelo=representativo).count()
    if ocupacion >= representativo.capacidad_maxima:
        return (False, "El Paralelo no presenta cupos disponibles")

    cohorte = _obtener_o_crear_cohorte(periodo, carrera)
    with transaction.atomic():
        ya_matriculado = set(
            MatriculaParalelo.objects.filter(
                estudiante=estudiante_db, paralelo__in=paralelos_grupo
            ).values_list("paralelo_id", flat=True)
        )
        nuevas = [
            MatriculaParalelo(estudiante=estudiante_db, paralelo=p, cohorte_de_matricula=cohorte)
            for p in paralelos_grupo if p.id not in ya_matriculado
        ]
        if nuevas:
            MatriculaParalelo.objects.bulk_create(nuevas)

    servicio_recalcular_cohorte_de_carrera(periodo, carrera)
    return (True, "El Estudiante fue agregado al Paralelo correctamente")


# ==========================================
# B.2 - Horarios
# ==========================================
def _construir_horario_poo(horario_db):
    return HorarioPOO(
        dia_semana=obtener_enum_flexible(DiaDeSemana, horario_db.dia_semana),
        hora_inicio=horario_db.hora_inicio,
        hora_fin=horario_db.hora_fin,
        espacio_de_imparticion=horario_db.espacio_de_imparticion,
    )

def _construir_paralelo_poo_con_horarios(paralelo_db):
    from poo.clases.paralelo import Paralelo as ParaleloBase

    paralelo_poo = ParaleloBase(
        codigo_de_paralelo=paralelo_db.codigo_de_paralelo,
        nombre=paralelo_db.nombre,
        jornada=obtener_enum_flexible(Jornada, paralelo_db.jornada),
        modalidad=obtener_enum_flexible(Modalidad, paralelo_db.modalidad),
        capacidad_maxima=paralelo_db.capacidad_maxima,
    )
    for horario_db in Horario.objects.filter(paralelo=paralelo_db):
        paralelo_poo.agregar_horario(_construir_horario_poo(horario_db))
    return paralelo_poo

def servicio_horas_agendadas_paralelo(paralelo_db):
    return _construir_paralelo_poo_con_horarios(paralelo_db).calcular_horas_agendadas()

def _horarios_externos_para_paralelo(paralelo_db):
    # Sesiones del mismo paralelo lógico (otras unidades del grupo, que comparten estudiantes)
    # + sesiones del docente responsable en otros paralelos del periodo (evita doble reserva).
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera
    paralelos_del_grupo = Paralelo.objects.filter(
        periodo_de_nivelacion=paralelo_db.periodo_de_nivelacion,
        jornada=paralelo_db.jornada,
        nombre=paralelo_db.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    )
    horarios_db = list(
        Horario.objects.filter(paralelo__in=paralelos_del_grupo).exclude(paralelo=paralelo_db)
    )
    if paralelo_db.docente_responsable_id:
        ids_vistos = {h.id for h in horarios_db}
        for h in Horario.objects.filter(
            paralelo__periodo_de_nivelacion=paralelo_db.periodo_de_nivelacion,
            paralelo__docente_responsable_id=paralelo_db.docente_responsable_id,
        ).exclude(paralelo=paralelo_db):
            if h.id not in ids_vistos:
                horarios_db.append(h)
                ids_vistos.add(h.id)
    return horarios_db

def periodo_en_planificacion(periodo_db):
    from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
    return periodo_db.estado == EstadoDePeriodo.PLANIFICACION.value

def servicio_registrar_horario(paralelo_db, dia_semana, hora_inicio, hora_fin, espacio):
    from poo.clases.franja_horaria import sesion_dentro_de_franja, DURACIONES_VALIDAS, texto_franja

    if not periodo_en_planificacion(paralelo_db.periodo_de_nivelacion):
        return (False, "Solo se pueden gestionar horarios en un Periodo en planificación")

    try:
        enum_dia = obtener_enum_flexible(DiaDeSemana, dia_semana)
    except ValueError:
        return (False, "Día no válido")

    # Una sola sesión de la misma unidad por día.
    if Horario.objects.filter(paralelo=paralelo_db, dia_semana=dia_semana).exists():
        return (False, "La Unidad curricular ya tiene una sesión ese día (solo se permite una por día)")

    nuevo_horario_poo = HorarioPOO(
        dia_semana=enum_dia,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
        espacio_de_imparticion=espacio,
    )

    errores_horario = nuevo_horario_poo.validar_datos_de_registro()
    if errores_horario:
        return (False, list(errores_horario.values())[0])

    jornada = obtener_enum_flexible(Jornada, paralelo_db.jornada)
    if not sesion_dentro_de_franja(jornada, hora_inicio, hora_fin):
        return (False, f"La sesión debe estar dentro de la franja de la jornada {paralelo_db.jornada} ({texto_franja(jornada)})")

    if nuevo_horario_poo.determinar_duracion_horas() not in DURACIONES_VALIDAS:
        return (False, "La duración de la sesión debe ser de 1, 2 o 3 horas")

    horarios_externos = [_construir_horario_poo(h) for h in _horarios_externos_para_paralelo(paralelo_db)]

    paralelo_poo = _construir_paralelo_poo_con_horarios(paralelo_db)
    horas_semanales = _horas_sincronicas_semanales(paralelo_db.unidad_curricular, paralelo_db.periodo_de_nivelacion)
    resultado = paralelo_poo.validar_nuevo_horario(
        nuevo_horario_poo, horas_semanales, horarios_externos
    )

    if not resultado["ok"]:
        if resultado["motivo"] == "conflicto":
            conflicto = resultado["horario_en_conflicto"]
            return (
                False,
                f"Conflicto horario en {conflicto.dia_semana.value} "
                f"{conflicto.hora_inicio.strftime('%H:%M')}–{conflicto.hora_fin.strftime('%H:%M')}"
            )
        return (
            False,
            "La sesión excede las horas sincrónicas semanales de la unidad curricular")

    Horario.objects.create(
        paralelo=paralelo_db,
        dia_semana=dia_semana,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
        espacio_de_imparticion=espacio,
    )
    return (True, "El horario ha sido registrado correctamente")


def _sugerir_bloque_libre(dia_poo, franja_inicio, franja_fin, bloque_horas, ocupados):
    from datetime import datetime, timedelta
    base = datetime(2000, 1, 1, franja_inicio.hour, franja_inicio.minute)
    tope = datetime(2000, 1, 1, franja_fin.hour, franja_fin.minute)
    dur = timedelta(hours=bloque_horas)
    paso = timedelta(minutes=60)
    actual = base
    while actual + dur <= tope:
        h_ini = actual.time()
        h_fin = (actual + dur).time()
        candidato = HorarioPOO(
            dia_semana=dia_poo, hora_inicio=h_ini, hora_fin=h_fin,
            espacio_de_imparticion="",
        )
        if not any(candidato.verificar_conflicto_horario(o) for o in ocupados):
            return (h_ini, h_fin)
        actual += paso
    return None


def _distribucion_simetrica(horas, max_dias=7, min_h=1, max_h=3):
    # Reparte 'horas' en bloques de HORAS ENTERAS lo más iguales posible, uno por día,
    # cada bloque entre min_h y max_h, usando hasta max_dias días.
    import math
    h = int(round(horas))
    if h < min_h:
        return []
    dias = min(max_dias, h)
    dias = max(dias, math.ceil(h / max_h))
    dias = min(dias, max_dias)
    dias = max(dias, 1)
    base = h // dias
    extra = h % dias
    return [min(base + (1 if i < extra else 0), max_h) for i in range(dias)]


def _generar_horario_para_unidad(paralelo_unidad_db):
    from poo.clases.franja_horaria import obtener_franja

    jornada = obtener_enum_flexible(Jornada, paralelo_unidad_db.jornada)
    franja = obtener_franja(jornada)
    if not franja:
        return (0, 0.0)
    franja_inicio, franja_fin = franja

    requeridas = _horas_sincronicas_semanales(
        paralelo_unidad_db.unidad_curricular, paralelo_unidad_db.periodo_de_nivelacion
    )
    paralelo_poo = _construir_paralelo_poo_con_horarios(paralelo_unidad_db)
    restante = round(requeridas - paralelo_poo.calcular_horas_agendadas(), 2)
    if restante <= 0:
        return (0, 0.0)

    ocupados = list(paralelo_poo.horarios) + [
        _construir_horario_poo(h) for h in _horarios_externos_para_paralelo(paralelo_unidad_db)
    ]
    bloques = _distribucion_simetrica(restante)
    dias = list(DiaDeSemana)

    nuevos_db = []
    generadas = 0.0
    for dia, bloque in zip(dias, bloques):
        candidato = _sugerir_bloque_libre(dia, franja_inicio, franja_fin, bloque, ocupados)
        if not candidato:
            continue
        h_ini, h_fin = candidato
        ocupados.append(HorarioPOO(
            dia_semana=dia, hora_inicio=h_ini, hora_fin=h_fin,
            espacio_de_imparticion="",
        ))
        nuevos_db.append(Horario(
            paralelo=paralelo_unidad_db, dia_semana=dia.value, hora_inicio=h_ini, hora_fin=h_fin,
            espacio_de_imparticion="",
        ))
        generadas = round(generadas + bloque, 2)

    if nuevos_db:
        Horario.objects.bulk_create(nuevos_db)
    return (len(nuevos_db), round(restante - generadas, 2))


def servicio_generar_horario_sugerido(representativo_db):
    # Genera el horario de TODO el paralelo lógico (todas sus unidades), de forma simétrica.
    if not periodo_en_planificacion(representativo_db.periodo_de_nivelacion):
        return (False, "Solo se pueden gestionar horarios en un Periodo en planificación")
    unidades = list(
        _paralelos_del_grupo_de_estudiantes(representativo_db).select_related("unidad_curricular")
    )

    total_creados = 0
    faltantes = []
    for row in unidades:
        creados, faltante = _generar_horario_para_unidad(row)
        total_creados += creados
        if faltante > 0:
            faltantes.append((row.unidad_curricular.nombre, faltante))

    if total_creados == 0 and not faltantes:
        return (False, "El horario ya cubre las horas sincrónicas semanales de todas las unidades")
    if total_creados == 0:
        return (False, "No se encontraron espacios libres dentro de la franja para generar el horario")

    mensaje = f"Se generaron {total_creados} sesiones automáticamente"
    if faltantes:
        detalle = "; ".join(f"{nombre} (faltan {horas} h)" for nombre, horas in faltantes)
        mensaje += f". Sin espacio suficiente en la franja para: {detalle}"
    return (True, mensaje)


def servicio_editar_horario(horario_db, dia_semana, hora_inicio, hora_fin, espacio):
    from poo.clases.franja_horaria import sesion_dentro_de_franja, DURACIONES_VALIDAS, texto_franja

    paralelo_db = horario_db.paralelo
    if not periodo_en_planificacion(paralelo_db.periodo_de_nivelacion):
        return (False, "Solo se pueden gestionar horarios en un Periodo en planificación")
    try:
        enum_dia = obtener_enum_flexible(DiaDeSemana, dia_semana)
    except ValueError:
        return (False, "Día no válido")

    # Una sola sesión de la misma unidad por día (excluyendo la sesión editada).
    if Horario.objects.filter(paralelo=paralelo_db, dia_semana=dia_semana).exclude(id=horario_db.id).exists():
        return (False, "La Unidad curricular ya tiene una sesión ese día (solo se permite una por día)")

    nuevo = HorarioPOO(
        dia_semana=enum_dia, hora_inicio=hora_inicio, hora_fin=hora_fin,
        espacio_de_imparticion=espacio,
    )
    errores = nuevo.validar_datos_de_registro()
    if errores:
        return (False, list(errores.values())[0])

    jornada = obtener_enum_flexible(Jornada, paralelo_db.jornada)
    if not sesion_dentro_de_franja(jornada, hora_inicio, hora_fin):
        return (False, f"La sesión debe estar dentro de la franja de la jornada {paralelo_db.jornada} ({texto_franja(jornada)})")

    dur = nuevo.determinar_duracion_horas()
    if dur not in DURACIONES_VALIDAS:
        return (False, "La duración de la sesión debe ser de 1, 2 o 3 horas")

    propios = [_construir_horario_poo(h) for h in Horario.objects.filter(paralelo=paralelo_db).exclude(id=horario_db.id)]
    externos = [_construir_horario_poo(h) for h in _horarios_externos_para_paralelo(paralelo_db)]
    for o in propios + externos:
        if nuevo.verificar_conflicto_horario(o):
            return (False, f"Conflicto horario en {o.dia_semana.value} {o.hora_inicio.strftime('%H:%M')}–{o.hora_fin.strftime('%H:%M')}")

    horas_otras = round(sum(h.determinar_duracion_horas() for h in propios), 2)
    requeridas = _horas_sincronicas_semanales(paralelo_db.unidad_curricular, paralelo_db.periodo_de_nivelacion)
    if horas_otras + dur > requeridas:
        return (False, "La sesión excede las horas sincrónicas semanales de la unidad curricular")

    horario_db.dia_semana = dia_semana
    horario_db.hora_inicio = hora_inicio
    horario_db.hora_fin = hora_fin
    horario_db.espacio_de_imparticion = espacio
    horario_db.save()
    return (True, "El horario ha sido actualizado correctamente")

def servicio_obtener_matriz_de_horarios(periodo_db, paralelos_db):
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica

    periodo_poo = _construir_periodo(periodo_db)
    paralelos_poo = [_construir_paralelo_poo_con_horarios(p) for p in paralelos_db]

    facade = CentroDeOperacionAcademica()
    matriz = facade.obtener_matriz_de_horarios(periodo_poo, paralelos_poo)

    mapa_unidad = {
        p.codigo_de_paralelo: p.unidad_curricular.nombre for p in paralelos_db
    }
    for fila in matriz:
        for bloque in fila["bloques"]:
            bloque["unidad"] = mapa_unidad.get(bloque["codigo_de_paralelo"], "—")

    return matriz


def _semanas_de_periodo(periodo_db):
    dias = (periodo_db.fecha_fin - periodo_db.fecha_inicio).days
    semanas = dias // 7
    return semanas if semanas > 0 else 1


def _horas_sincronicas_semanales(unidad, periodo_db):
    # Fuente de verdad: las horas sincrónicas SEMANALES definidas en la unidad.
    # Retrocompatibilidad: si no se definieron, se derivan de las totales y las semanas.
    if getattr(unidad, "horas_sincronicas_semanales", 0):
        return round(unidad.horas_sincronicas_semanales, 2)
    return round(unidad.horas_sincronicas / _semanas_de_periodo(periodo_db), 2)


def _construir_docente_poo_para_periodo(docente_db, periodo_db, paralelo_excluir_id=None):
    from usuarios.services import _crear_docente
    from poo.clases.enums.estado_de_vinculacion import EstadoDeVinculacion as EnumEstadoDeVinculacion

    docente_poo = _crear_docente(docente_db)
    docente_poo.establecer_estado_de_vinculacion(
        obtener_enum_flexible(EnumEstadoDeVinculacion, docente_db.estado_de_vinculacion)
    )
    docente_poo.definir_especialidades(docente_db.especialidades or [])

    paralelos_actuales = Paralelo.objects.filter(
        periodo_de_nivelacion=periodo_db, docente_responsable=docente_db
    ).select_related("unidad_curricular")
    if paralelo_excluir_id:
        paralelos_actuales = paralelos_actuales.exclude(id=paralelo_excluir_id)

    carga_actual = 0.0
    for paralelo_actual in paralelos_actuales:
        carga_actual += _horas_sincronicas_semanales(paralelo_actual.unidad_curricular, periodo_db)
        for horario_db in Horario.objects.filter(paralelo=paralelo_actual):
            docente_poo.agregar_horario_ocupado(_construir_horario_poo(horario_db))

    carga_actual = round(carga_actual, 2)
    docente_poo.registrar_carga_actual(carga_actual)
    return docente_poo, carga_actual

def _texto_motivo_no_asignable(resultado):
    motivo = resultado.get("motivo")
    if motivo == "inactivo":
        return "El docente no está activo"
    if motivo == "conflicto":
        conflicto = resultado["horario_en_conflicto"]
        return (
            f"Conflicto de Horario el día '{conflicto.dia_semana.value}'. {conflicto.hora_inicio}–{conflicto.hora_fin}"
        )
    if motivo == "carga":
        return (
            f"Excede la carga máxima ({resultado['carga_actual']} + {resultado['horas_nuevas']} h > {resultado['carga_maxima']} h máx.)"
        )
    return "El Docente no puede ser asignado."

def servicio_evaluar_docentes_para_paralelo(paralelo_db):
    from usuarios.models import PerfilDocente

    periodo = paralelo_db.periodo_de_nivelacion
    unidad = paralelo_db.unidad_curricular
    areas = []
    horas_unidad = _horas_sincronicas_semanales(unidad, periodo)

    docentes = PerfilDocente.objects.filter(
        universidad=periodo.universidad
    ).select_related("usuario_de_sistema").order_by(
        "usuario_de_sistema__apellidos", "usuario_de_sistema__nombres"
    )

    evaluaciones = []
    for docente_db in docentes:
        # Solo se listan los docentes COMPATIBLES CON LA JORNADA del paralelo.
        # La carga horaria (y las demás restricciones) NO se validan aquí para
        # listar: eso se verifica al pulsar "Asignar" (servicio_asignar_docente).
        if paralelo_db.jornada not in (docente_db.jornadas or []):
            continue
        docente_poo, carga_actual = _construir_docente_poo_para_periodo(
            docente_db, periodo, paralelo_excluir_id=paralelo_db.id
        )
        es_actual = paralelo_db.docente_responsable_id == docente_db.id
        evaluaciones.append({
            "docente": docente_db,
            "es_actual": es_actual,
            "carga_real": round(carga_actual + (horas_unidad if es_actual else 0), 2),
            "carga_maxima": docente_db.carga_horaria_maxima,
            "horas_unidad": horas_unidad,
            "activo": docente_poo.esta_activo(),
        })
    return evaluaciones

def servicio_asignar_docente(paralelo_db, docente_db):
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica

    periodo = paralelo_db.periodo_de_nivelacion
    unidad = paralelo_db.unidad_curricular
    areas = []
    horas_unidad = _horas_sincronicas_semanales(unidad, periodo)

    if not periodo_en_planificacion(periodo):
        return (False, "Solo se puede asignar docentes en un Periodo en planificación", None)

    if not Horario.objects.filter(paralelo=paralelo_db).exists():
        return (False, "Debe registrar el horario de la Unidad curricular antes de asignar un Docente", None)

    # Exigir que el horario esté COMPLETO (horas agendadas >= horas semanales requeridas).
    horas_agendadas = servicio_horas_agendadas_paralelo(paralelo_db)
    if horas_agendadas < horas_unidad:
        return (False, f"El horario de la Unidad curricular no está completo ({horas_agendadas}/{horas_unidad} h semanales)", None)

    docente_poo, carga_actual = _construir_docente_poo_para_periodo(
        docente_db, periodo, paralelo_excluir_id=paralelo_db.id
    )
    paralelo_poo = _construir_paralelo_poo_con_horarios(paralelo_db)

    if paralelo_db.jornada not in (docente_db.jornadas or []):
        return (False, "El Docente no atiende la jornada del paralelo", None)

    facade = CentroDeOperacionAcademica()
    resultado = facade.validar_asignacion_docente(docente_poo, paralelo_poo, horas_unidad, areas)

    if not resultado["ok"]:
        return (False, _texto_motivo_no_asignable(resultado), None)

    paralelo_db.docente_responsable = docente_db
    paralelo_db.save(update_fields=["docente_responsable"])

    advertencia = None
    if resultado.get("advertencia") == "especialidad":
        advertencia = ("El área de conocimiento de la Unidad curricular no coincide con las especialidades registradas del Docente")
    return (True, "El Docente ha sido asignado correctamente", advertencia)

def servicio_quitar_docente(paralelo_db):
    if not periodo_en_planificacion(paralelo_db.periodo_de_nivelacion):
        return (False, "Solo se puede modificar la asignación en un Periodo en planificación")
    docente_db = paralelo_db.docente_responsable
    if not docente_db:
        return (False, "El paralelo no tiene un docente designado actualmente")

    paralelo_db.docente_responsable = None
    paralelo_db.save(update_fields=["docente_responsable"])

    return (True, "El Docente ha sido desvinculado correctamente")