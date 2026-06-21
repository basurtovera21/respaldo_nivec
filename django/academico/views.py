from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
from usuarios.utils import generar_identificador_siguiente

from .models import (Universidad, Campus, Carrera, MallaCurricular, UnidadCurricular,
                     PeriodoDeNivelacion, Paralelo, Horario, CohorteDeMatricula,
                     MatriculaParalelo, ConsolidadoAcademico, EvaluacionAcademica,
                     IncidenciaAcademica, EvaluacionDeDesempeno, InformeGeneral)
from .forms import (FormularioUniversidad, FormularioCampus, FormularioCarrera,
                    FormularioMallaCurricular, FormularioUnidadCurricular,
                    FormularioPeriodoDeNivelacion, FormularioParalelo, FormularioHorario,
                    FormularioCohorteDeMatricula, FormularioMatriculaParalelo,
                    FormularioConsolidadoAcademico, FormularioEvaluacionAcademica,
                    FormularioIncidenciaAcademica, FormularioEvaluacionDeDesempeno,
                    FormularioInformeGeneral)
from . import services
from datetime import datetime, date

#Entidades base
@login_required
def detalle_universidad(request):
    universidad = request.user.perfil_administrativo.universidad
    if not universidad:
        return redirect("registrar_universidad")
        
    return render(request, "entidades/detalle_universidad.html", {
        "universidad": universidad,
        "titulo_pagina": "Universidad - NIVEC"
    })

@login_required
def registrar_universidad(request):
    if request.user.perfil_administrativo.universidad:
        return redirect("panel_principal")

    if request.method == "POST":
        formulario_universidad = FormularioUniversidad(request.POST, request.FILES)
        if formulario_universidad.is_valid():
            nueva_universidad = formulario_universidad.save()
            perfil = request.user.perfil_administrativo
            perfil.universidad = nueva_universidad
            perfil.save()
            
            messages.success(request, "La universidad ha sido registrada correctamente")
            return redirect("panel_principal")
    else:
        formulario_universidad = FormularioUniversidad()
        
    return render(request, "entidades/formulario_universidad.html", {
        "formulario": formulario_universidad,
        "titulo_pagina": "Universidad - NIVEC",
        "titulo": "Registrar universidad",
        "boton_texto": "Registrar",
        "url_cancelar": "panel_principal",
        "mostrar_carga_masiva": False
    })

@login_required
def modificar_universidad(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    if request.method == "POST":
        formulario = FormularioUniversidad(request.POST, request.FILES, instance=universidad_usuario)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "La universidad ha sido modificada correctamente")
            return redirect("detalle_universidad")
    else:
        formulario = FormularioUniversidad(instance=universidad_usuario)
        
    return render(request, "entidades/formulario_universidad.html", {
        "formulario": formulario,
        "titulo_pagina": "Universidad - NIVEC",
        "titulo": "Modificar universidad",
        "boton_texto": "Modificar",
        "url_cancelar": "detalle_universidad",
        "mostrar_carga_masiva": False
    })


@login_required
def listar_campus(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    campus = Campus.objects.filter(universidad=universidad_usuario)
    
    return render(request, "entidades/listar_campus.html", {
        "campus": campus,
        "titulo_pagina": "Campus - NIVEC",
        "titulo": "Campus",
        "url_registrar": "registrar_campus",
        "texto_registrar": "Registrar",
        "url_volver": "panel_principal"
    })

@login_required
def descargar_plantilla_campus(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campus"

    cabeceras = ["Nombre del campus", "Dirección física", "Provincia"]
    ws.append(cabeceras)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_campus_nivec.xlsx"'
    wb.save(response)
    return response

@login_required
def registrar_campus(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    if request.method == "POST":
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_campus")

            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                
                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    nombre = fila[0]
                    direccion = fila[1]
                    provincia = fila[2]
                    
                    if not nombre and not direccion and not provincia:
                        continue
                        
                    if not nombre or not direccion or not provincia:
                        messages.warning(request, f"El registro de la fila {numero_fila} fue omitido por falta de información")
                        continue
                    
                    nuevo_codigo = generar_identificador_siguiente(Campus, 'CAM', 'codigo_de_campus')

                    Campus.objects.create(
                        universidad=universidad_usuario,
                        codigo_de_campus=nuevo_codigo,
                        nombre=nombre,
                        direccion_fisica=direccion,
                        provincia=provincia
                    )
                    registros_exitosos += 1
                
                messages.success(request, f"{registros_exitosos} campus han sido registrados correctamente")
                return redirect("listar_campus")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento ({str(e)})")
                return redirect("registrar_campus")

        else:
            formulario = FormularioCampus(request.POST)
            
            if 'codigo_de_campus' in formulario.fields:
                formulario.fields['codigo_de_campus'].required = False

            if formulario.is_valid():
                nuevo_campus = formulario.save(commit=False)
                nuevo_campus.universidad = universidad_usuario
                nuevo_campus.codigo_de_campus = generar_identificador_siguiente(Campus, 'CAM', 'codigo_de_campus')
                nuevo_campus.save()
                messages.success(request, "El campus ha sido registrado correctamente")
                return redirect("listar_campus")
    else:
        formulario = FormularioCampus()
        
    return render(request, "entidades/formulario_campus.html", {
        "formulario": formulario,
        "titulo_pagina": "Campus - NIVEC",
        "titulo": "Registrar campus",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_campus",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_campus"
    })
    
@login_required
def modificar_campus(request, campus_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    campus = get_object_or_404(Campus, id=campus_id, universidad=universidad_usuario)

    if request.method == "POST":
        formulario = FormularioCampus(request.POST, instance=campus)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "El campus ha sido modificado correctamente")
            return redirect("listar_campus")
    else:
        formulario = FormularioCampus(instance=campus)
        
    return render(request, "entidades/formulario_campus.html", {
        "formulario": formulario,
        "titulo_pagina": "Campus - NIVEC",
        "titulo": "Modificar campus",
        "boton_texto": "Modificar",
        "url_cancelar": "listar_campus",
        "mostrar_carga_masiva": False
    })

@login_required
def eliminar_campus(request, campus_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    campus = get_object_or_404(Campus, id=campus_id, universidad=universidad_usuario)
    campus.delete()
    
    messages.success(request, "El campus ha sido eliminado correctamente")
    return redirect("listar_campus")


@login_required
def listar_carreras(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    if not Campus.objects.filter(universidad=universidad_usuario).exists():
        messages.info(request, "No hay registros de campus actualmente")
        return redirect("registrar_campus")

    carreras = Carrera.objects.filter(campus__universidad=universidad_usuario).select_related("campus")
    
    return render(request, "entidades/listar_carreras.html", {
        "carreras": carreras,
        "titulo_pagina": "Carrera - NIVEC",
        "titulo": "Carreras",
        "url_registrar": "registrar_carrera",
        "texto_registrar": "Registrar",
        "url_volver": "panel_principal"
    })

@login_required
def descargar_plantilla_carrera(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carreras"

    cabeceras = [
        "Código de campus (CAM...)", 
        "Nombre de la carrera", 
        "Modalidad (Virtual, Presencial, Semipresencial)",
        "Facultad", 
        "Vigencia SNIESE (AAAA-MM-DD)"
    ]
    ws.append(cabeceras)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40  
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_carreras_nivec.xlsx"'
    wb.save(response)
    return response


@login_required
def registrar_carrera(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    campus_existente = Campus.objects.filter(universidad=universidad_usuario)
    
    if not campus_existente.exists():
        messages.warning(request, "No hay registros de campus actualmente")
        return redirect("registrar_campus")

    if request.method == "POST":

        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_carrera")

            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                
                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    codigo_campus = fila[0]
                    nombre = fila[1]
                    modalidad = fila[2]
                    facultad = fila[3]
                    vigencia = fila[4]
                    
                    if not codigo_campus and not nombre and not modalidad and not facultad and not vigencia:
                        continue
                        
                    if not codigo_campus or not nombre or not modalidad or not facultad or not vigencia:
                        messages.warning(request, f"El registro de la fila {numero_fila} fue omitido por falta de información")
                        continue
                    
                    opciones_validas = ["Virtual", "Presencial", "Semipresencial"]
                    modalidad_str = str(modalidad).strip().capitalize()
                    
                    if modalidad_str not in opciones_validas:
                        messages.warning(request, f"El registro de la fila {numero_fila} fue omitido ('{modalidad}' no reconocida)")
                        continue
                    modalidad = modalidad_str
                    
                    if isinstance(vigencia, (datetime, date)):
                        if isinstance(vigencia, datetime):
                            vigencia = vigencia.date()
                    elif isinstance(vigencia, str):
                        try:
                            vigencia = datetime.strptime(vigencia.strip(), "%Y-%m-%d").date()
                        except ValueError:
                            messages.warning(request, f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                            continue
                    else:
                        messages.warning(request, f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                        continue

                    campus_obj = campus_existente.filter(codigo_de_campus=codigo_campus).first()
                    
                    if not campus_obj:
                        messages.warning(request, f"El registro de la fila {numero_fila} fue omitido por falta de información (código de campus no válido)")
                        continue
                    
                    nuevo_codigo = generar_identificador_siguiente(Carrera, 'CAR', 'codigo_de_carrera')

                    Carrera.objects.create(
                        campus=campus_obj,
                        codigo_de_carrera=nuevo_codigo,
                        nombre=nombre,
                        modalidad=modalidad, 
                        facultad=facultad,
                        vigencia_sniese=vigencia
                    )
                    registros_exitosos += 1
                
                if registros_exitosos > 0:
                    messages.success(request, f"{registros_exitosos} carreras han sido registradas correctamente")
                else:
                    messages.warning(request, "No se registraron carreras nuevas")
                    
                return redirect("listar_carreras")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento ({str(e)})")
                return redirect("registrar_carrera")

        else:
            formulario = FormularioCarrera(request.POST)
            
            if 'codigo_de_carrera' in formulario.fields:
                formulario.fields['codigo_de_carrera'].required = False

            if formulario.is_valid():
                nueva_carrera = formulario.save(commit=False)
                nueva_carrera.codigo_de_carrera = generar_identificador_siguiente(Carrera, 'CAR', 'codigo_de_carrera')
                nueva_carrera.save()
                
                messages.success(request, "La carrera ha sido registrada correctamente")
                return redirect("listar_carreras")
    else:
        formulario = FormularioCarrera()
        formulario.fields['campus'].queryset = campus_existente
        
    return render(request, "entidades/formulario_carrera.html", {
        "formulario": formulario,
        "titulo_pagina": "Carrera - NIVEC",
        "titulo": "Registrar carrera",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_carreras",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_carrera"
    })

@login_required
def modificar_carrera(request, carrera_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    carrera = get_object_or_404(Carrera, id=carrera_id, campus__universidad=universidad_usuario)

    if request.method == "POST":
        formulario = FormularioCarrera(request.POST, instance=carrera)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "La carrera ha sido modificada correctamente")
            return redirect("listar_carreras")
    else:
        formulario = FormularioCarrera(instance=carrera)
        formulario.fields['campus'].queryset = Campus.objects.filter(universidad=universidad_usuario)

    return render(request, "entidades/formulario_carrera.html", {
        "formulario": formulario,
        "titulo_pagina": "Carrera - NIVEC",
        "titulo": "Modificar carrera",
        "boton_texto": "Modificar",
        "url_cancelar": "listar_carreras",
        "mostrar_carga_masiva": False
    })

@login_required
def eliminar_carrera(request, carrera_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    carrera = get_object_or_404(Carrera, id=carrera_id, campus__universidad=universidad_usuario)
    carrera.delete()
    
    messages.success(request, "La carrera ha sido eliminada correctamente")
    return redirect("listar_carreras")


#Estructura académica
@login_required
def listar_mallas(request):
    mallas = MallaCurricular.objects.all().select_related("carrera")
    return render(request, "academico/listar_mallas.html", {"mallas": mallas})


@login_required
def registrar_malla(request):
    if request.method == "POST":
        formulario = FormularioMallaCurricular(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Malla curricular registrada correctamente.")
            return redirect("listar_mallas")
    else:
        formulario = FormularioMallaCurricular()
    return render(request, "academico/formulario_malla.html", {
        "formulario": formulario,
        "titulo": "Registrar malla curricular",
    })


@login_required
def listar_unidades(request):
    unidades = UnidadCurricular.objects.all().select_related("malla_curricular")
    return render(request, "academico/listar_unidades.html", {"unidades": unidades})


@login_required
def registrar_unidad(request):
    if request.method == "POST":
        formulario = FormularioUnidadCurricular(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Unidad curricular registrada correctamente.")
            return redirect("listar_unidades")
    else:
        formulario = FormularioUnidadCurricular()
    return render(request, "academico/formulario_unidad.html", {
        "formulario": formulario,
        "titulo": "Registrar unidad curricular",
    })


@login_required
def listar_periodos(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    periodos = PeriodoDeNivelacion.objects.filter(universidad=universidad_usuario).order_by("-anio", "-numero_periodo")
    
    return render(request, "academico/listar_periodos.html", {
        "periodos": periodos,
        "titulo_pagina": "Periodo de nivelación - NIVEC",
        "titulo": "Periodos de Nivelación",
        "url_registrar": "registrar_periodo",
        "texto_registrar": "Registrar",
        "url_volver": "panel_principal"
    })

@login_required
def registrar_periodo(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente.")
        return redirect("registrar_universidad")

    if request.method == "POST":
        formulario = FormularioPeriodoDeNivelacion(request.POST, universidad=universidad_usuario)
        if formulario.is_valid():
            nuevo_periodo = formulario.save(commit=False)
            nuevo_periodo.universidad = universidad_usuario
            nuevo_periodo.codigo_periodo = generar_identificador_siguiente(PeriodoDeNivelacion, 'PNV', 'codigo_periodo')
            nuevo_periodo.periodo = f"{nuevo_periodo.anio}-{nuevo_periodo.numero_periodo}"
            nuevo_periodo.save()
            
            messages.success(request, "El periodo de nivelación ha sido registrado correctamente")
            return redirect("listar_periodos")
    else:
        formulario = FormularioPeriodoDeNivelacion(universidad=universidad_usuario)
        
    return render(request, "academico/formulario_periodo.html", {
        "formulario": formulario,
        "titulo_pagina": "Periodos - NIVEC",
        "titulo": "Registrar periodo de nivelación",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_periodos",
        "mostrar_carga_masiva": False,
    })

@login_required
def modificar_periodo(request, periodo_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    periodo = get_object_or_404(PeriodoDeNivelacion, id=periodo_id, universidad=universidad_usuario)

    if request.method == "POST":

        formulario = FormularioPeriodoDeNivelacion(request.POST, instance=periodo, universidad=universidad_usuario)
        if formulario.is_valid():
            periodo_modificado = formulario.save(commit=False)
            periodo_modificado.periodo = f"{periodo_modificado.anio}-{periodo_modificado.numero_periodo}"
            periodo_modificado.save()
            messages.success(request, "El periodo de nivelación ha sido modificado correctamente")
            return redirect("listar_periodos")
    else:
        formulario = FormularioPeriodoDeNivelacion(instance=periodo, universidad=universidad_usuario)
        
    return render(request, "academico/formulario_periodo.html", {
        "formulario": formulario,
        "titulo_pagina": "Modificar Periodo - NIVEC",
        "titulo": "Modificar periodo de nivelación", 
        "boton_texto": "Modificar",                  
        "url_cancelar": "listar_periodos",
        "mostrar_carga_masiva": False,
        "periodo": periodo
    })
    
@login_required
def eliminar_periodo(request, periodo_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("registrar_universidad")

    periodo = get_object_or_404(PeriodoDeNivelacion, id=periodo_id, universidad=universidad_usuario)
    periodo.delete()
    messages.success(request, "El periodo de nivelación ha sido eliminado correctamente")
    return redirect("listar_periodos")

@login_required
def iniciar_periodo(request, periodo_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    periodo = get_object_or_404(PeriodoDeNivelacion, pk=periodo_id, universidad=universidad_usuario)
    resultado = services.servicio_iniciar_periodo_de_nivelacion(periodo)
    if resultado:
        messages.success(request, f"El periodo de nivelación {periodo.periodo} ha iniciado")
    else:
        messages.error(request, "No se ha podido iniciar el periodo de nivelación")
        
    return redirect("listar_periodos")

@login_required
def finalizar_periodo(request, periodo_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    periodo = get_object_or_404(PeriodoDeNivelacion, pk=periodo_id, universidad=universidad_usuario)
    resultado = services.servicio_finalizar_periodo_de_nivelacion(periodo)
    if resultado:
        messages.success(request, f"El periodo de nivelación {periodo.periodo} ha finalizado")
    else:
        messages.error(request, "No se ha podido finalizar el periodo de nivelación")
        
    return redirect("listar_periodos")


@login_required
def listar_paralelos(request):
    paralelos = Paralelo.objects.all().select_related(
        "periodo_de_nivelacion", "unidad_curricular", "docente_responsable"
    )
    return render(request, "academico/listar_paralelos.html", {"paralelos": paralelos})

@login_required
def registrar_paralelo(request):
    if request.method == "POST":
        formulario = FormularioParalelo(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Paralelo registrado correctamente.")
            return redirect("listar_paralelos")
    else:
        formulario = FormularioParalelo()
    return render(request, "academico/formulario_paralelo.html", {
        "formulario": formulario,
        "titulo": "Registrar paralelo",
    })


@login_required
def registrar_horario(request):
    if request.method == "POST":
        formulario = FormularioHorario(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Horario registrado correctamente.")
            return redirect("listar_paralelos")
    else:
        formulario = FormularioHorario()
    return render(request, "academico/formulario_horario.html", {
        "formulario": formulario,
        "titulo": "Registrar horario",
    })


#Procesos académicos
@login_required
def listar_cohortes(request):
    cohortes = CohorteDeMatricula.objects.all().select_related(
        "periodo_de_nivelacion", "carrera_registrada"
    )
    return render(request, "academico/listar_cohortes.html", {"cohortes": cohortes})


@login_required
def registrar_cohorte(request):
    if request.method == "POST":
        formulario = FormularioCohorteDeMatricula(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Cohorte de matrícula registrada correctamente.")
            return redirect("listar_cohortes")
    else:
        formulario = FormularioCohorteDeMatricula()
    return render(request, "academico/formulario_cohorte.html", {
        "formulario": formulario,
        "titulo": "Registrar cohorte de matrícula",
    })


@login_required
def registrar_matricula(request):
    if request.method == "POST":
        formulario = FormularioMatriculaParalelo(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Matrícula en paralelo registrada correctamente.")
            return redirect("listar_paralelos")
    else:
        formulario = FormularioMatriculaParalelo()
    return render(request, "academico/formulario_matricula.html", {
        "formulario": formulario,
        "titulo": "Registrar matrícula en paralelo",
    })


@login_required
def procesar_mtn(request):
    from academico.forms import FormularioConsolidadoAcademico
    if request.method == "POST" and request.FILES.get("documento_mtn"):
        periodo_id = request.POST.get("periodo")
        periodo = get_object_or_404(PeriodoDeNivelacion, pk=periodo_id)
        resultado = services.servicio_procesar_mtn(request.FILES["documento_mtn"], periodo)
        messages.success(
            request,
            f"MTN procesada: {resultado['registros_validos']} válidos, "
            f"{resultado['registros_observados']} observados de {resultado['registros_totales']}."
        )
        for observacion in resultado["filas_observadas"]:
            messages.warning(request, observacion)
        return redirect("listar_estudiantes")
    periodos = PeriodoDeNivelacion.objects.filter(estado="Planificación")
    return render(request, "dan/cargar_mtn.html", {
        "periodos": periodos,
        "titulo": "Procesar Matriz de Tercer Nivel (MTN)",
    })


@login_required
def distribuir_estudiantes(request, periodo_id):
    periodo = get_object_or_404(PeriodoDeNivelacion, pk=periodo_id)
    no_asignados = services.servicio_distribuir_estudiantes(periodo)
    if no_asignados:
        messages.warning(request, f"{len(no_asignados)} estudiante(s) sin paralelo disponible.")
    else:
        messages.success(request, "Todos los estudiantes fueron asignados correctamente.")
    return redirect("listar_paralelos")


@login_required
def listar_evaluaciones(request):
    evaluaciones = EvaluacionAcademica.objects.all().select_related(
        "estudiante", "unidad_curricular"
    )
    return render(request, "academico/listar_evaluaciones.html", {"evaluaciones": evaluaciones})

@login_required
def registrar_evaluacion(request):
    if request.method == "POST":
        formulario = FormularioEvaluacionAcademica(request.POST)
        if formulario.is_valid():
            evaluacion = formulario.save(commit=False)
            evaluacion.save()
            services.servicio_registrar_evaluacion(evaluacion)
            messages.success(request, "Evaluación registrada correctamente.")
            return redirect("listar_evaluaciones")
    else:
        formulario = FormularioEvaluacionAcademica()
    return render(request, "academico/formulario_evaluacion.html", {
        "formulario": formulario,
        "titulo": "Registrar evaluación académica",
    })


@login_required
def listar_incidencias(request):
    incidencias = IncidenciaAcademica.objects.all().select_related(
        "docente_implicado", "responsable_autorizacion"
    )
    return render(request, "academico/listar_incidencias.html", {"incidencias": incidencias})

@login_required
def registrar_incidencia(request):
    if request.method == "POST":
        formulario = FormularioIncidenciaAcademica(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Incidencia registrada correctamente.")
            return redirect("listar_incidencias")
    else:
        formulario = FormularioIncidenciaAcademica()
    return render(request, "academico/formulario_incidencia.html", {
        "formulario": formulario,
        "titulo": "Registrar incidencia académica",
    })


@login_required
def registrar_desempeno(request):
    if request.method == "POST":
        formulario = FormularioEvaluacionDeDesempeno(request.POST)
        if formulario.is_valid():
            evaluacion = formulario.save(commit=False)
            from poo.servicios.estrategia_de_evaluacion_estandar import EstrategiaDeEvaluacionEstandar
            estrategia = EstrategiaDeEvaluacionEstandar(
                porcentaje_horas = 0.25,
                porcentaje_notas = 0.25,
                porcentaje_aprobacion = 0.25,
                porcentaje_evaluacion_estudiantil = 0.25
            )
            from poo.clases.evaluacion_de_desempeno import EvaluacionDeDesempeno as EvaluacionDeDesempenoPOO
            evaluacion_poo = EvaluacionDeDesempenoPOO(
                docente_responsable = None,
                porcentaje_de_horas_cumplidas = evaluacion.porcentaje_de_horas_cumplidas,
                entrega_oportuna_de_calificaciones = evaluacion.entrega_oportuna_de_calificaciones,
                porcentaje_de_aprobacion_paralelo = evaluacion.porcentaje_de_aprobacion_paralelo,
                resultado_de_evaluacion_estudiantil = evaluacion.resultado_de_evaluacion_estudiantil
            )
            evaluacion.puntaje_final = evaluacion_poo.procesar_evaluacion(estrategia)
            evaluacion.save()
            messages.success(request, "Evaluación de desempeño registrada correctamente.")
            return redirect("listar_docentes")
    else:
        formulario = FormularioEvaluacionDeDesempeno()
    return render(request, "academico/formulario_desempeno.html", {
        "formulario": formulario,
        "titulo": "Registrar evaluación de desempeño",
    })


#Informes
@login_required
def listar_informes(request):
    informes = InformeGeneral.objects.all().select_related("periodo_academico")
    return render(request, "academico/listar_informes.html", {"informes": informes})


@login_required
def registrar_informe(request):
    if request.method == "POST":
        formulario = FormularioInformeGeneral(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Informe registrado correctamente.")
            return redirect("listar_informes")
    else:
        formulario = FormularioInformeGeneral()
    return render(request, "academico/formulario_informe.html", {
        "formulario": formulario,
        "titulo": "Registrar informe general",
    })


@login_required
def emitir_informe(request, informe_id):
    informe = get_object_or_404(InformeGeneral, pk=informe_id)
    from poo.clases.informe_general import InformeGeneral as InformeGeneralPOO
    from poo.clases.enums.tipo_de_informe import TipoDeInforme
    from poo.clases.periodo_de_nivelacion import PeriodoDeNivelacion as PeriodoDeNivelacionBase
    from poo.clases.enums.modalidad import Modalidad

    periodo_poo = PeriodoDeNivelacionBase(
        codigo_periodo = informe.periodo_academico.codigo_periodo,
        anio = informe.periodo_academico.anio,
        periodo = informe.periodo_academico.periodo,
        fecha_inicio = informe.periodo_academico.fecha_inicio,
        fecha_fin = informe.periodo_academico.fecha_fin,
        modalidad = Modalidad(informe.periodo_academico.modalidad),
        numero_periodo = informe.periodo_academico.numero_periodo
    )
    periodo_poo._estado = __import__(
        'poo.clases.enums.estado_de_periodo', fromlist=['EstadoDePeriodo']
    ).EstadoDePeriodo(informe.periodo_academico.estado)

    informe_poo = InformeGeneralPOO(
        codigo_de_informe = informe.codigo_de_informe,
        periodo_academico = periodo_poo,
        tipo_de_informe = TipoDeInforme(informe.tipo_de_informe)
    )
    resultado = informe_poo.emitir_informe_de_nivelacion()
    if resultado:
        from datetime import date
        informe.estado_de_informe = "Revisión"
        informe.fecha_de_emision = date.today()
        informe.save()
        messages.success(request, "Informe emitido correctamente.")
    else:
        messages.error(request, "No se pudo emitir el informe. Verifique que el periodo esté cerrado.")
    return redirect("listar_informes")


@login_required
def exportar_informe(request, informe_id):
    informe = get_object_or_404(InformeGeneral, pk=informe_id)
    formato = request.GET.get("formato", "excel")
    return services.servicio_exportar_informe(informe, formato)