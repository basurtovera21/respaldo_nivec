from django.shortcuts import render
from django.shortcuts import redirect
from django.shortcuts import get_object_or_404
import openpyxl
from django.http import HttpResponse

from django.contrib.auth.decorators import login_required
from django.contrib import messages

from django.views.decorators.cache import never_cache

from .models import UsuarioDeSistema
from .models import PerfilDocente
from .models import PerfilEstudiante
from .models import PerfilAdministrativo

from .forms import (
    FormularioUsuarioDeSistema, 
    FormularioModificarUsuarioDeSistema, 
    FormularioPerfilEstudiante, 
    FormularioPerfilDocente, 
    FormularioRegistrarPerfilAdministrativo,
    FormularioModificarPerfilAdministrativo,
    FormularioRegistrarCoordinadorDAN,
    FormularioRegistrarCoordinadorUA,
    FormularioDatosDocenteUA,
    FormularioRegistrarDocente,
)

from . import services

from usuarios.utils import generar_identificador_siguiente

from poo.clases.enums.perfil_administrativo import PerfilAdministrativo as EnumPerfilAdministrativo
from poo.clases.usuarios.usuario_administrativo import UsuarioAdministrativo as UsuarioAdministrativoBase
from poo.clases.enums.estado_de_usuario import EstadoDeUsuario

from poo.clases.enums.tipo_de_vinculacion import TipoDeVinculacion
from poo.clases.enums.tiempo_de_dedicacion import TiempoDeDedicacion
from poo.clases.enums.estado_de_vinculacion import EstadoDeVinculacion
from poo.clases.enums.estado_de_matricula import EstadoDeMatricula
from poo.clases.enums.jornada import Jornada
from poo.clases.enums.registro_de_cupo import RegistroDeCupo

#Autenticación
@never_cache
def iniciar_sesion(request):
    if request.user.is_authenticated:
        return redirect("panel_principal")
    
    if request.method == "POST":
        correo_institucional = request.POST.get("correo_institucional")
        contrasena = request.POST.get("contrasena")
        
        resultado_inicio = services.servicio_iniciar_sesion(request, correo_institucional, contrasena)
        
        if resultado_inicio["exito"]:
            return redirect("panel_principal")

        messages.error(request, resultado_inicio["mensaje"])
    
    return render(request, "autenticacion/iniciar_sesion.html")


@login_required
def cerrar_sesion(request):
    services.servicio_cerrar_sesion(request)
    return redirect("iniciar_sesion")


def panel_sin_perfil(request):
    return render(request, "autenticacion/sin_perfil.html")


@login_required
@never_cache
def panel_principal(request):
    usuario = request.user
    
    perfil_administrativo = PerfilAdministrativo.objects.filter(usuario_de_sistema=usuario).first()
    
    if perfil_administrativo:
        tipo_de_perfil = perfil_administrativo.perfil_administrativo

        if tipo_de_perfil == EnumPerfilAdministrativo.COORDINADOR_UA.value:
            return redirect('panel_ua')
        
        elif tipo_de_perfil == EnumPerfilAdministrativo.DIRECTOR_DAN.value:
            return redirect('panel_director_dan')
            
        elif tipo_de_perfil == EnumPerfilAdministrativo.COORDINADOR_DAN.value:
            return redirect('panel_dan')
            
        else:
            return redirect('panel_administrativo')
            
            
    elif PerfilDocente.objects.filter(usuario_de_sistema=usuario).exists():
        return redirect('panel_docente')
        
    elif PerfilEstudiante.objects.filter(usuario_de_sistema=usuario).exists():
        return redirect('panel_estudiante')
        
    return redirect('sin_perfil')


@login_required
@never_cache
def panel_director_dan(request):
    return render(request, "administrativo/panel_director_dan.html")


#Panel por tipo de usuario de sistema
@login_required
def panel_dan(request):
    from academico.models import PeriodoDeNivelacion, ConsolidadoAcademico
    periodos_de_nivelacion = PeriodoDeNivelacion.objects.all().order_by("-anio")
    consolidados_academicos = ConsolidadoAcademico.objects.all()
    return render(request, "dan/panel_dan.html", {
        "periodos_de_nivelacion": periodos_de_nivelacion,
        "consolidados_academicos": consolidados_academicos,
    })


@login_required
def panel_coordinador_ua(request):
    from academico.models import Paralelo
    paralelos = Paralelo.objects.all().order_by("periodo_de_nivelacion")
    docentes = PerfilDocente.objects.filter(estado_de_vinculacion="Activo")
    return render(request, "coordinador_ua/panel_coordinador_ua.html", {
        "paralelos": paralelos,
        "docentes": docentes,
    })


@login_required
def panel_docente(request):
    try:
        perfil_docente = request.user.perfil_docente
        from academico.models import Paralelo
        paralelos = Paralelo.objects.filter(docente_responsable=perfil_docente)
        carga_academica = services.servicio_obtener_carga_academica(perfil_docente)
        return render(request, "docente/panel_docente.html", {
            "perfil_docente": perfil_docente,
            "paralelos": paralelos,
            "carga": carga_academica,
        })
    except PerfilDocente.DoesNotExist:
        return redirect("panel_principal")


@login_required
def panel_estudiante(request):
    try:
        perfil_estudiante = request.user.perfil_estudiante
        from academico.models import MatriculaParalelo, EvaluacionAcademica
        matriculas = MatriculaParalelo.objects.filter(
            estudiante=perfil_estudiante
        ).select_related("paralelo", "paralelo__unidad_curricular")
        evaluaciones = EvaluacionAcademica.objects.filter(estudiante=perfil_estudiante)
        registro = services.servicio_obtener_registro_institucional(perfil_estudiante)
        return render(request, "estudiante/panel_estudiante.html", {
            "perfil_estudiante": perfil_estudiante,
            "matriculas": matriculas,
            "evaluaciones": evaluaciones,
            "registro": registro,
        })
    except PerfilEstudiante.DoesNotExist:
        return redirect("panel_principal")

@login_required
def panel_administrativo(request):
    return render(request, "administrativo/panel_director_dan.html", {
        "titulo": "Panel Administrativo",
    })


#Usuarios
@login_required
def listar_usuarios(request):
    usuarios = UsuarioDeSistema.objects.all()
    return render(request, "usuarios/listar_usuarios.html", {"usuarios": usuarios})


@login_required
def registrar_usuario(request):
    if request.method == "POST":
        formulario = FormularioUsuarioDeSistema(request.POST)
        if formulario.is_valid():
            usuario = formulario.save(commit=False)
            
            usuario.set_password(usuario.identificacion)
            
            usuario.save()
            messages.success(request, "El usuario ha sido registrado correctamente.")
            return redirect("listar_usuarios")
    else:
        formulario = FormularioUsuarioDeSistema()
    return render(request, "usuarios/formulario_usuario.html", {
        "formulario": formulario,
        "titulo": "Registrar usuario",
    })


@login_required
def listar_docentes(request):
    docentes = PerfilDocente.objects.all().select_related("usuario_de_sistema")
    return render(request, "usuarios/listar_docentes.html", {"docentes": docentes})


@login_required
def registrar_docente(request):
    if request.method == "POST":
        formulario_usuario = FormularioUsuarioDeSistema(request.POST)
        formulario_perfil = FormularioPerfilDocente(request.POST)
        
        formulario_perfil.fields['identificador_institucional'].required = False
        
        if formulario_usuario.is_valid() and formulario_perfil.is_valid():
            usuario = formulario_usuario.save(commit=False)
            
            usuario.set_password(usuario.identificacion)
            
            usuario.save()
            
            perfil = formulario_perfil.save(commit=False)
            perfil.usuario_de_sistema = usuario
            
            perfil.identificador_institucional = generar_identificador_siguiente(
                PerfilDocente, 'DC', 'identificador_institucional'
            )
            
            perfil.save()
            messages.success(request, "Docente registrado correctamente.")
            return redirect("listar_docentes")
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        
        siguiente_id = generar_identificador_siguiente(PerfilDocente, 'DC', 'identificador_institucional')
        formulario_perfil = FormularioPerfilDocente(initial={
            'identificador_institucional': siguiente_id
        })
        
    return render(request, "usuarios/formulario_docente.html", {
        "formulario_usuario": formulario_usuario,
        "formulario_perfil": formulario_perfil,
        "titulo": "Registrar docente",
    })


@login_required
def listar_estudiantes(request):
    estudiantes = PerfilEstudiante.objects.all().select_related(
        "usuario_de_sistema", "carrera_registrada", "campus_registrado"
    )
    return render(request, "usuarios/listar_estudiantes.html", {"estudiantes": estudiantes})


@login_required
def registrar_estudiante(request):
    if request.method == "POST":
        formulario_usuario = FormularioUsuarioDeSistema(request.POST)
        formulario_perfil = FormularioPerfilEstudiante(request.POST)
        
        formulario_perfil.fields['identificador_institucional'].required = False
        formulario_perfil.fields['numero_de_matricula'].required = False
        
        if formulario_usuario.is_valid() and formulario_perfil.is_valid():
            usuario = formulario_usuario.save(commit=False)
            
            usuario.set_password(usuario.identificacion)
            
            usuario.save()
            
            perfil = formulario_perfil.save(commit=False)
            perfil.usuario_de_sistema = usuario
            
            perfil.identificador_institucional = generar_identificador_siguiente(
                PerfilEstudiante, 'ES', 'identificador_institucional'
            )
            perfil.numero_de_matricula = generar_identificador_siguiente(
                PerfilEstudiante, 'MAT', 'numero_de_matricula'
            )
            
            perfil.save()
            messages.success(request, "Estudiante registrado correctamente.")
            return redirect("listar_estudiantes")
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        
        siguiente_id = generar_identificador_siguiente(PerfilEstudiante, 'ES', 'identificador_institucional')
        siguiente_matricula = generar_identificador_siguiente(PerfilEstudiante, 'MAT', 'numero_de_matricula')
        
        formulario_perfil = FormularioPerfilEstudiante(initial={
            'identificador_institucional': siguiente_id,
            'numero_de_matricula': siguiente_matricula
        })
        
    return render(request, "usuarios/formulario_estudiante.html", {
        "formulario_usuario": formulario_usuario,
        "formulario_perfil": formulario_perfil,
        "titulo": "Registrar estudiante",
    })


@login_required
def listar_administrativos(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente.")
        return redirect("panel_principal")

    administrativos = PerfilAdministrativo.objects.filter(
        universidad=universidad_usuario
    ).exclude(
        perfil_administrativo__in=[
            EnumPerfilAdministrativo.COORDINADOR_DAN.value,
            EnumPerfilAdministrativo.COORDINADOR_UA.value
        ]
    ).select_related("usuario_de_sistema")
    
    return render(request, "usuarios/listar_administrativos.html", {
        "administrativos": administrativos,
        "titulo_pagina": "Administrativo - NIVEC",
        "titulo": "Usuarios administrativos",
        "url_registrar": "registrar_administrativo",
        "texto_registrar": "Registrar",
        "url_volver": "panel_principal"
    })

@login_required
def descargar_plantilla_administrativo(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Administrativos"

    cabeceras = [
        "Tipo de identificación (Cédula, Pasaporte, Cédula extranjera)", 
        "Número de identificación", 
        "Nombres",
        "Apellidos", 
        "Correo institucional",
        "Perfil administrativo (Rector, Vicerrector académico)"
    ]
    ws.append(cabeceras)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40  
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_administrativo_nivec.xlsx"'
    wb.save(response)
    return response

@login_required
def registrar_administrativo(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    if request.method == "POST":
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_administrativo")
                
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                
                mapeo_prefijos = {
                    EnumPerfilAdministrativo.RECTOR.value: "AD",
                    EnumPerfilAdministrativo.VICERRECTOR_ACADEMICO.value: "AD",
                    EnumPerfilAdministrativo.DIRECTOR_DAN.value: "DAN",
                    EnumPerfilAdministrativo.COORDINADOR_DAN.value: "CAN",
                    EnumPerfilAdministrativo.COORDINADOR_UA.value: "CUA",
                }

                perfiles_validos_diccionario = {str(e.value).strip().lower(): e.value for e in EnumPerfilAdministrativo}

                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        tipo_id = fila[0]
                        identificacion = fila[1]
                        nombres = fila[2]
                        apellidos = fila[3]
                        correo = fila[4]
                        perfil_str = fila[5]

                        if not any([tipo_id, identificacion, nombres, apellidos, correo, perfil_str]):
                            continue
                        
                        if not all([tipo_id, identificacion, nombres, apellidos, correo, perfil_str]):
                            messages.warning(request, f"Fila {numero_fila} omitida por falta de información")
                            continue
                            
                        tipo_id_str = str(tipo_id).strip().capitalize()
                        identificacion_str = str(identificacion).strip()
                        perfil_ingresado_limpio = str(perfil_str).strip().lower()

                        if perfil_ingresado_limpio not in perfiles_validos_diccionario:
                            messages.warning(request, f"Fila {numero_fila} omitida (perfil '{perfil_str}' no válido)")
                            continue
                        
                        perfil_exacto = perfiles_validos_diccionario[perfil_ingresado_limpio]

                        try:
                            UsuarioAdministrativoBase.validar_creacion_de_usuario_administrativo(
                                identificacion=identificacion_str, 
                                contrasena=identificacion_str
                            )
                        except ValueError as e:
                            messages.warning(request, f"Fila {numero_fila} omitida: {str(e)}")
                            continue

                        if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                            messages.warning(request, f"Fila {numero_fila} omitida (el usuario administrativo ya se encuentra registrado)")
                            continue

                        usuario = UsuarioDeSistema.objects.create(
                            tipo_de_identificacion=tipo_id_str,
                            identificacion=identificacion_str,
                            nombres=str(nombres).strip(),
                            apellidos=str(apellidos).strip(),
                            correo_institucional=str(correo).strip(),
                            estado_de_usuario=EstadoDeUsuario.ACTIVO.value
                        )
                        usuario.set_password(identificacion_str)
                        usuario.save()

                        prefijo = mapeo_prefijos.get(perfil_exacto, "AD")
                        nuevo_identificador = generar_identificador_siguiente(
                            PerfilAdministrativo, prefijo, 'identificador_administrativo'
                        )

                        PerfilAdministrativo.objects.create(
                            usuario_de_sistema=usuario,
                            universidad=universidad_usuario,
                            identificador_administrativo=nuevo_identificador,
                            perfil_administrativo=perfil_exacto
                        )
                        
                        registros_exitosos += 1
                        
                    except Exception as error_de_fila:
                        messages.warning(request, f"Fila {numero_fila} omitida ({str(error_de_fila)})")
                        continue
                    
                if registros_exitosos > 0:
                    messages.success(request, f"{registros_exitosos} usuarios administrativos registrados correctamente")
                else:
                    messages.warning(request, "No se procesaron registros")
                    
                return redirect("listar_administrativos")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento")
                return redirect("registrar_administrativo")

        else:
            formulario_usuario = FormularioUsuarioDeSistema(request.POST)
            formulario_perfil = FormularioRegistrarPerfilAdministrativo(request.POST)

            if formulario_usuario.is_valid() and formulario_perfil.is_valid():
                usuario = formulario_usuario.save(commit=False)
                usuario.set_password(usuario.identificacion)
                usuario.save()
                
                perfil = formulario_perfil.save(commit=False)
                perfil.usuario_de_sistema = usuario
                perfil.universidad = universidad_usuario
                
                mapeo_prefijos = {
                    EnumPerfilAdministrativo.RECTOR.value: "AD",
                    EnumPerfilAdministrativo.VICERRECTOR_ACADEMICO.value: "AD",
                    EnumPerfilAdministrativo.DIRECTOR_DAN.value: "DAN",
                    EnumPerfilAdministrativo.COORDINADOR_DAN.value: "CAN",
                    EnumPerfilAdministrativo.COORDINADOR_UA.value: "CUA",
                }
                
                prefijo = mapeo_prefijos.get(perfil.perfil_administrativo, "AD")
                perfil.identificador_administrativo = generar_identificador_siguiente(
                    PerfilAdministrativo, prefijo, 'identificador_administrativo'
                )
                
                perfil.save()
                messages.success(request, "El usuario administrativo ha sido registrado correctamente")
                return redirect("listar_administrativos")
            
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        formulario_perfil = FormularioRegistrarPerfilAdministrativo()

    return render(request, "usuarios/formulario_administrativo.html", {
        "form_usuario": formulario_usuario,
        "form_perfil": formulario_perfil,
        "titulo": "Registrar usuario administrativo",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_administrativos",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_administrativo"
    })

@login_required
def modificar_administrativo(request, admin_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    perfil = get_object_or_404(PerfilAdministrativo, pk=admin_id, universidad=universidad_usuario)
    
    if perfil.perfil_administrativo == "Director de dirección de admisión y nivelación":
        messages.error(request, "El usuario no puede ser modificado")
        return redirect("listar_administrativos")

    usuario = perfil.usuario_de_sistema

    if request.method == "POST":
        formulario_usuario = FormularioModificarUsuarioDeSistema(request.POST, instance=usuario)
        formulario_perfil = FormularioModificarPerfilAdministrativo(request.POST, instance=perfil)

        if formulario_usuario.is_valid() and formulario_perfil.is_valid():
            usuario_modificado = formulario_usuario.save(commit=False)
            
            nueva_contrasena = formulario_usuario.cleaned_data.get("contrasena")
            if nueva_contrasena:
                usuario_modificado.set_password(nueva_contrasena)
            
            usuario_modificado.save()
            formulario_perfil.save()
            
            if perfil.perfil_administrativo == EnumPerfilAdministrativo.COORDINADOR_DAN.value:
                messages.success(request, "El coordinador de dirección de admisión y nivelación ha sido modificado correctamente")
                return redirect("listar_coordinadores_dan")
            elif perfil.perfil_administrativo == EnumPerfilAdministrativo.COORDINADOR_UA.value:
                messages.success(request, "El coordinador de unidad académica ha sido modificado correctamente")
                return redirect("listar_coordinadores_ua")
            else:
                messages.success(request, "El usuario administrativo ha sido modificado correctamente")
                return redirect("listar_administrativos")
    else:
        formulario_usuario = FormularioModificarUsuarioDeSistema(instance=usuario)
        formulario_perfil = FormularioModificarPerfilAdministrativo(instance=perfil)

    if perfil.perfil_administrativo == EnumPerfilAdministrativo.COORDINADOR_DAN.value:
        url_cancelar = "listar_coordinadores_dan"
        titulo_pagina = "Coordinador de dirección de admisión y nivelación - NIVEC"
        titulo_seccion = "Modificar coordinador de dirección de admisión y nivelación"
    elif perfil.perfil_administrativo == EnumPerfilAdministrativo.COORDINADOR_UA.value:
        url_cancelar = "listar_coordinadores_ua"
        titulo_pagina = "Coordinador de unidad académica - NIVEC"
        titulo_seccion = "Modificar coordinador de unidad académica"
    else:
        url_cancelar = "listar_administrativos"
        titulo_pagina = "Administrativo - NIVEC"
        titulo_seccion = "Modificar usuario administrativo"

    return render(request, "usuarios/formulario_administrativo.html", {
        "form_usuario": formulario_usuario,
        "form_perfil": formulario_perfil,
        "titulo": titulo_seccion,
        "subtitulo": f"{usuario.nombres} {usuario.apellidos}",
        "boton_texto": "Modificar",
        "url_cancelar": url_cancelar,
        "url_volver": url_cancelar,
        "titulo_pagina": titulo_pagina,
        "mostrar_carga_masiva": False
    })
    
@login_required
def eliminar_administrativo(request, admin_id):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    perfil = get_object_or_404(PerfilAdministrativo, pk=admin_id, universidad=universidad_usuario)
    
    if perfil.perfil_administrativo == "Director de dirección de admisión y nivelación":
        messages.error(request, "El usuario no puede ser eliminado")
        return redirect("listar_administrativos")

    rol_eliminado = perfil.perfil_administrativo
    usuario = perfil.usuario_de_sistema
    usuario.delete()
    
    if rol_eliminado == EnumPerfilAdministrativo.COORDINADOR_DAN.value:
        messages.success(request, "El coordinador de dirección de admisión y nivelación ha sido eliminado correctamente")
        return redirect("listar_coordinadores_dan")
    elif rol_eliminado == EnumPerfilAdministrativo.COORDINADOR_UA.value:
        messages.success(request, "El coordinador de unidad académica ha sido eliminado correctamente")
        return redirect("listar_coordinadores_ua")
    else:
        messages.success(request, "El usuario administrativo ha sido eliminado correctamente")
        return redirect("listar_administrativos")

@login_required
def listar_coordinadores_dan(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente.")
        return redirect("panel_principal")

    coordinadores = PerfilAdministrativo.objects.filter(
        universidad=universidad_usuario,
        perfil_administrativo=EnumPerfilAdministrativo.COORDINADOR_DAN.value
    ).select_related("usuario_de_sistema")
    
    return render(request, "usuarios/listar_coordinadores_dan.html", {
        "coordinadores": coordinadores,
        "titulo_pagina": "Coordinador de dirección de admisión y nivelación - NIVEC",
        "titulo": "Coordinadores de dirección de admisión y nivelación",
        "url_registrar": "registrar_coordinador_dan",
        "texto_registrar": "Registrar",
        "url_volver": "panel_director_dan"
    })

@login_required
def descargar_plantilla_coordinador_dan(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coordinadores DAN"

    cabeceras = [
        "Tipo de identificación (Cédula, Pasaporte, Cédula extranjera)", 
        "Número de identificación", 
        "Nombres",
        "Apellidos", 
        "Correo institucional"
    ]
    ws.append(cabeceras)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40  
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_coordinador_dan_nivec.xlsx"'
    wb.save(response)
    return response

@login_required
def registrar_coordinador_dan(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    if request.method == "POST":
        
        # === 1. LÓGICA DE CARGA MASIVA (EXCEL) ===
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_coordinador_dan")
                
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                
                # El rol es fijo para todo este documento
                rol_fijo = EnumPerfilAdministrativo.COORDINADOR_DAN.value

                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        tipo_id = fila[0]
                        identificacion = fila[1]
                        nombres = fila[2]
                        apellidos = fila[3]
                        correo = fila[4]
                        
                        if not any([tipo_id, identificacion, nombres, apellidos, correo]):
                            continue
                            
                        if not all([tipo_id, identificacion, nombres, apellidos, correo]):
                            messages.warning(request, f"Fila {numero_fila} omitida por falta de información")
                            continue
                            
                        tipo_id_str = str(tipo_id).strip().capitalize()
                        identificacion_str = str(identificacion).strip()

                        try:
                            UsuarioAdministrativoBase.validar_creacion_de_usuario_administrativo(
                                identificacion=identificacion_str, 
                                contrasena=identificacion_str
                            )
                        except ValueError as e:
                            messages.warning(request, f"Fila {numero_fila} omitida: {str(e)}")
                            continue

                        if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                            messages.warning(request, f"Fila {numero_fila} omitida (el coordinador de dirección de admisión y nivelación ya se encuentra registrado)")
                            continue

                        usuario = UsuarioDeSistema.objects.create(
                            tipo_de_identificacion=tipo_id_str,
                            identificacion=identificacion_str,
                            nombres=str(nombres).strip(),
                            apellidos=str(apellidos).strip(),
                            correo_institucional=str(correo).strip(),
                            estado_de_usuario=EstadoDeUsuario.ACTIVO.value
                        )
                        usuario.set_password(identificacion_str)
                        usuario.save()

                        # Generamos AMBOS identificadores
                        nuevo_id_admin = generar_identificador_siguiente(
                            PerfilAdministrativo, "CAN", 'identificador_administrativo'
                        )
                        nuevo_id_dan = generar_identificador_siguiente(
                            PerfilAdministrativo, "DAN", 'identificador_coordinador_dan'
                        )

                        PerfilAdministrativo.objects.create(
                            usuario_de_sistema=usuario,
                            universidad=universidad_usuario,
                            identificador_administrativo=nuevo_id_admin,
                            identificador_coordinador_dan=nuevo_id_dan,
                            perfil_administrativo=rol_fijo
                        )
                        
                        registros_exitosos += 1
                        
                    except Exception as error_de_fila:
                        messages.warning(request, f"Fila {numero_fila} omitida ({str(error_de_fila)})")
                        continue
                    
                if registros_exitosos > 0:
                    messages.success(request, f"{registros_exitosos} coordinadores de dirección de admisión y nivelación registrados correctamente")
                else:
                    messages.warning(request, "No se procesaron registros")
                    
                return redirect("listar_coordinadores_dan")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento")
                return redirect("registrar_coordinador_dan")

        # === 2. LÓGICA DE FORMULARIO MANUAL ===
        else:
            formulario_usuario = FormularioUsuarioDeSistema(request.POST)
            formulario_perfil = FormularioRegistrarCoordinadorDAN(request.POST)

            if formulario_usuario.is_valid() and formulario_perfil.is_valid():
                usuario = formulario_usuario.save(commit=False)
                usuario.set_password(usuario.identificacion)
                usuario.save()
                
                perfil = formulario_perfil.save(commit=False)
                perfil.usuario_de_sistema = usuario
                perfil.universidad = universidad_usuario
                
                perfil.identificador_administrativo = generar_identificador_siguiente(
                    PerfilAdministrativo, "CAN", 'identificador_administrativo'
                )
                perfil.identificador_coordinador_dan = generar_identificador_siguiente(
                    PerfilAdministrativo, "DAN", 'identificador_coordinador_dan'
                )
                
                perfil.save()
                messages.success(request, "El coordinador de dirección de admisión y nivelación ha sido registrado correctamente")
                return redirect("listar_coordinadores_dan")
                
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        formulario_perfil = FormularioRegistrarCoordinadorDAN()

    return render(request, "usuarios/formulario_administrativo.html", {
        "form_usuario": formulario_usuario,
        "form_perfil": formulario_perfil,
        "titulo_pagina": "Coordinador de dirección de admisión y nivelación - NIVEC",
        "titulo": "Registrar coordinador de dirección de admisión y nivelación",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_coordinadores_dan",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_coordinador_dan"
    })

@login_required
def descargar_plantilla_coordinador_ua(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coordinadores UA"

    # Actualizamos las cabeceras con las opciones exactas de tus Enums
    cabeceras = [
        "Tipo de identificación (Cédula, Pasaporte, Cédula extranjera)", 
        "Número de identificación", 
        "Nombres",
        "Apellidos", 
        "Correo institucional",
        "Unidad académica",
        "Tipo de vinculación (Nombramiento, Contrato, Ocasional, Honorario)",
        "Tiempo de dedicación (Tiempo completo, Medio tiempo, Tiempo parcial)",
        "Carga horaria máxima (número decimal)"
    ]
    ws.append(cabeceras)

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 32
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_coordinador_ua_nivec.xlsx"'
    wb.save(response)
    return response


@login_required
def registrar_coordinador_ua(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente")
        return redirect("panel_principal")

    if request.method == "POST":
        
        # === 1. LÓGICA DE CARGA MASIVA (EXCEL) ===
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_coordinador_ua")
                
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                rol_fijo = EnumPerfilAdministrativo.COORDINADOR_UA.value
                
                # Estos diccionarios leerán automáticamente "Tiempo completo", "Nombramiento", etc.
                vinc_validas = {str(e.value).strip().lower(): e.value for e in TipoDeVinculacion}
                dedic_validas = {str(e.value).strip().lower(): e.value for e in TiempoDeDedicacion}

                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Extraemos las 9 columnas de la fila
                        tipo_id, identificacion, nombres, apellidos, correo, ua, tipo_vinc, tiempo_dedic, carga_max = fila[:9]
                        
                        if not any([tipo_id, identificacion, nombres, apellidos, correo, ua, tipo_vinc, tiempo_dedic, carga_max]):
                            continue
                            
                        if not all([tipo_id, identificacion, nombres, apellidos, correo, ua, tipo_vinc, tiempo_dedic, carga_max]):
                            messages.warning(request, f"Fila {numero_fila} omitida por falta de información")
                            continue
                            
                        tipo_id_str = str(tipo_id).strip().capitalize()
                        identificacion_str = str(identificacion).strip()
                        ua_str = str(ua).strip()
                        tipo_vinc_limpio = str(tipo_vinc).strip().lower()
                        tiempo_dedic_limpio = str(tiempo_dedic).strip().lower()

                        # Validar Enums Docentes inteligentemente
                        if tipo_vinc_limpio not in vinc_validas:
                            messages.warning(request, f"Fila {numero_fila} omitida (tipo de vinculación '{tipo_vinc}' no válido)")
                            continue
                        if tiempo_dedic_limpio not in dedic_validas:
                            messages.warning(request, f"Fila {numero_fila} omitida (tiempo de dedicación '{tiempo_dedic}' no válido)")
                            continue
                            
                        # Rescatamos el valor oficial (ej: "Medio tiempo")
                        vinc_exacta = vinc_validas[tipo_vinc_limpio]
                        dedic_exacta = dedic_validas[tiempo_dedic_limpio]

                        try:
                            carga_max_float = float(carga_max)
                        except ValueError:
                            messages.warning(request, f"Fila {numero_fila} omitida (la carga horaria debe ser un número decimal)")
                            continue

                        try:
                            UsuarioAdministrativoBase.validar_creacion_de_usuario_administrativo(
                                identificacion=identificacion_str, 
                                contrasena=identificacion_str
                            )
                        except ValueError as e:
                            messages.warning(request, f"Fila {numero_fila} omitida: {str(e)}")
                            continue

                        if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                            messages.warning(request, f"Fila {numero_fila} omitida (el coordinador de unidad académica ya se encuentra registrado)")
                            continue

                        # 1. Crear Usuario
                        usuario = UsuarioDeSistema.objects.create(
                            tipo_de_identificacion=tipo_id_str,
                            identificacion=identificacion_str,
                            nombres=str(nombres).strip(),
                            apellidos=str(apellidos).strip(),
                            correo_institucional=str(correo).strip(),
                            estado_de_usuario=EstadoDeUsuario.ACTIVO.value
                        )
                        usuario.set_password(identificacion_str)
                        usuario.save()

                        # 2. Crear Perfil Administrativo (UA)
                        nuevo_id_admin = generar_identificador_siguiente(PerfilAdministrativo, "CUA", 'identificador_administrativo')
                        nuevo_id_ua = generar_identificador_siguiente(PerfilAdministrativo, "UA", 'identificador_coordinador_ua')

                        PerfilAdministrativo.objects.create(
                            usuario_de_sistema=usuario,
                            universidad=universidad_usuario,
                            identificador_administrativo=nuevo_id_admin,
                            identificador_coordinador_ua=nuevo_id_ua,
                            unidad_academica=ua_str,
                            perfil_administrativo=rol_fijo
                        )

                        # 3. Crear Perfil Docente
                        nuevo_id_docente = generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional')
                        
                        PerfilDocente.objects.create(
                            usuario_de_sistema=usuario,
                            identificador_institucional=nuevo_id_docente,
                            tipo_de_vinculacion=vinc_exacta,
                            tiempo_de_dedicacion=dedic_exacta,
                            carga_horaria_maxima=carga_max_float,
                            estado_de_vinculacion=EstadoDeVinculacion.ACTIVO.value 
                        )
                        
                        registros_exitosos += 1
                        
                    except Exception as error_de_fila:
                        messages.warning(request, f"Fila {numero_fila} omitida ({str(error_de_fila)})")
                        continue
                    
                if registros_exitosos > 0:
                    messages.success(request, f"{registros_exitosos} coordinadores de unidad académica registrados correctamente")
                else:
                    messages.warning(request, "No se procesaron registros")
                    
                return redirect("listar_coordinadores_ua")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento")
                return redirect("registrar_coordinador_ua")

        # === 2. LÓGICA DE FORMULARIO MANUAL ===
        else:
            formulario_usuario = FormularioUsuarioDeSistema(request.POST)
            formulario_perfil = FormularioRegistrarCoordinadorUA(request.POST)
            formulario_docente = FormularioDatosDocenteUA(request.POST) 

            if formulario_usuario.is_valid() and formulario_perfil.is_valid() and formulario_docente.is_valid():
                usuario = formulario_usuario.save(commit=False)
                usuario.set_password(usuario.identificacion)
                usuario.save()
                
                perfil = formulario_perfil.save(commit=False)
                perfil.usuario_de_sistema = usuario
                perfil.universidad = universidad_usuario
                perfil.identificador_administrativo = generar_identificador_siguiente(PerfilAdministrativo, "CUA", 'identificador_administrativo')
                perfil.identificador_coordinador_ua = generar_identificador_siguiente(PerfilAdministrativo, "UA", 'identificador_coordinador_ua')
                perfil.save()

                docente = formulario_docente.save(commit=False)
                docente.usuario_de_sistema = usuario
                docente.identificador_institucional = generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional')
                docente.estado_de_vinculacion = EstadoDeVinculacion.ACTIVO.value
                docente.save()

                messages.success(request, "El coordinador de unidad académica ha sido registrado correctamente")
                return redirect("listar_coordinadores_ua")
                
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        formulario_perfil = FormularioRegistrarCoordinadorUA()
        formulario_docente = FormularioDatosDocenteUA()

    return render(request, "usuarios/formulario_administrativo.html", {
        "form_usuario": formulario_usuario,
        "form_perfil": formulario_perfil,
        "form_docente": formulario_docente, 
        "titulo": "Registrar coordinador de unidad académica",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_coordinadores_ua",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_coordinador_ua"
    })


@login_required
def listar_coordinadores_ua(request):
    universidad_usuario = request.user.perfil_administrativo.universidad
    if not universidad_usuario:
        messages.warning(request, "La universidad no ha sido registrada actualmente.")
        return redirect("panel_principal")

    # Optimizamos la consulta trayendo select_related del usuario y prefetch_related de su perfil docente
    coordinadores = PerfilAdministrativo.objects.filter(
        universidad=universidad_usuario,
        perfil_administrativo=EnumPerfilAdministrativo.COORDINADOR_UA.value
    ).select_related("usuario_de_sistema").prefetch_related("usuario_de_sistema__perfil_docente")
    
    return render(request, "usuarios/listar_coordinadores_ua.html", {
        "coordinadores": coordinadores,
        "titulo_pagina": "Coordinador de unidad académica - NIVEC",
        "titulo": "Coordinadores de unidades académicas",
        "url_registrar": "registrar_coordinador_ua",
        "texto_registrar": "Registrar",
        "url_volver": "panel_director_dan"
    })

@login_required
def descargar_plantilla_docente(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"

    cabeceras = [
        "Tipo de identificación (Cédula, Pasaporte, Cédula extranjera)", 
        "Número de identificación", 
        "Nombres",
        "Apellidos", 
        "Correo institucional",
        "Tipo de vinculación (Nombramiento, Contrato, Ocasional, Honorario)",
        "Tiempo de dedicación (Tiempo completo, Medio tiempo, Tiempo parcial)",
        "Carga horaria máxima (número decimal)",
        "Especialidades (información separada por comas)"
    ]
    ws.append(cabeceras)

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 32
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['I'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_docentes_nivec.xlsx"'
    wb.save(response)
    return response

@login_required
def listar_docentes(request):
    docentes = PerfilDocente.objects.all().select_related("usuario_de_sistema")
    
    return render(request, "usuarios/listar_docentes.html", {
        "docentes": docentes,
        "titulo_pagina": "Docente - NIVEC",
        "titulo": "Docentes",
        "url_registrar": "registrar_docente",
        "texto_registrar": "Registrar",
        "url_volver": "panel_principal" # Ajusta esto a tu panel correcto
    })

@login_required
def registrar_docente(request):
    if request.method == "POST":
        
        # === 1. CARGA MASIVA ===
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Documento con formato no válido")
                return redirect("registrar_docente")
                
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                registros_exitosos = 0
                
                vinc_validas = {str(e.value).strip().lower(): e.value for e in TipoDeVinculacion}
                dedic_validas = {str(e.value).strip().lower(): e.value for e in TiempoDeDedicacion}

                for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max, especialidades_str = fila[:9]
                        
                        if not any([tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max]):
                            continue
                            
                        if not all([tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max]):
                            messages.warning(request, f"Fila {numero_fila} omitida por falta de información")
                            continue
                            
                        tipo_id_str = str(tipo_id).strip().capitalize()
                        identificacion_str = str(identificacion).strip()
                        tipo_vinc_limpio = str(tipo_vinc).strip().lower()
                        tiempo_dedic_limpio = str(tiempo_dedic).strip().lower()

                        if tipo_vinc_limpio not in vinc_validas:
                            messages.warning(request, f"Fila {numero_fila} omitida (tipo de vinculación no válido)")
                            continue
                        if tiempo_dedic_limpio not in dedic_validas:
                            messages.warning(request, f"Fila {numero_fila} omitida (tiempo de dedicación no válido)")
                            continue
                            
                        vinc_exacta = vinc_validas[tipo_vinc_limpio]
                        dedic_exacta = dedic_validas[tiempo_dedic_limpio]

                        try:
                            carga_max_float = float(carga_max)
                        except ValueError:
                            messages.warning(request, f"Fila {numero_fila} omitida (la carga horaria debe ser un número decimal)")
                            continue

                        # Procesar lista de especialidades
                        lista_especialidades = []
                        if especialidades_str:
                            lista_especialidades = [esp.strip() for esp in str(especialidades_str).split(',') if esp.strip()]

                        try:
                            UsuarioAdministrativoBase.validar_creacion_de_usuario_administrativo(
                                identificacion=identificacion_str, 
                                contrasena=identificacion_str
                            )
                        except ValueError as e:
                            messages.warning(request, f"Fila {numero_fila} omitida: {str(e)}")
                            continue

                        if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                            messages.warning(request, f"Fila {numero_fila} omitida (el usuario ya se encuentra registrado)")
                            continue

                        # 1. Crear Usuario
                        usuario = UsuarioDeSistema.objects.create(
                            tipo_de_identificacion=tipo_id_str,
                            identificacion=identificacion_str,
                            nombres=str(nombres).strip(),
                            apellidos=str(apellidos).strip(),
                            correo_institucional=str(correo).strip(),
                            estado_de_usuario=EstadoDeUsuario.ACTIVO.value
                        )
                        usuario.set_password(identificacion_str)
                        usuario.save()

                        # 2. Crear Perfil Docente
                        nuevo_id_docente = generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional')
                        PerfilDocente.objects.create(
                            usuario_de_sistema=usuario,
                            identificador_institucional=nuevo_id_docente,
                            tipo_de_vinculacion=vinc_exacta,
                            tiempo_de_dedicacion=dedic_exacta,
                            carga_horaria_maxima=carga_max_float,
                            estado_de_vinculacion=EstadoDeVinculacion.ACTIVO.value,
                            especialidades=lista_especialidades
                        )
                        
                        registros_exitosos += 1
                        
                    except Exception as error_de_fila:
                        messages.warning(request, f"Fila {numero_fila} omitida ({str(error_de_fila)})")
                        continue
                    
                if registros_exitosos > 0:
                    messages.success(request, f"{registros_exitosos} docentes registrados correctamente")
                else:
                    messages.warning(request, "No se procesaron registros")
                    
                return redirect("listar_docentes")
                
            except Exception as e:
                messages.error(request, f"Ha ocurrido un error al procesar el documento")
                return redirect("registrar_docente")

        # === 2. REGISTRO MANUAL ===
        else:
            formulario_usuario = FormularioUsuarioDeSistema(request.POST)
            formulario_docente = FormularioRegistrarDocente(request.POST) 

            if formulario_usuario.is_valid() and formulario_docente.is_valid():
                usuario = formulario_usuario.save(commit=False)
                usuario.set_password(usuario.identificacion)
                usuario.save()
                
                docente = formulario_docente.save(commit=False)
                docente.usuario_de_sistema = usuario
                docente.identificador_institucional = generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional')
                docente.estado_de_vinculacion = EstadoDeVinculacion.ACTIVO.value
                docente.especialidades = formulario_docente.cleaned_data.get('especialidades', [])
                docente.save()

                messages.success(request, "El docente ha sido registrado correctamente")
                return redirect("listar_docentes")
                
    else:
        formulario_usuario = FormularioUsuarioDeSistema()
        formulario_docente = FormularioRegistrarDocente()

    return render(request, "usuarios/formulario_docente.html", {
        "form_usuario": formulario_usuario,
        "form_docente": formulario_docente, 
        "titulo": "Registrar docente",
        "boton_texto": "Registrar",
        "url_cancelar": "listar_docentes",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_docente"
    })

@login_required
def modificar_docente(request, docente_id):
    # Obtenemos el perfil docente y su usuario asociado
    docente = get_object_or_404(PerfilDocente, id=docente_id)
    usuario = docente.usuario_de_sistema

    if request.method == "POST":
        # Usamos los mismos formularios pero pasándole la instancia existente (para que se pre-llenen y actualicen)
        formulario_usuario = FormularioUsuarioDeSistema(request.POST, instance=usuario)
        formulario_docente = FormularioRegistrarDocente(request.POST, instance=docente)

        # Como es modificación, quitamos la obligatoriedad de contraseñas si tu formulario base lo pide
        if 'contrasena' in formulario_usuario.fields:
            formulario_usuario.fields['contrasena'].required = False

        if formulario_usuario.is_valid() and formulario_docente.is_valid():
            usuario_guardado = formulario_usuario.save(commit=False)
            # Solo actualizamos la contraseña si el usuario escribió una nueva
            nueva_contrasena = formulario_usuario.cleaned_data.get('contrasena')
            if nueva_contrasena:
                usuario_guardado.set_password(nueva_contrasena)
            usuario_guardado.save()
            
            docente_guardado = formulario_docente.save(commit=False)
            # Procesamos nuevamente la lista de especialidades
            docente_guardado.especialidades = formulario_docente.cleaned_data.get('especialidades', [])
            docente_guardado.save()

            messages.success(request, "El docente ha sido modificado correctamente")
            return redirect("listar_docentes")
            
    else:
        formulario_usuario = FormularioUsuarioDeSistema(instance=usuario)
        formulario_docente = FormularioRegistrarDocente(instance=docente)

    return render(request, "usuarios/formulario_docente.html", {
        "form_usuario": formulario_usuario,
        "form_docente": formulario_docente, 
        "titulo": "Modificar docente",
        "boton_texto": "Modificar",
        "url_cancelar": "listar_docentes",
        "mostrar_carga_masiva": False # Ocultamos el cuadro de Excel al modificar
    })

@login_required
def eliminar_docente(request, docente_id):
    docente = get_object_or_404(PerfilDocente, id=docente_id)
    usuario = docente.usuario_de_sistema
    
    try:
        # Al eliminar el UsuarioDeSistema, la eliminación en cascada (CASCADE) 
        # borrará automáticamente el PerfilDocente asociado en la base de datos.
        usuario.delete()
        messages.success(request, "El docente ha sido eliminado correctamente")
    except Exception as e:
        messages.error(request, f"Error ({str(e)})")
        
    return redirect("listar_docentes")


@login_required
def inhabilitar_docente(request, docente_id):
    docente = get_object_or_404(PerfilDocente, id=docente_id)
    usuario = docente.usuario_de_sistema
    
    # Aplicamos tu lógica de POO: cambiar el estado a inactivo
    docente.estado_de_vinculacion = EstadoDeVinculacion.INACTIVO.value
    docente.save()
    
    # Desactivamos su acceso al sistema
    usuario.estado_de_usuario = EstadoDeUsuario.INACTIVO.value
    usuario.is_active = False 
    usuario.save()
    
    messages.success(request, f"El docente {usuario.nombres} {usuario.apellidos} ha sido inhabilitado correctamente")
    return redirect("listar_docentes")


@login_required
def habilitar_docente(request, docente_id):
    docente = get_object_or_404(PerfilDocente, id=docente_id)
    usuario = docente.usuario_de_sistema
    
    # Revertimos los estados al valor ACTIVO de tus Enums
    docente.estado_de_vinculacion = EstadoDeVinculacion.ACTIVO.value
    docente.save()
    
    usuario.estado_de_usuario = EstadoDeUsuario.ACTIVO.value
    usuario.is_active = True  # Le devolvemos el acceso al sistema
    usuario.save()
    
    messages.success(request, f"El docente {usuario.nombres} {usuario.apellidos} ha sido habilitado correctamente")
    return redirect("listar_docentes")

@login_required
def descargar_plantilla_estudiante(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    cabeceras = [
        "Tipo de identificación Tipo de identificación (Cédula, Pasaporte, Cédula extranjera)", "Número de identificación", "Nombres", 
        "Apellidos", "Correo institucional", "Jornada registrada (Matutina, Vespertina, Nocturna)", 
        "Registro de cupo (Registro regular, Segunda matrícula, Proceso de exoneración)", "Carrera registrada", "Campus registrado"
    ]
    ws.append(cabeceras)
    # Formato de celdas
    for col in range(1, 10): 
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formato_estudiantes_nivec.xlsx"'
    wb.save(response)
    return response

@login_required
def listar_estudiantes(request):
    estudiantes = PerfilEstudiante.objects.all().select_related("usuario_de_sistema", "carrera_registrada", "campus_registrado")
    return render(request, "usuarios/listar_estudiantes.html", {
        "estudiantes": estudiantes,
        "titulo": "Estudiantes",
        "url_registrar": "registrar_estudiante",
        "url_volver": "panel_principal"
    })

@login_required
def registrar_estudiante(request):
    if request.method == "POST":
        # === 1. CARGA MASIVA ===
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            registros_exitosos = 0
            
            # Mapas para validación de Enums
            jornadas_map = {str(j.value).strip().lower(): j.value for j in Jornada}
            cupos_map = {str(c.value).strip().lower(): c.value for c in RegistroDeCupo}

            for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    tipo_id, ident, nom, ape, corr, jor, cupo, carr, camp = fila
                    
                    if not all([tipo_id, ident, nom, ape, corr]): continue
                    
                    # Validar datos foráneos y enums
                    carrera = Carrera.objects.filter(nombre__iexact=str(carr).strip()).first()
                    campus = Campus.objects.filter(nombre__iexact=str(camp).strip()).first()
                    jor_val = jornadas_map.get(str(jor).strip().lower())
                    cupo_val = cupos_map.get(str(cupo).strip().lower())

                    if not carrera or not campus or not jor_val or not cupo_val:
                        messages.warning(request, f"Fila {numero_fila}: Datos de carrera, campus, jornada o cupo inválidos.")
                        continue

                    # Crear Usuario
                    usuario = UsuarioDeSistema.objects.create(
                        tipo_de_identificacion=str(tipo_id).strip().capitalize(),
                        identificacion=str(ident).strip(),
                        nombres=str(nom).strip(),
                        apellidos=str(ape).strip(),
                        correo_institucional=str(corr).strip(),
                        estado_de_usuario=EstadoDeUsuario.ACTIVO.value
                    )
                    usuario.set_password(str(ident).strip())
                    usuario.save()

                    # Crear Estudiante
                    PerfilEstudiante.objects.create(
                        usuario_de_sistema=usuario,
                        identificador_institucional=generar_identificador_siguiente(PerfilEstudiante, "ES", 'identificador_institucional'),
                        numero_de_matricula=generar_identificador_siguiente(PerfilEstudiante, "MAT", 'numero_de_matricula'),
                        jornada=jor_val,
                        registro_de_cupo=cupo_val,
                        carrera_registrada=carrera,
                        campus_registrado=campus,
                        estado_de_matricula=EstadoDeMatricula.MATRICULADO.value
                    )
                    registros_exitosos += 1
                except Exception as e:
                    messages.warning(request, f"Fila {numero_fila} omitida: {e}")
            
            messages.success(request, f"{registros_exitosos} estudiantes registrados.")
            return redirect("listar_estudiantes")

        # === 2. REGISTRO MANUAL ===
        else:
            form_u = FormularioUsuarioDeSistema(request.POST)
            form_e = FormularioPerfilEstudiante(request.POST)
            
            if form_u.is_valid() and form_e.is_valid():
                usuario = form_u.save(commit=False)
                usuario.set_password(usuario.identificacion)
                usuario.save()
                
                estudiante = form_e.save(commit=False)
                estudiante.usuario_de_sistema = usuario
                estudiante.identificador_institucional = generar_identificador_siguiente(PerfilEstudiante, "ES", 'identificador_institucional')
                estudiante.numero_de_matricula = generar_identificador_siguiente(PerfilEstudiante, "MAT", 'numero_de_matricula')
                estudiante.estado_de_matricula = EstadoDeMatricula.MATRICULADO.value
                estudiante.save()
                
                messages.success(request, "Estudiante registrado exitosamente.")
                return redirect("listar_estudiantes")
    
    else:
        form_u = FormularioUsuarioDeSistema()
        form_e = FormularioPerfilEstudiante()

    return render(request, "usuarios/formulario_estudiante.html", {
        "form_usuario": form_u,
        "form_estudiante": form_e,
        "titulo": "Registrar estudiante",
        "boton_texto": "Registrar",  # <--- CORREGIDO: Define el texto del botón
        "url_cancelar": "listar_estudiantes",
        "mostrar_carga_masiva": True,
        "url_plantilla": "descargar_plantilla_estudiante"
    })

@login_required
def modificar_estudiante(request, estudiante_id):
    est = get_object_or_404(PerfilEstudiante, id=estudiante_id)
    usuario = est.usuario_de_sistema
    
    if request.method == "POST":
        form_u = FormularioUsuarioDeSistema(request.POST, instance=usuario)
        form_e = FormularioPerfilEstudiante(request.POST, instance=est)
        
        # Quitamos obligatoriedad de contraseña en modificación
        form_u.fields['contrasena'].required = False 

        if form_u.is_valid() and form_e.is_valid():
            user_saved = form_u.save(commit=False)
            if form_u.cleaned_data.get('contrasena'):
                user_saved.set_password(form_u.cleaned_data['contrasena'])
            user_saved.save()
            form_e.save()
            
            messages.success(request, "Estudiante modificado correctamente.")
            return redirect("listar_estudiantes")
    else:
        form_u = FormularioUsuarioDeSistema(instance=usuario)
        form_e = FormularioPerfilEstudiante(instance=est)
        
    return render(request, "usuarios/formulario_estudiante.html", {
        "form_usuario": form_u, 
        "form_estudiante": form_e, 
        "titulo": "Modificar estudiante",
        "boton_texto": "Modificar", # <--- CORREGIDO: Define el texto del botón
        "url_cancelar": "listar_estudiantes"
    })

@login_required
def eliminar_estudiante(request, estudiante_id):
    get_object_or_404(PerfilEstudiante, id=estudiante_id).usuario_de_sistema.delete()
    messages.success(request, "Estudiante eliminado.")
    return redirect("listar_estudiantes")

# === TRANSICIONES ===
@login_required
def formalizar_matricula(request, estudiante_id):
    est = get_object_or_404(PerfilEstudiante, id=estudiante_id)
    est.estado_de_matricula = EstadoDeMatricula.MATRICULADO.value
    est.save()
    messages.success(request, "Matrícula formalizada.")
    return redirect("listar_estudiantes")

@login_required
def anular_matricula(request, estudiante_id):
    est = get_object_or_404(PerfilEstudiante, id=estudiante_id)
    est.estado_de_matricula = EstadoDeMatricula.ANULADO.value
    est.save()
    return redirect("listar_estudiantes")

@login_required
def solicitar_retiro(request, estudiante_id):
    est = get_object_or_404(PerfilEstudiante, id=estudiante_id)
    if est.estado_de_matricula == EstadoDeMatricula.MATRICULADO.value:
        est.estado_de_matricula = EstadoDeMatricula.RETIRADO.value
        est.save()
    return redirect("listar_estudiantes")