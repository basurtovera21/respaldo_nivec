import openpyxl
import unicodedata
from django.db import transaction
from django.contrib.auth import authenticate, login, logout

from usuarios.models import UsuarioDeSistema, PerfilDocente, PerfilAdministrativo, PerfilEstudiante
from academico.models import Carrera, Campus

from usuarios.utils import generar_identificador_siguiente

from poo.clases.usuarios.docente import Docente
from poo.clases.usuarios.estudiante import Estudiante
from poo.clases.usuarios.usuario_de_sistema import UsuarioDeSistema as UsuarioDeSistemaBase
from poo.clases.usuarios.usuario_administrativo import UsuarioAdministrativo as UsuarioAdministrativoBase

from poo.clases.enums.perfil_administrativo import PerfilAdministrativo as EnumPerfilAdministrativo
from poo.clases.enums.tipo_de_identificacion import TipoDeIdentificacion
from poo.clases.enums.estado_de_usuario import EstadoDeUsuario as EnumEstadoDeUsuario
from poo.clases.enums.estado_de_vinculacion import EstadoDeVinculacion as EnumEstadoDeVinculacion
from poo.clases.enums.tipo_de_vinculacion import TipoDeVinculacion as EnumTipoDeVinculacion
from poo.clases.enums.tiempo_de_dedicacion import TiempoDeDedicacion as EnumTiempoDeDedicacion
from poo.clases.enums.jornada import Jornada as EnumJornada
from poo.clases.enums.registro_de_cupo import RegistroDeCupo as EnumRegistroDeCupo
from poo.clases.enums.estado_de_matricula import EstadoDeMatricula as EnumEstadoDeMatricula


def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto

def obtener_enum_flexible(enum_class, valor_sucio):
    if not valor_sucio:
        return None
    valor_normalizado = normalizar_texto(valor_sucio)
    for opcion in enum_class:
        if normalizar_texto(opcion.value) == valor_normalizado:
            return opcion
    raise ValueError(f"'{valor_sucio}' registro no válido ({enum_class.__name__})")

def servicio_iniciar_sesion(request, correo_institucional, contrasena):
    usuario_de_sistema_django = authenticate(request, username=correo_institucional, password=contrasena)
    if usuario_de_sistema_django is None:
        return {"exito": False, "mensaje": "Las credenciales registradas no son válidas"}

    estado_valido = UsuarioDeSistemaBase.validar_estado_de_usuario(usuario_de_sistema_django.estado_de_usuario)
    if not estado_valido:
        mensaje = UsuarioDeSistemaBase.obtener_mensaje_estado_no_valido(usuario_de_sistema_django.estado_de_usuario)
        return {"exito": False, "mensaje": mensaje}
    
    login(request, usuario_de_sistema_django)
    return {"exito": True, "mensaje": ""}

def servicio_cerrar_sesion(request):
    logout(request)
    
def servicio_administrativo_registrar_masivo_desde_excel(archivo, universidad_usuario):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo); ws = wb.active
        perfiles_permitidos = [
            EnumPerfilAdministrativo.RECTOR,
            EnumPerfilAdministrativo.VICERRECTOR_ACADEMICO,
        ]
        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                tipo_id, identificacion, nombres, apellidos, correo, perfil_str = fila[:6]
                if not any([tipo_id, identificacion, nombres, apellidos, correo, perfil_str]): continue
                if not all([tipo_id, identificacion, nombres, apellidos, correo, perfil_str]):
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información"); continue
                tipo_id_str, identificacion_str = str(tipo_id).strip().capitalize(), str(identificacion).strip()
                correo_str = str(correo).strip()

                # Validar perfil via POO
                perfil_exacto = UsuarioAdministrativoBase.obtener_perfil_exacto(perfil_str, perfiles_permitidos)
                if not perfil_exacto:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (perfil no válido)"); continue

                # Validar identificación via POO
                try: UsuarioDeSistemaBase.validar_contrasena(identificacion_str)
                except ValueError as e: resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})"); continue

                # Validar correo via POO
                if not UsuarioDeSistemaBase.validar_correo_institucional(correo_str):
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (correo institucional no válido)"); continue

                # Verificar duplicados
                if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (el Administrativo ya ha sido registrado)"); continue
                if UsuarioDeSistema.objects.filter(correo_institucional__iexact=correo_str).exists():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (el correo institucional ya ha sido registrado)"); continue

                with transaction.atomic():
                    usuario = UsuarioDeSistema.objects.create(tipo_de_identificacion=tipo_id_str, identificacion=identificacion_str, nombres=str(nombres).strip(), apellidos=str(apellidos).strip(), correo_institucional=correo_str, estado_de_usuario=EnumEstadoDeUsuario.ACTIVO.value)
                    usuario.set_password(identificacion_str); usuario.save()
                    enum_perfil = EnumPerfilAdministrativo(perfil_exacto)
                    prefijo = UsuarioAdministrativoBase.definir_prefijo_identificador(enum_perfil)
                    nuevo_identificador = generar_identificador_siguiente(PerfilAdministrativo, prefijo, 'identificador_administrativo')
                    PerfilAdministrativo.objects.create(usuario_de_sistema=usuario, universidad=universidad_usuario, identificador_administrativo=nuevo_identificador, perfil_administrativo=perfil_exacto)
                    resultado["exitosos"] += 1
            except Exception as error_de_fila:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(error_de_fila)})")
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
    return resultado


def servicio_coordinador_dan_registrar_masivo_desde_excel(archivo, universidad_usuario):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo); ws = wb.active
        rol_fijo = EnumPerfilAdministrativo.COORDINADOR_DAN.value
        prefijo = UsuarioAdministrativoBase.definir_prefijo_identificador(EnumPerfilAdministrativo.COORDINADOR_DAN)
        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                tipo_id, identificacion, nombres, apellidos, correo = fila[:5]
                if not any([tipo_id, identificacion, nombres, apellidos, correo]): continue
                if not all([tipo_id, identificacion, nombres, apellidos, correo]):
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información"); continue
                tipo_id_str, identificacion_str = str(tipo_id).strip().capitalize(), str(identificacion).strip()
                try: UsuarioDeSistemaBase.validar_contrasena(identificacion_str)
                except ValueError as e: resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})"); continue
                if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (el Coordinador de dirección de admisión y nivelación ya ha sido registrado)"); continue
                with transaction.atomic():
                    usuario = UsuarioDeSistema.objects.create(tipo_de_identificacion=tipo_id_str, identificacion=identificacion_str, nombres=str(nombres).strip(), apellidos=str(apellidos).strip(), correo_institucional=str(correo).strip(), estado_de_usuario=EnumEstadoDeUsuario.ACTIVO.value)
                    usuario.set_password(identificacion_str); usuario.save()
                    PerfilAdministrativo.objects.create(usuario_de_sistema=usuario, universidad=universidad_usuario, identificador_administrativo=generar_identificador_siguiente(PerfilAdministrativo, prefijo, 'identificador_administrativo'), identificador_coordinador_dan=generar_identificador_siguiente(PerfilAdministrativo, prefijo, 'identificador_coordinador_dan'), perfil_administrativo=rol_fijo)
                    resultado["exitosos"] += 1
            except Exception as error_de_fila:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(error_de_fila)})")
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
    return resultado

def servicio_coordinador_ua_registrar_masivo_desde_excel(archivo, universidad_usuario):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo); ws = wb.active
        rol_fijo = EnumPerfilAdministrativo.COORDINADOR_UA.value
        vinc_validas = {str(e.value).strip().lower(): e.value for e in EnumTipoDeVinculacion}
        dedic_validas = {str(e.value).strip().lower(): e.value for e in EnumTiempoDeDedicacion}
        prefijo_ua = UsuarioAdministrativoBase.definir_prefijo_identificador(EnumPerfilAdministrativo.COORDINADOR_UA)
        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                tipo_id, identificacion, nombres, apellidos, correo, codigo_carrera, tipo_vinc, tiempo_dedic, carga_max, especialidades_str, jornadas_str = fila[:11]
                if not any([tipo_id, identificacion, nombres, apellidos, correo, codigo_carrera, tipo_vinc, tiempo_dedic, carga_max]): continue
                if not all([tipo_id, identificacion, nombres, apellidos, correo, codigo_carrera, tipo_vinc, tiempo_dedic, carga_max]):
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información"); continue
                tipo_id_str, identificacion_str, codigo_carrera_str = str(tipo_id).strip().capitalize(), str(identificacion).strip(), str(codigo_carrera).strip()
                tipo_vinc_limpio, tiempo_dedic_limpio = str(tipo_vinc).strip().lower(), str(tiempo_dedic).strip().lower()
                if tipo_vinc_limpio not in vinc_validas or tiempo_dedic_limpio not in dedic_validas:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (información no válida)"); continue
                vinc_exacta, dedic_exacta = vinc_validas[tipo_vinc_limpio], dedic_validas[tiempo_dedic_limpio]
                try: carga_max_float = float(carga_max)
                except ValueError: resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Carga horaria no válida)"); continue
                from usuarios.forms import validar_jornadas_continuas
                lista_especialidades = [e.strip() for e in str(especialidades_str).split(',') if e.strip()] if especialidades_str else []
                lista_jornadas = [j.strip().capitalize() for j in str(jornadas_str).split(',') if j.strip()] if jornadas_str else []
                error_jornadas = validar_jornadas_continuas(lista_jornadas)
                if error_jornadas:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (jornadas no válidas: {error_jornadas})"); continue
                try: UsuarioDeSistemaBase.validar_contrasena(identificacion_str)
                except ValueError as e: resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})"); continue
                if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (el Coordinador de unidad académica ya ha sido registrado)"); continue
                from academico.models import Carrera
                carrera_obj = Carrera.objects.filter(codigo_de_carrera=codigo_carrera_str, campus__universidad=universidad_usuario).first()
                if not carrera_obj:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (código de Carrera no válido)"); continue
                with transaction.atomic():
                    usuario = UsuarioDeSistema.objects.create(tipo_de_identificacion=tipo_id_str, identificacion=identificacion_str, nombres=str(nombres).strip(), apellidos=str(apellidos).strip(), correo_institucional=str(correo).strip(), estado_de_usuario=EnumEstadoDeUsuario.ACTIVO.value)
                    usuario.set_password(identificacion_str); usuario.save()
                    PerfilAdministrativo.objects.create(usuario_de_sistema=usuario, universidad=universidad_usuario, identificador_administrativo=generar_identificador_siguiente(PerfilAdministrativo, prefijo_ua, 'identificador_administrativo'), identificador_coordinador_ua=generar_identificador_siguiente(PerfilAdministrativo, prefijo_ua, 'identificador_coordinador_ua'), carrera_asignada=carrera_obj, perfil_administrativo=rol_fijo)
                    PerfilDocente.objects.create(usuario_de_sistema=usuario, universidad=universidad_usuario, identificador_institucional=generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional'), tipo_de_vinculacion=vinc_exacta, tiempo_de_dedicacion=dedic_exacta, carga_horaria_maxima=carga_max_float, estado_de_vinculacion=EnumEstadoDeVinculacion.ACTIVO.value, especialidades=lista_especialidades, jornadas=lista_jornadas)
                    resultado["exitosos"] += 1
            except Exception as e: resultado["advertencias"].append(f"Fila {numero_fila} omitida ({str(e)})")
    except Exception: resultado["error"] = "Ha ocurrido un error al procesar el documento"
    return resultado


def servicio_docente_registrar_masivo_desde_excel(archivo, universidad):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        vinc_validas = {str(e.value).strip().lower(): e.value for e in EnumTipoDeVinculacion}
        dedic_validas = {str(e.value).strip().lower(): e.value for e in EnumTiempoDeDedicacion}

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max, especialidades_str, jornadas_str = fila[:10]
                
                if not any([tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max]):
                    continue
                
                if not all([tipo_id, identificacion, nombres, apellidos, correo, tipo_vinc, tiempo_dedic, carga_max]):
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información")
                    continue
                    
                tipo_id_str = str(tipo_id).strip().capitalize()
                identificacion_str = str(identificacion).strip()
                tipo_vinc_limpio = str(tipo_vinc).strip().lower()
                tiempo_dedic_limpio = str(tiempo_dedic).strip().lower()

                if tipo_vinc_limpio not in vinc_validas:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Tipo de vinculación no válido)")
                    continue
                if tiempo_dedic_limpio not in dedic_validas:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Tiempo de dedicación no válido)")
                    continue
                    
                try:
                    carga_max_float = float(carga_max)
                except ValueError:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Carga horaria no válida)")
                    continue

                try: 
                    UsuarioDeSistemaBase.validar_contrasena(identificacion_str)
                except ValueError as e: 
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})")
                    continue

                if UsuarioDeSistema.objects.filter(identificacion=identificacion_str).exists():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (el Docente ya se encuentra registrado)")
                    continue
                
                lista_especialidades = [esp.strip() for esp in str(especialidades_str).split(',') if esp.strip()] if especialidades_str else []

                from usuarios.forms import validar_jornadas_continuas
                lista_jornadas = [j.strip().capitalize() for j in str(jornadas_str).split(',') if j.strip()] if jornadas_str else []
                error_jornadas = validar_jornadas_continuas(lista_jornadas)
                if error_jornadas:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Jornadas no válidas: {error_jornadas})")
                    continue

                with transaction.atomic():
                    usuario = UsuarioDeSistema.objects.create(
                        tipo_de_identificacion=tipo_id_str,
                        identificacion=identificacion_str,
                        nombres=str(nombres).strip(),
                        apellidos=str(apellidos).strip(),
                        correo_institucional=str(correo).strip(),
                        estado_de_usuario=EnumEstadoDeUsuario.ACTIVO.value
                    )
                    usuario.set_password(identificacion_str)
                    usuario.save()

                    PerfilDocente.objects.create(
                        usuario_de_sistema=usuario,
                        universidad=universidad,
                        identificador_institucional=generar_identificador_siguiente(PerfilDocente, "DC", 'identificador_institucional'),
                        tipo_de_vinculacion=vinc_validas[tipo_vinc_limpio],
                        tiempo_de_dedicacion=dedic_validas[tiempo_dedic_limpio],
                        carga_horaria_maxima=carga_max_float,
                        estado_de_vinculacion=EnumEstadoDeVinculacion.ACTIVO.value,
                        especialidades=lista_especialidades,
                        jornadas=lista_jornadas
                    )
                    resultado["exitosos"] += 1
            
            except Exception as e:
                resultado["advertencias"].append(f"Registro de la fila {numero_fila} omitido ({str(e)})")
                continue
                
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
        
    return resultado

def servicio_estudiante_registrar_masivo_desde_excel(archivo, universidad_usuario, periodo_de_nivelacion=None):
    from poo.clases.criterios_filtro.criterio_cedula_formato import CriterioCedulaFormato
    from poo.clases.servicios.depurador_de_sincronizacion import DepuradorDeSincronizacion

    resultado = {
        "exitosos": 0,
        "advertencias": [],
        "error": None,
        "observados": 0,
        "clasificacion": {"regular": 0, "segunda": 0, "exoneracion": 0},
        "identificaciones_validas": [],
    }
    try:
        wb = openpyxl.load_workbook(archivo); ws = wb.active
        jornadas_map = {str(j.value).strip().lower(): j.value for j in EnumJornada}
        cupos_map = {str(c.value).strip().lower(): c.value for c in EnumRegistroDeCupo}

        registros = []
        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tipo_id, ident, nom, ape, corr, jor, cupo, cod_carr = fila[:8]
            if not any([tipo_id, ident, nom, ape, corr, jor, cupo, cod_carr]):
                continue
            if not all([tipo_id, ident, nom, ape, corr, jor, cupo, cod_carr]):
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información")
                continue
            registros.append({
                "fila": numero_fila,
                "tipo_de_identificacion": str(tipo_id).strip().capitalize(),
                "cedula": str(ident).strip(),
                "nombres": str(nom).strip(),
                "apellidos": str(ape).strip(),
                "correo": str(corr).strip(),
                "jornada": str(jor).strip(),
                "cupo": str(cupo).strip(),
                "codigo_carrera": str(cod_carr).strip(),
            })

        depurador = DepuradorDeSincronizacion([CriterioCedulaFormato()])
        depurador.procesar_matriz_externa(registros)

        for registro in depurador.registros_con_observaciones:
            resultado["advertencias"].append(
                f"El registro de la fila {registro['fila']} fue omitido (Número de cédula no válido)"
            )

        for registro in depurador.registros_validos:
            try:
                ident_str = registro["cedula"]
                if UsuarioDeSistema.objects.filter(identificacion=ident_str).exists():
                    resultado["advertencias"].append(
                        f"El registro de la fila {registro['fila']} fue omitido (el Estudiante ya se encuentra registrado)"
                    )
                    continue


                carrera = Carrera.objects.filter(
                    codigo_de_carrera__iexact=registro["codigo_carrera"], campus__universidad=universidad_usuario
                ).select_related("campus").first()
                jor_val = jornadas_map.get(registro["jornada"].lower())
                cupo_val = cupos_map.get(registro["cupo"].lower())

                if not carrera or not jor_val or not cupo_val:
                    resultado["advertencias"].append(
                        f"El registro de la fila {registro['fila']} fue omitido (información no válida)"
                    )
                    continue

                with transaction.atomic():
                    usuario = UsuarioDeSistema.objects.create(
                        tipo_de_identificacion=registro["tipo_de_identificacion"],
                        identificacion=ident_str,
                        nombres=registro["nombres"],
                        apellidos=registro["apellidos"],
                        correo_institucional=registro["correo"],
                        estado_de_usuario=EnumEstadoDeUsuario.ACTIVO.value
                    )
                    usuario.set_password(ident_str); usuario.save()
                    PerfilEstudiante.objects.create(
                        usuario_de_sistema=usuario,
                        identificador_institucional=generar_identificador_siguiente(PerfilEstudiante, "ES", 'identificador_institucional'),
                        numero_de_matricula=generar_identificador_siguiente(PerfilEstudiante, "MAT", 'numero_de_matricula'),
                        jornada=jor_val,
                        registro_de_cupo=cupo_val,
                        carrera_registrada=carrera,
                        campus_registrado=carrera.campus,
                        estado_de_matricula=EnumEstadoDeMatricula.MATRICULADO.value,
                        periodo_de_nivelacion=periodo_de_nivelacion
                    )
                    resultado["exitosos"] += 1
                    resultado["identificaciones_validas"].append(ident_str)

                if cupo_val == EnumRegistroDeCupo.REGULAR.value:
                    resultado["clasificacion"]["regular"] += 1
                elif cupo_val == EnumRegistroDeCupo.SEGUNDA_MATRICULA.value:
                    resultado["clasificacion"]["segunda"] += 1
                elif cupo_val == EnumRegistroDeCupo.EXONERACION.value:
                    resultado["clasificacion"]["exoneracion"] += 1
            except Exception as e:
                resultado["advertencias"].append(f"Fila {registro['fila']} omitida ({str(e)})")
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"

    resultado["observados"] = len(resultado["advertencias"])
    return resultado


def _crear_usuario_administrativo(perfil_administrativo: PerfilAdministrativo):
    usuario_de_sistema = perfil_administrativo.usuario_de_sistema
    return UsuarioAdministrativoBase(
        tipo_de_identificacion=obtener_enum_flexible(TipoDeIdentificacion, usuario_de_sistema.tipo_de_identificacion),
        identificacion=usuario_de_sistema.identificacion,
        nombres=usuario_de_sistema.nombres,
        apellidos=usuario_de_sistema.apellidos,
        correo_institucional=usuario_de_sistema.correo_institucional,
        contrasena="temporal",
        fecha_de_nacimiento=usuario_de_sistema.fecha_de_nacimiento,
        sexo=usuario_de_sistema.sexo,
        etnia=usuario_de_sistema.etnia,
        porcentaje_de_discapacidad=usuario_de_sistema.porcentaje_de_discapacidad,
        celular=usuario_de_sistema.celular,
        direccion=usuario_de_sistema.direccion,
        identificador_administrativo=perfil_administrativo.identificador_administrativo,
        perfil_administrativo=obtener_enum_flexible(EnumPerfilAdministrativo, perfil_administrativo.perfil_administrativo),
        universidad=perfil_administrativo.universidad
    )

def _crear_docente(perfil_docente: PerfilDocente):
    usuario_de_sistema = perfil_docente.usuario_de_sistema
    return Docente(
        tipo_de_identificacion=obtener_enum_flexible(TipoDeIdentificacion, usuario_de_sistema.tipo_de_identificacion),
        identificacion=usuario_de_sistema.identificacion,
        nombres=usuario_de_sistema.nombres,
        apellidos=usuario_de_sistema.apellidos,
        correo_institucional=usuario_de_sistema.correo_institucional,
        contrasena="temporal",
        fecha_de_nacimiento=None,
        sexo=usuario_de_sistema.sexo if hasattr(usuario_de_sistema, 'sexo') else "No especificado",
        etnia=usuario_de_sistema.etnia if hasattr(usuario_de_sistema, 'etnia') else "No especificado",
        porcentaje_de_discapacidad=0.0,
        celular=usuario_de_sistema.celular if hasattr(usuario_de_sistema, 'celular') else "",
        direccion=usuario_de_sistema.direccion if hasattr(usuario_de_sistema, 'direccion') else "",
        identificador_institucional=perfil_docente.identificador_institucional,
        universidad=perfil_docente.universidad,
        tipo_de_vinculacion=obtener_enum_flexible(EnumTipoDeVinculacion, perfil_docente.tipo_de_vinculacion),
        tiempo_de_dedicacion=obtener_enum_flexible(EnumTiempoDeDedicacion, perfil_docente.tiempo_de_dedicacion),
        carga_horaria_maxima=perfil_docente.carga_horaria_maxima
    )
    
def _crear_estudiante(perfil_est):
    usuario = perfil_est.usuario_de_sistema
    return Estudiante(
        tipo_de_identificacion=obtener_enum_flexible(TipoDeIdentificacion, usuario.tipo_de_identificacion),
        identificacion=usuario.identificacion,
        nombres=usuario.nombres,
        apellidos=usuario.apellidos,
        correo_institucional=usuario.correo_institucional,
        contrasena="temporal",
        fecha_de_nacimiento=None,
        sexo=usuario.sexo if hasattr(usuario, 'sexo') else "No especificado",
        etnia=usuario.etnia if hasattr(usuario, 'etnia') else "No especificado",
        porcentaje_de_discapacidad=0.0,
        celular=usuario.celular if hasattr(usuario, 'celular') else "",
        direccion=usuario.direccion if hasattr(usuario, 'direccion') else "",
        identificador_institucional=perfil_est.identificador_institucional,
        numero_de_matricula=perfil_est.numero_de_matricula,
        jornada=obtener_enum_flexible(EnumJornada, perfil_est.jornada),
        registro_de_cupo=obtener_enum_flexible(EnumRegistroDeCupo, perfil_est.registro_de_cupo),
        carrera_registrada=perfil_est.carrera_registrada,
        campus_registrado=perfil_est.campus_registrado,
        estado_de_matricula=obtener_enum_flexible(EnumEstadoDeMatricula, perfil_est.estado_de_matricula)
    )